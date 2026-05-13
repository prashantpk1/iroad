from django.utils.translation import gettext as _
from django.contrib import messages
from django.template import Context, Template
from django.http import HttpResponse, JsonResponse
from django.contrib.sessions.models import Session
from django.contrib.auth import login, logout
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth.mixins import LoginRequiredMixin
from django.conf import settings
from django.db import connection
from django.db import transaction as db_transaction
from django.db.models import (
    Sum, Count, F, Q, Window, Case, When, Value, IntegerField, DecimalField,
    CharField,
)
from django.db.models.functions import Coalesce
from django.db.models.expressions import ExpressionWrapper
from django.db.models.fields import DecimalField
from django.db.models.functions import Abs, RowNumber, ExtractHour, ExtractWeekDay
from superadmin.redis_helpers import (
    count_active_admin_sessions
)
from django.shortcuts import get_object_or_404, redirect, render
from django.core.paginator import Paginator
from django.core.files.storage import default_storage
from django.core import signing
from django.urls import reverse
from urllib.parse import urlencode
from django.utils import timezone
from django.utils.html import strip_tags
from django.template.loader import render_to_string
from django.utils.dateparse import parse_date
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views import View
from django.views.generic import TemplateView
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
import json
import logging
import os
import re
import secrets
import subprocess
import uuid
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - optional dependency safety
    Workbook = None
    load_workbook = None

logger = logging.getLogger(__name__)
TENANT_USER_PASSWORD_RESET_TOKEN_SALT = 'tenant-user-password-reset'
TENANT_USER_PASSWORD_RESET_TOKEN_MAX_AGE_SECONDS = 3600


def _rename_txn_attachment(uploaded_file, txn_id):
    ext = os.path.splitext((uploaded_file.name or '').strip())[1].lower() or '.bin'
    uploaded_file.name = f'txn_{txn_id}_{uuid.uuid4().hex[:10]}{ext}'
    return uploaded_file


def _write_txn_approve_debug(message):
    """Append transaction-approve debug output to a local txt file."""
    debug_file = os.path.join(settings.BASE_DIR, 'transaction_approve_debug.txt')
    stamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        with open(debug_file, 'a', encoding='utf-8') as fh:
            fh.write(f'[{stamp}] {message}\n')
    except Exception:
        logger.exception('Failed writing transaction approve debug log')


def _build_tenant_user_password_reset_token(*, tenant_id, email):
    payload = {
        'tenant_id': str(tenant_id),
        'email': (email or '').strip().lower(),
    }
    return signing.dumps(payload, salt=TENANT_USER_PASSWORD_RESET_TOKEN_SALT)


def _read_tenant_user_password_reset_token(raw_token):
    try:
        return signing.loads(
            raw_token,
            salt=TENANT_USER_PASSWORD_RESET_TOKEN_SALT,
            max_age=TENANT_USER_PASSWORD_RESET_TOKEN_MAX_AGE_SECONDS,
        )
    except signing.SignatureExpired:
        return None
    except signing.BadSignature:
        return None


from .billing_helpers import (
    calculate_addon_prorata,
    calculate_promo_discount,
    calculate_pro_rata_credit,
    complete_order_payment_as_system,
    fulfill_paid_order,
    get_fx_snapshot,
    get_plan_cycle_days,
    get_standard_billing_cycle_days,
    get_tax_code_for_tenant,
    refresh_order_projected_fields,
    resolve_upgrade_credit_basis_price,
    send_invoice_paid_notification,
    sync_or_create_order_payment_transaction,
    get_live_bill_to_snapshot,
    validate_downgrade_order,
)
from .auth_helpers import (
    check_brute_force,
    create_auth_token,
    get_security_settings,
    log_access,
    record_failed_attempt,
    reset_failed_attempts,
    send_auth_email,
)
from .audit_helpers import (
    create_session,
    close_session,
    get_client_ip,
    log_audit_action,
)
from .redis_helpers import create_admin_session
from .tenant_portal_auth import (
    get_tenant_portal_cookie_payload,
    set_tenant_portal_cookie,
)
from .communication_helpers import (
    _extract_sender_address,
    _normalize_from_email_header,
    dispatch_internal_alerts,
    dispatch_event_notification,
    ensure_default_notification_templates,
    send_named_notification_email,
    send_transactional_email,
)
from .communication_helpers import (
    dispatch_event_notification,
    ensure_default_notification_templates,
    send_named_notification_email,
    send_transactional_email,
)
from .forms import (
    AddOnsPricingPolicyForm,
    AdminSecuritySettingsForm,
    AdminUserForm,
    CountryForm,
    CurrencyForm,
    CommGatewayForm,
    BaseCurrencyForm,
    BankAccountForm,
    ExchangeRateForm,
    EventMappingForm,
    ForgotPasswordForm,
    LoginForm,
    MyAccountForm,
    OTPVerificationForm,
    GeneralTaxSettingsForm,
    GlobalSystemRulesForm,
    RoleForm,
    SetPasswordForm,
    LegalIdentityForm,
    NotificationTemplateForm,
    PlanPricingCycleForm,
    PaymentGatewayForm,
    PaymentMethodForm,
    PromoCodeForm,
    PushNotificationForm,
    SystemBannerForm,
    SubscriptionPlanForm,
    TaxCodeForm,
    InternalAlertRouteForm,
    TenantProfileCreateForm,
    TenantProfileUpdateForm,
    TenantSecuritySettingsForm,
    SupportCategoryForm,
    CannedResponseForm,
    SubscriptionFAQForm,
    SupportTicketForm,
    TicketAssignForm,
    TicketPriorityForm,
    AdminReplyForm,
)
from .models import (
    AccessLog,
    ActiveSession,
    AdminSecuritySettings,
    AddOnsPricingPolicy,
    AdminAuthToken,
    AdminUser,
    AuditLog,
    BaseCurrencyConfig,
    BankAccount,
    Country,
    CommGateway,
    CommLog,
    CRMNote,
    Currency,
    EventMapping,
    InternalAlertNotification,
    InternalAlertRoute,
    GeneralTaxSettings,
    GlobalSystemRules,
    LegalIdentity,
    NotificationTemplate,
    OrderAddonLine,
    OrderPlanLine,
    PaymentGateway,
    PaymentMethod,
    PlanPricingCycle,
    PromoCode,
    PushNotification,
    Role,
    StandardInvoice,
    SubscriptionOrder,
    SubscriptionPlan,
    SystemBanner,
    TaxCode,
    TenantProfile,
    TenantSecuritySettings,
    Transaction,
    ExchangeRate,
    FXRateChangeLog,
    SupportCategory,
    CannedResponse,
    SubscriptionFAQ,
    SupportTicket,
    TicketReply,
)
from iroad_tenants.models import TenantAuthToken, TenantRegistry


def _client_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def _resolve_tenant_by_email(raw_email):
    """Resolve tenant login target from email with duplicate-safety."""
    email = (raw_email or '').strip()
    if not email:
        return None, 'not_found'

    qs = TenantProfile.objects.filter(primary_email__iexact=email)
    if not qs.exists():
        return None, 'not_found'

    active_qs = qs.filter(account_status='Active')
    active_count = active_qs.count()
    if active_count == 1:
        return active_qs.first(), None
    if active_count > 1:
        return None, 'ambiguous_active'

    if qs.count() == 1:
        return qs.first(), None
    return None, 'ambiguous_inactive'


def _resolve_tenant_user_by_tid_and_email(raw_tid, raw_email):
    tenant_id = str(raw_tid or '').strip()
    email = (raw_email or '').strip().lower()
    if not email:
        return None, None, 'missing_inputs'

    from tenant_workspace.models import TenantRole, TenantUser

    if not tenant_id:
        registries = list(
            TenantRegistry.objects.select_related('tenant_profile')
            .filter(tenant_profile__account_status='Active')
            .order_by('tenant_profile__company_name')
        )
        matches = []
        inactive_matches = []
        role_inactive_matches = []
        connection.set_schema_to_public()
        try:
            for registry in registries:
                tenant = registry.tenant_profile
                connection.set_tenant(registry)
                tenant_user = (
                    TenantUser.objects.filter(email__iexact=email)
                    .order_by('-updated_at')
                    .first()
                )
                if not tenant_user:
                    continue
                if tenant_user.status != TenantUser.Status.ACTIVE:
                    inactive_matches.append((tenant, tenant_user))
                    continue
                role = (
                    TenantRole.objects.filter(
                        role_name_en__iexact=(tenant_user.role_name or '').strip()
                    ).first()
                )
                if role and role.status != TenantRole.Status.ACTIVE:
                    role_inactive_matches.append((tenant, tenant_user))
                    continue
                matches.append((tenant, tenant_user))
        finally:
            connection.set_schema_to_public()

        if len(matches) == 1:
            return matches[0][0], matches[0][1], None
        if len(matches) > 1:
            return None, None, 'tenant_user_ambiguous'
        if inactive_matches:
            return inactive_matches[0][0], None, 'tenant_user_inactive'
        if role_inactive_matches:
            return role_inactive_matches[0][0], None, 'tenant_user_role_inactive'
        return None, None, 'tenant_user_not_found'

    tenant = TenantProfile.objects.filter(pk=tenant_id).first()
    if not tenant:
        return None, None, 'tenant_not_found'
    if tenant.account_status != 'Active':
        return tenant, None, 'tenant_inactive'

    registry = (
        TenantRegistry.objects.select_related('tenant_profile')
        .filter(tenant_profile_id=tenant_id)
        .first()
    )
    if not registry:
        return tenant, None, 'tenant_registry_missing'

    connection.set_schema_to_public()
    try:
        connection.set_tenant(registry)
        tenant_user = (
            TenantUser.objects.filter(email__iexact=email)
            .order_by('-updated_at')
            .first()
        )
        if not tenant_user:
            return tenant, None, 'tenant_user_not_found'
        if tenant_user.status != TenantUser.Status.ACTIVE:
            return tenant, None, 'tenant_user_inactive'
        role = (
            TenantRole.objects.filter(role_name_en__iexact=(tenant_user.role_name or '').strip())
            .first()
        )
        if role and role.status != TenantRole.Status.ACTIVE:
            return tenant, None, 'tenant_user_role_inactive'
        return tenant, tenant_user, None
    finally:
        connection.set_schema_to_public()


ADMIN_LOGIN_OTP_SESSION_KEY = 'admin_login_otp'
ADMIN_LOGIN_OTP_TTL_SECONDS = 300
ADMIN_LOGIN_OTP_MAX_ATTEMPTS = 5
TENANT_LOGIN_OTP_SESSION_KEY = 'tenant_login_otp'
TENANT_LOGIN_OTP_TTL_SECONDS = 300
TENANT_LOGIN_OTP_MAX_ATTEMPTS = 5


def _issue_admin_login_otp(request, user):
    ensure_default_notification_templates(
        created_by=user if getattr(user, 'is_authenticated', False) else None
    )
    otp_code = f'{secrets.randbelow(1000000):06d}'
    print("otp_code", otp_code)
    
    # PCS FRM-CP-11-01 — fetch dynamic OTP TTL from settings row.
    from .auth_helpers import get_security_settings
    settings_obj = get_security_settings()
    ttl_seconds = getattr(settings_obj, 'otp_timeout_seconds', ADMIN_LOGIN_OTP_TTL_SECONDS)

    expires_at = timezone.now() + timezone.timedelta(seconds=ttl_seconds)
    request.session[ADMIN_LOGIN_OTP_SESSION_KEY] = {
        'admin_id': str(getattr(user, 'admin_id', user.id)),
        'email': (user.email or '').strip().lower(),
        'otp_code': otp_code,
        'expires_at': expires_at.isoformat(),
        'attempts': 0,
    }
    request.session.modified = True

    ctx = {
        'otp_code': otp_code,
        'otp': otp_code,
        'user_name': (getattr(user, 'full_name', '') or user.email or 'Admin User'),
    }
    sent = send_named_notification_email(
        'AUTH_LOGIN_OTP',
        recipient_email=user.email,
        context_dict=ctx,
        default_subject='Your iRoad Login Verification Code',
        trigger_source='TemplateName: AUTH_LOGIN_OTP',
        force_django_smtp=False,
    )
    if sent:
        return True

    sent = dispatch_event_notification(
        'OTP_Requested',
        recipient_email=user.email,
        context_dict=ctx,
        use_async_tasks=False,
    )
    if sent:
        return True

    # Fallback when EventMapping/template is not configured.
    send_transactional_email(
        user.email,
        'Your iRoad OTP verification code',
        f'Your verification code is {otp_code}. It expires in 5 minutes.',
        (
            f'<p>Your verification code is <strong>{otp_code}</strong>.</p>'
            '<p>This code expires in 5 minutes.</p>'
        ),
        trigger_source='Direct: Login OTP',
    )
    return True


def _issue_tenant_login_otp(request, tenant, *, recipient_email=None, recipient_name=None):
    ensure_default_notification_templates(created_by=None)
    otp_code = f'{secrets.randbelow(1000000):06d}'
    expires_at = timezone.now() + timezone.timedelta(seconds=TENANT_LOGIN_OTP_TTL_SECONDS)
    target_email = (recipient_email or tenant.primary_email or '').strip().lower()
    request.session[TENANT_LOGIN_OTP_SESSION_KEY] = {
        'tenant_id': str(tenant.tenant_id),
        'email': target_email,
        'otp_code': otp_code,
        'expires_at': expires_at.isoformat(),
        'attempts': 0,
    }
    request.session.modified = True

    tenant_name = recipient_name or (
        f"{(tenant.first_name or '').strip()} {(tenant.last_name or '').strip()}".strip()
        or (tenant.company_name or tenant.primary_email or 'Tenant User')
    )
    ctx = {
        'otp_code': otp_code,
        'otp': otp_code,
        'user_name': tenant_name,
        'company_name': (tenant.company_name or tenant_name),
    }
    sent = send_named_notification_email(
        'AUTH_LOGIN_OTP',
        recipient_email=target_email,
        context_dict=ctx,
        default_subject='Your iRoad Login Verification Code',
        trigger_source='TemplateName: AUTH_LOGIN_OTP',
        force_django_smtp=False,
    )
    if sent:
        return True

    sent = dispatch_event_notification(
        'OTP_Requested',
        recipient_email=target_email,
        context_dict=ctx,
        use_async_tasks=False,
    )
    if sent:
        return True

    send_transactional_email(
        target_email,
        'Your iRoad OTP verification code',
        f'Your verification code is {otp_code}. It expires in 5 minutes.',
        (
            f'<p>Your verification code is <strong>{otp_code}</strong>.</p>'
            '<p>This code expires in 5 minutes.</p>'
        ),
        trigger_source='Direct: Tenant Login OTP',
    )
    return True


@method_decorator(ensure_csrf_cookie, name='dispatch')
class LoginView(View):
    template_name = 'auth/login.html'

    def get(self, request):
        email = request.GET.get('email', '').strip().lower()
        settings_obj = get_security_settings()
        context = {
            'form': LoginForm(initial={'email': email}),
            'max_failed_logins': settings_obj.max_failed_logins,
            'attempts_remaining': settings_obj.max_failed_logins,
            'security_timeout': settings_obj.otp_timeout_seconds,
        }

        if email:
            brute = check_brute_force(email)
            if brute['is_locked']:
                context.update({
                    'is_locked': True,
                    'remaining_seconds': brute.get('remaining_seconds', 0),
                    'lockout_duration_total': settings_obj.lockout_duration_minutes * 60,
                    'attempts_remaining': 0,
                    'locked_email': email,
                    'error': 'Account locked due to too many failed attempts.',
                })
            else:
                context['attempts_remaining'] = max(0, settings_obj.max_failed_logins - brute['failed_count'])

        return render(request, self.template_name, context)

    def post(self, request):
        form = LoginForm(request.POST)
        email = request.POST.get('email', '').lower().strip()
        ip = _client_ip(request)

        # STEP 1: Check brute force FIRST
        brute = check_brute_force(email)
        settings_obj = get_security_settings()
        
        # Calculate initial remaining attempts based on pre-post state
        current_remaining = max(0, settings_obj.max_failed_logins - brute['failed_count'])

        base_context = {
            'max_failed_logins': settings_obj.max_failed_logins,
            'attempts_remaining': current_remaining,
            'security_timeout': settings_obj.otp_timeout_seconds,
            'is_locked': brute['is_locked'],
        }

        def render_login(extra_context=None):
            context = {'form': form, **base_context}
            if extra_context:
                context.update(extra_context)
            return render(request, self.template_name, context)

        if brute['is_locked']:
            log_access('Login', 'Blocked', email, ip)
            return render_login({
                'error': 'Account locked due to too many failed attempts.',
                'is_locked': True,
                'remaining_seconds': brute.get('remaining_seconds', 0),
                'lockout_duration_total': settings_obj.lockout_duration_minutes * 60,
                'attempts_remaining': 0,
                'locked_email': email,
            })

        if not form.is_valid():
            if form.errors.get('email'):
                error_msg = 'Please enter a valid email address.'
            elif form.errors.get('password'):
                error_msg = 'Please enter your password.'
            else:
                error_msg = 'Please enter valid email and password.'
            return render_login({'error': error_msg})

        # STEP 2: Check user exists
        try:
            user = AdminUser.objects.get(
                email=form.cleaned_data['email'].lower().strip()
            )
        except AdminUser.DoesNotExist:
            tid = (request.GET.get('tid') or request.POST.get('tid') or '').strip()
            if tid:
                tenant_by_tid, tenant_user_by_tid, tenant_user_resolution = _resolve_tenant_user_by_tid_and_email(
                    tid,
                    email,
                )
                if tenant_by_tid and tenant_user_by_tid:
                    if not check_password(
                        form.cleaned_data['password'],
                        tenant_user_by_tid.password_hash,
                    ):
                        record_failed_attempt(email)
                        log_access('Login', 'Failed', email, ip)
                        brute_after = check_brute_force(email)
                        if brute_after['is_locked']:
                            return render_login({
                                'error': 'Account locked due to too many failed attempts.',
                                'is_locked': True,
                                'remaining_seconds': brute_after.get('remaining_seconds', 0),
                                'lockout_duration_total': settings_obj.lockout_duration_minutes * 60,
                                'attempts_remaining': 0,
                                'locked_email': email,
                            })
                        remaining_attempts = max(0, settings_obj.max_failed_logins - brute_after['failed_count'])
                        return render_login({
                            'error': f"Invalid email or password. {remaining_attempts} attempts remaining before lockout.",
                            'failed_count': brute_after['failed_count'],
                            'attempts_remaining': remaining_attempts,
                        })
                    temp_expires_at = getattr(tenant_user_by_tid, 'temp_password_expires_at', None)
                    if temp_expires_at and temp_expires_at <= timezone.now():
                        log_access('Login', 'Failed', email, ip)
                        return render_login({
                            'error': (
                                'Temporary password has expired after 24 hours. '
                                'Please contact your tenant administrator for a new password.'
                            ),
                        })
                    request.session['pending_tenant_id'] = str(tenant_by_tid.tenant_id)
                    request.session['pending_tenant_user_id'] = str(tenant_user_by_tid.user_id)
                    request.session['pending_tenant_email'] = (tenant_user_by_tid.email or '').strip().lower()
                    request.session['pending_tenant_ip'] = ip
                    request.session.modified = True
                    try:
                        _issue_tenant_login_otp(
                            request,
                            tenant_by_tid,
                            recipient_email=tenant_user_by_tid.email,
                            recipient_name=tenant_user_by_tid.full_name,
                        )
                    except Exception:
                        logger.exception('Failed to dispatch tenant-user login OTP for %s', tenant_user_by_tid.email)
                        return render_login({
                            'error': (
                                'Could not send verification code right now. '
                                'Please try again.'
                            ),
                        })
                    messages.success(
                        request,
                        'A verification code has been sent to your email. Enter it to continue.',
                    )
                    return redirect('otp_verify')
                if tenant_user_resolution == 'tenant_user_ambiguous':
                    log_access('Login', 'Failed', email, ip)
                    return render_login({
                        'error': (
                            'This email is linked to multiple tenant workspaces. '
                            'Please use your tenant-specific login link.'
                        ),
                    })
                if tenant_user_resolution in ('tenant_user_inactive', 'tenant_user_role_inactive'):
                    log_access('Login', 'Failed', email, ip)
                    return render_login({'error': 'User account is not active for login.'})

            tenant, tenant_resolution = _resolve_tenant_by_email(email)
            if not tenant:
                tenant, tenant_user, tenant_user_resolution = _resolve_tenant_user_by_tid_and_email(
                    tid,
                    email,
                )
                if tenant and tenant_user:
                    if not check_password(
                        form.cleaned_data['password'],
                        tenant_user.password_hash,
                    ):
                        record_failed_attempt(email)
                        log_access('Login', 'Failed', email, ip)
                        brute_after = check_brute_force(email)
                        if brute_after['is_locked']:
                            return render_login({
                                'error': 'Account locked due to too many failed attempts.',
                                'is_locked': True,
                                'remaining_seconds': brute_after.get('remaining_seconds', 0),
                                'lockout_duration_total': settings_obj.lockout_duration_minutes * 60,
                                'attempts_remaining': 0,
                                'locked_email': email,
                            })
                        remaining_attempts = max(0, settings_obj.max_failed_logins - brute_after['failed_count'])
                        return render_login({
                            'error': f"Invalid email or password. {remaining_attempts} attempts remaining before lockout.",
                            'failed_count': brute_after['failed_count'],
                            'attempts_remaining': remaining_attempts,
                        })
                    temp_expires_at = getattr(tenant_user, 'temp_password_expires_at', None)
                    if temp_expires_at and temp_expires_at <= timezone.now():
                        log_access('Login', 'Failed', email, ip)
                        return render_login({
                            'error': (
                                'Temporary password has expired after 24 hours. '
                                'Please contact your tenant administrator for a new password.'
                            ),
                        })
                    request.session['pending_tenant_id'] = str(tenant.tenant_id)
                    request.session['pending_tenant_user_id'] = str(tenant_user.user_id)
                    request.session['pending_tenant_email'] = (tenant_user.email or '').strip().lower()
                    request.session['pending_tenant_ip'] = ip
                    request.session.modified = True
                    try:
                        _issue_tenant_login_otp(
                            request,
                            tenant,
                            recipient_email=tenant_user.email,
                            recipient_name=tenant_user.full_name,
                        )
                    except Exception:
                        logger.exception('Failed to dispatch tenant-user login OTP for %s', tenant_user.email)
                        return render_login({
                            'error': (
                                'Could not send verification code right now. '
                                'Please try again.'
                            ),
                        })
                    messages.success(
                        request,
                        'A verification code has been sent to your email. Enter it to continue.',
                    )
                    return redirect('otp_verify')

                if tenant_resolution in ('ambiguous_active', 'ambiguous_inactive'):
                    log_access('Login', 'Failed', email, ip)
                    return render_login({
                        'error': (
                            'This email is linked to multiple tenant accounts. '
                            'Please contact administrator support.'
                        ),
                    })
                if tenant_user_resolution == 'tenant_user_ambiguous':
                    log_access('Login', 'Failed', email, ip)
                    return render_login({
                        'error': (
                            'This email is linked to multiple tenant workspaces. '
                            'Please use your tenant-specific login link.'
                        ),
                    })
                if tenant_user_resolution in ('tenant_user_inactive', 'tenant_user_role_inactive'):
                    log_access('Login', 'Failed', email, ip)
                    return render_login({'error': 'User account is not active for login.'})
                record_failed_attempt(email)
                log_access('Login', 'Failed', email, ip)
                brute_after = check_brute_force(email)
                
                if brute_after['is_locked']:
                    return render_login({
                        'error': 'Account locked due to too many failed attempts.',
                        'is_locked': True,
                        'remaining_seconds': brute_after.get('remaining_seconds', 0),
                        'lockout_duration_total': settings_obj.lockout_duration_minutes * 60,
                        'attempts_remaining': 0,
                        'locked_email': email,
                    })

                remaining_attempts = max(0, settings_obj.max_failed_logins - brute_after['failed_count'])
                return render_login({
                    'error': f"Invalid email or password. {remaining_attempts} attempts remaining before lockout.",
                    'failed_count': brute_after['failed_count'],
                    'attempts_remaining': remaining_attempts,
                })
            if tenant.account_status != 'Active':
                log_access('Login', 'Failed', email, ip)
                return render_login({'error': 'Tenant account is not active.'})
            if not (tenant.portal_bootstrap_password_hash or '').strip():
                log_access('Login', 'Failed', email, ip)
                return render_login({'error': 'Tenant password is not provisioned yet.'})
            if not check_password(
                form.cleaned_data['password'],
                tenant.portal_bootstrap_password_hash,
            ):
                record_failed_attempt(email)
                log_access('Login', 'Failed', email, ip)
                brute_after = check_brute_force(email)
                
                if brute_after['is_locked']:
                    return render_login({
                        'error': 'Account locked due to too many failed attempts.',
                        'is_locked': True,
                        'remaining_seconds': brute_after.get('remaining_seconds', 0),
                        'lockout_duration_total': settings_obj.lockout_duration_minutes * 60,
                        'attempts_remaining': 0,
                        'locked_email': email,
                    })

                remaining_attempts = max(0, settings_obj.max_failed_logins - brute_after['failed_count'])
                return render_login({
                    'error': f"Invalid email or password. {remaining_attempts} attempts remaining before lockout.",
                    'failed_count': brute_after['failed_count'],
                    'attempts_remaining': remaining_attempts,
                })

            request.session['pending_tenant_id'] = str(tenant.tenant_id)
            request.session.pop('pending_tenant_user_id', None)
            request.session['pending_tenant_email'] = (tenant.primary_email or '').strip().lower()
            request.session['pending_tenant_ip'] = ip
            request.session.modified = True
            try:
                _issue_tenant_login_otp(request, tenant)
            except Exception:
                logger.exception('Failed to dispatch tenant login OTP for %s', tenant.primary_email)
                return render_login({
                    'error': (
                        'Could not send verification code right now. '
                        'Please try again.'
                    ),
                })
            messages.success(
                request,
                'A verification code has been sent to your email. Enter it to continue.',
            )
            return redirect('otp_verify')

        # STEP 3: Check status
        if user.status == 'Suspended':
            log_access('Login', 'Failed', email, ip)
            return render_login({
                'error': (
                    'Your account has been suspended. '
                    'Contact your administrator.'
                ),
            })

        if user.status == 'Pending_Activation':
            log_access('Login', 'Failed', email, ip)
            return render_login({
                'error': (
                    'Your account is not yet activated. '
                    'Please use the invite link sent to you.'
                ),
            })

        # STEP 4: Check password (never log plaintext password)
        password_ok = user.check_password(form.cleaned_data['password'])
        if not password_ok:
            logger.warning(
                'Admin login: password mismatch (user_id=%s email=%s ip=%s '
                'has_usable_password=%s password_len=%s)',
                user.pk,
                email,
                ip,
                user.has_usable_password(),
                len(form.cleaned_data.get('password') or ''),
            )
            record_failed_attempt(email)
            log_access('Login', 'Failed', email, ip)

            # Re-check AFTER recording — may have just hit the limit
            brute_after = check_brute_force(email)
            if brute_after['is_locked']:
                return render_login({
                    'error': 'Account locked due to too many failed attempts.',
                    'is_locked': True,
                    'remaining_seconds': brute_after.get('remaining_seconds', 0),
                    'lockout_duration_total': settings_obj.lockout_duration_minutes * 60,
                    'attempts_remaining': 0,
                    'locked_email': email,
                })

            settings_obj = get_security_settings()
            remaining_attempts = (
                settings_obj.max_failed_logins - brute_after['failed_count']
            )
            return render_login({
                'error': f"Invalid email or password. {remaining_attempts} attempts remaining before lockout.",
                'failed_count': brute_after['failed_count'],
                'attempts_remaining': remaining_attempts,
            })

        # STEP 5: Password verified — continue with OTP verification
        request.session['pending_admin_id'] = str(getattr(user, 'admin_id', user.id))
        request.session['pending_admin_email'] = (user.email or '').strip().lower()
        request.session['pending_admin_ip'] = ip
        request.session.modified = True
        try:
            _issue_admin_login_otp(request, user)
        except Exception:
            logger.exception('Failed to dispatch login OTP for %s', user.email)
            return render_login({
                'error': (
                    'Could not send verification code right now. '
                    'Please try again.'
                ),
            })
        messages.success(
            request,
            'A verification code has been sent to your email. Enter it to continue.',
        )
        return redirect('otp_verify')


@method_decorator(ensure_csrf_cookie, name='dispatch')
class OTPVerificationView(View):
    template_name = 'auth/otp_verify.html'

    @staticmethod
    def _otp_page_context(domain):
        is_tenant = domain == 'tenant'
        return {
            'otp_page_title': 'Tenant OTP Verification' if is_tenant else 'OTP Verification',
            'otp_meta_description': (
                'iRoad Tenant Portal - OTP Verification'
                if is_tenant
                else 'iRoad Admin Dashboard - OTP Verification'
            ),
            'otp_portal_label': 'Tenant Portal' if is_tenant else 'Admin Portal',
        }

    @staticmethod
    def _clear_admin_pending(request):
        request.session.pop(ADMIN_LOGIN_OTP_SESSION_KEY, None)
        request.session.pop('pending_admin_id', None)
        request.session.pop('pending_admin_email', None)
        request.session.pop('pending_admin_ip', None)

    @staticmethod
    def _clear_tenant_pending(request):
        request.session.pop(TENANT_LOGIN_OTP_SESSION_KEY, None)
        request.session.pop('pending_tenant_id', None)
        request.session.pop('pending_tenant_user_id', None)
        request.session.pop('pending_tenant_email', None)
        request.session.pop('pending_tenant_ip', None)

    @staticmethod
    def _pending_state(request):
        admin_otp_payload = request.session.get(ADMIN_LOGIN_OTP_SESSION_KEY) or {}
        admin_id = request.session.get('pending_admin_id')
        admin_email = (request.session.get('pending_admin_email') or '').strip().lower()
        if admin_otp_payload and admin_id and admin_email:
            otp_payload = admin_otp_payload
            domain = 'admin'
            identity_id = admin_id
            email = admin_email
            max_attempts = ADMIN_LOGIN_OTP_MAX_ATTEMPTS
        else:
            tenant_otp_payload = request.session.get(TENANT_LOGIN_OTP_SESSION_KEY) or {}
            tenant_id = request.session.get('pending_tenant_id')
            tenant_email = (request.session.get('pending_tenant_email') or '').strip().lower()
            if not tenant_otp_payload or not tenant_id or not tenant_email:
                return None
            otp_payload = tenant_otp_payload
            domain = 'tenant'
            identity_id = tenant_id
            email = tenant_email
            max_attempts = TENANT_LOGIN_OTP_MAX_ATTEMPTS

        expires_raw = otp_payload.get('expires_at')
        try:
            expires_at = datetime.fromisoformat(expires_raw)
        except Exception:
            return None
        if timezone.is_naive(expires_at):
            expires_at = timezone.make_aware(expires_at, timezone.get_current_timezone())
        remaining = max(0, int((expires_at - timezone.now()).total_seconds()))

        from .auth_helpers import get_security_settings
        settings_obj = get_security_settings()
        total_ttl = getattr(settings_obj, 'otp_timeout_seconds', 300)

        return {
            'otp_payload': otp_payload,
            'identity_id': identity_id,
            'email': email,
            'remaining': remaining,
            'total_ttl': total_ttl,
            'domain': domain,
            'max_attempts': max_attempts,
        }

    def get(self, request):
        pending = self._pending_state(request)
        if not pending:
            if request.user.is_authenticated:
                return redirect('dashboard')
            tenant_auth = get_tenant_portal_cookie_payload(request)
            if tenant_auth:
                return redirect(reverse('iroad_tenants:tenant_dashboard'))
            messages.error(request, 'Your verification session has expired. Please sign in again.')
            return redirect('login')
        otp_payload = pending['otp_payload']
        email = pending['email']
        remaining = pending['remaining']
        total_ttl = pending.get('total_ttl', 300)

        if remaining <= 0:
            if pending['domain'] == 'admin':
                self._clear_admin_pending(request)
            else:
                self._clear_tenant_pending(request)
            messages.error(request, 'OTP expired. Please sign in again to get a new code.')
            return redirect('login')

        return render(
            request,
            self.template_name,
            {
                'form': OTPVerificationForm(),
                'masked_email': email,
                'remaining_seconds': remaining,
                'total_ttl': total_ttl,
                'attempts_left': max(0, pending['max_attempts'] - int(otp_payload.get('attempts', 0))),
                **self._otp_page_context(pending['domain']),
            },
        )

    def post(self, request):
        action = request.POST.get('action')
        pending = self._pending_state(request)
        if not pending:
            if request.user.is_authenticated:
                return redirect('dashboard')
            tenant_auth = get_tenant_portal_cookie_payload(request)
            if tenant_auth:
                return redirect(reverse('iroad_tenants:tenant_dashboard'))
            messages.error(request, 'Your verification session has expired. Please sign in again.')
            return redirect('login')
        otp_payload = pending['otp_payload']
        identity_id = pending['identity_id']
        email = pending['email']
        remaining = pending['remaining']
        if remaining <= 0:
            if pending['domain'] == 'admin':
                self._clear_admin_pending(request)
            else:
                self._clear_tenant_pending(request)
            messages.error(request, 'OTP expired. Please sign in again to get a new code.')
            return redirect('login')

        if pending['domain'] == 'admin':
            user = AdminUser.objects.filter(pk=identity_id, email__iexact=email).first()
            if not user or user.status != 'Active':
                self._clear_admin_pending(request)
                messages.error(request, 'Account no longer eligible for login. Please contact administrator.')
                return redirect('login')
            tenant = None
        else:
            tenant = TenantProfile.objects.filter(pk=identity_id).first()
            if not tenant or tenant.account_status != 'Active':
                self._clear_tenant_pending(request)
                messages.error(request, 'Account no longer eligible for login. Please contact administrator.')
                return redirect('login')
            tenant_user_id = str(request.session.get('pending_tenant_user_id') or '').strip()
            tenant_user = None
            if tenant_user_id:
                registry = (
                    TenantRegistry.objects.select_related('tenant_profile')
                    .filter(tenant_profile_id=tenant.tenant_id)
                    .first()
                )
                if not registry:
                    self._clear_tenant_pending(request)
                    messages.error(request, 'Tenant workspace is unavailable. Please contact administrator.')
                    return redirect('login')
                from tenant_workspace.models import TenantRole, TenantUser
                connection.set_schema_to_public()
                try:
                    connection.set_tenant(registry)
                    tenant_user = TenantUser.objects.filter(
                        pk=tenant_user_id,
                        email__iexact=email,
                    ).first()
                    if not tenant_user or tenant_user.status != TenantUser.Status.ACTIVE:
                        self._clear_tenant_pending(request)
                        messages.error(request, 'Account no longer eligible for login. Please contact administrator.')
                        return redirect('login')
                    role = TenantRole.objects.filter(
                        role_name_en__iexact=(tenant_user.role_name or '').strip()
                    ).first()
                    if role and role.status != TenantRole.Status.ACTIVE:
                        self._clear_tenant_pending(request)
                        messages.error(request, 'Assigned role is inactive. Please contact administrator.')
                        return redirect('login')
                finally:
                    connection.set_schema_to_public()
            user = None

        if action == 'resend':
            if pending['domain'] == 'admin':
                try:
                    _issue_admin_login_otp(request, user)
                except Exception:
                    logger.exception('Failed to resend login OTP for %s', user.email)
                    messages.error(request, 'Failed to resend OTP right now. Please try again shortly.')
                else:
                    messages.success(request, 'A new OTP has been sent to your email.')
            else:
                try:
                    recipient_email = email
                    recipient_name = (
                        getattr(tenant_user, 'full_name', None)
                        if tenant_user is not None
                        else None
                    )
                    _issue_tenant_login_otp(
                        request,
                        tenant,
                        recipient_email=recipient_email,
                        recipient_name=recipient_name,
                    )
                except Exception:
                    logger.exception('Failed to resend tenant login OTP for %s', tenant.primary_email)
                    messages.error(request, 'Failed to resend OTP right now. Please try again shortly.')
                else:
                    messages.success(request, 'A new OTP has been sent to your email.')
            return redirect('otp_verify')

        form = OTPVerificationForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'masked_email': email,
                    'remaining_seconds': remaining,
                    'attempts_left': max(0, pending['max_attempts'] - int(otp_payload.get('attempts', 0))),
                    **self._otp_page_context(pending['domain']),
                },
            )

        if form.cleaned_data['otp'] != str(otp_payload.get('otp_code') or ''):
            attempts = int(otp_payload.get('attempts', 0)) + 1
            otp_payload['attempts'] = attempts
            session_key = (
                ADMIN_LOGIN_OTP_SESSION_KEY
                if pending['domain'] == 'admin'
                else TENANT_LOGIN_OTP_SESSION_KEY
            )
            request.session[session_key] = otp_payload
            request.session.modified = True
            attempts_left = max(0, pending['max_attempts'] - attempts)
            if attempts >= pending['max_attempts']:
                if pending['domain'] == 'admin':
                    self._clear_admin_pending(request)
                else:
                    self._clear_tenant_pending(request)
                messages.error(request, 'Too many invalid OTP attempts. Please sign in again.')
                return redirect('login')
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'error': f'Invalid OTP. {attempts_left} attempt(s) remaining.',
                    'masked_email': email,
                    'remaining_seconds': remaining,
                    'attempts_left': attempts_left,
                    **self._otp_page_context(pending['domain']),
                },
            )

        if pending['domain'] == 'admin':
            # OTP verified successfully -> complete admin login + session creation.
            reset_failed_attempts(email)
            user.last_login_at = timezone.now()
            user.two_factor_enabled = True
            user.save(update_fields=['last_login_at', 'two_factor_enabled'])
            login(request, user)

            settings_obj = get_security_settings()
            user_agent = request.META.get('HTTP_USER_AGENT', '')
            ip = request.session.get('pending_admin_ip') or _client_ip(request)
            jti = create_admin_session(
                admin_user=user,
                ip_address=ip,
                user_agent=user_agent,
                timeout_minutes=settings_obj.session_timeout_minutes,
            )
            request.session['jti'] = jti
            create_session(request, user, user_domain='Admin', redis_jti=jti)
            log_audit_action(
                request=request,
                action_type='Create',
                module_name='Auth - Login',
                record_id=str(getattr(user, 'admin_id', user.id)),
                new_instance=None,
            )
            log_access('Login', 'Success', email, ip)

            self._clear_admin_pending(request)
            request.session['notify_auth_tabs'] = True
            messages.success(request, 'OTP verified successfully.')
            return redirect('dashboard')

        from .tenant_jwt import sign_tenant_access_jwt
        from .redis_helpers import create_tenant_session, revoke_tenant_session_key
        from .models import TenantSecuritySettings

        sec = TenantSecuritySettings.objects.first()
        tenant_timeout_min = max(
            60,
            int(getattr(sec, 'tenant_web_timeout_hours', 12)) * 60,
        )
        ttl_seconds = max(300, int(tenant_timeout_min) * 60)
        subject_email = (
            (tenant_user.email or '').strip()
            if tenant_user
            else (tenant.primary_email or '').strip()
        )
        extra_claims = {
            'portal_actor': 'tenant_user' if tenant_user else 'tenant_admin',
            'email': subject_email,
            'tenant_user_id': str(tenant_user.user_id) if tenant_user else '',
            'sid': str(tenant_user.user_id) if tenant_user else '',
            'role_name': (tenant_user.role_name or '').strip() if tenant_user else 'Tenant Admin',
            'display_name': (
                (tenant_user.full_name or tenant_user.email or '').strip()
                if tenant_user
                else (tenant.company_name or tenant.primary_email or '').strip()
            ),
        }
        token_str, jti = sign_tenant_access_jwt(
            tenant_id=tenant.tenant_id,
            subject=subject_email or str(tenant.tenant_id),
            token_type='tenant_access',
            ttl_seconds=ttl_seconds,
            extra_claims=extra_claims,
        )
        # Keep existing tenant sessions for other tabs/profiles.
        # We only add/refresh the current tenant session cookie entry below.

        create_tenant_session(
            tenant_id=str(tenant.tenant_id),
            user_domain='Tenant_User',
            reference_id=str(tenant_user.user_id) if tenant_user else str(tenant.tenant_id),
            reference_name=(
                (tenant_user.full_name or tenant_user.email or '').strip()
                if tenant_user
                else (tenant.company_name or tenant.primary_email or '').strip()
            ),
            ip_address=request.session.get('pending_tenant_ip') or _client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            timeout_minutes=tenant_timeout_min,
            jti=jti,
        )

        if tenant_user:
            registry = (
                TenantRegistry.objects.select_related('tenant_profile')
                .filter(tenant_profile_id=tenant.tenant_id)
                .first()
            )
            if registry:
                from tenant_workspace.models import TenantUser
                connection.set_schema_to_public()
                try:
                    connection.set_tenant(registry)
                    TenantUser.objects.filter(pk=tenant_user.user_id).update(
                        last_login_at=timezone.now(),
                        login_attempts=0,
                        temp_password_expires_at=None,
                    )
                finally:
                    connection.set_schema_to_public()

        # Backward-compat cleanup for older session-based tenant bootstrap.
        for key in ('tenant_bootstrap_token', 'tenant_bootstrap_tenant_id', 'tenant_bootstrap_jti', 'tenant_bootstrap_expires_in'):
            request.session.pop(key, None)
        reset_failed_attempts(email)
        ip = request.session.get('pending_tenant_ip') or _client_ip(request)
        log_access('Login', 'Success', email, ip)
        self._clear_tenant_pending(request)
        messages.success(request, 'OTP verified successfully.')
        response = redirect(reverse('iroad_tenants:tenant_dashboard'))
        set_tenant_portal_cookie(
            response,
            tenant.tenant_id,
            jti,
            request=request,
            access_jwt=token_str,
        )
        return response


class LogoutView(View):
    def get(self, request):
        if request.user.is_authenticated:
            from superadmin.redis_helpers import revoke_admin_session
            from django.contrib.auth import logout as auth_logout

            email = request.user.email
            ip = _client_ip(request)

            jti = request.session.get('jti')
            if jti:
                revoke_admin_session(jti)

            close_session(request)
            auth_logout(request)
            log_access('Logout', 'Success', email, ip)
            # Bridge page notifies other tabs (localStorage) then goes to login.
            return render(request, 'auth/logout_bridge.html')
        return redirect('login')

    def post(self, request):
        if request.user.is_authenticated:
            from superadmin.redis_helpers import revoke_admin_session
            from django.contrib.auth import logout as auth_logout

            email = request.user.email
            ip = _client_ip(request)

            jti = request.session.get('jti')
            if jti:
                revoke_admin_session(jti)

            close_session(request)
            auth_logout(request)
            log_access('Logout', 'Success', email, ip)
            return render(request, 'auth/logout_bridge.html')
        return redirect('login')


class ForgotPasswordView(View):
    """Request password reset (email always gets same response text)."""

    template_request = 'auth/reset_password.html'
    success_message = (
        'If this email exists in our system, '
        'a reset link has been generated.'
    )

    def get(self, request):
        return render(
            request,
            self.template_request,
            {'form': ForgotPasswordForm()},
        )

    def post(self, request):
        form = ForgotPasswordForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_request,
                {'form': form},
            )

        email = form.cleaned_data['email'].lower().strip()
        reset_url = None

        try:
            user = AdminUser.objects.get(email=email)
        except AdminUser.DoesNotExist:
            user = None

        if user and user.status in ('Active', 'Pending_Activation'):
            # Generate token only for valid accounts
            token = create_auth_token(user, 'password_reset')
            reset_url = request.build_absolute_uri(
                f'/new-password/{token.token}/'
            )
            send_auth_email(
                user,
                "password_reset",
                {"admin_user": user, "reset_url": reset_url},
            )
        else:
            tenant_user_tenant, tenant_user, tenant_user_resolution = _resolve_tenant_user_by_tid_and_email(
                None,
                email,
            )
            if tenant_user_tenant and tenant_user and tenant_user_resolution is None:
                registry = (
                    TenantRegistry.objects.select_related('tenant_profile')
                    .filter(tenant_profile_id=tenant_user_tenant.tenant_id)
                    .first()
                )
                if registry:
                    from tenant_workspace.models import TenantUser
                    connection.set_schema_to_public()
                    try:
                        connection.set_tenant(registry)
                        tenant_user_record = TenantUser.objects.filter(
                            pk=tenant_user.user_id,
                            email__iexact=email,
                        ).first()
                        if tenant_user_record and tenant_user_record.status == TenantUser.Status.ACTIVE:
                            reset_token = _build_tenant_user_password_reset_token(
                                tenant_id=tenant_user_tenant.tenant_id,
                                email=tenant_user_record.email,
                            )
                            reset_url = request.build_absolute_uri(
                                reverse('tenant_user_new_password', args=[reset_token])
                            )
                            context_dict = {
                                'name': tenant_user_record.full_name,
                                'email': tenant_user_record.email,
                                'reset_url': reset_url,
                                'login_url': reset_url,
                                'user_name': tenant_user_record.full_name,
                            }
                            sent = send_named_notification_email(
                                'TENANT_USER_PASSWORD_RESET',
                                recipient_email=tenant_user_record.email,
                                context_dict=context_dict,
                                language='en',
                                default_subject='Reset Your iRoad Password',
                                trigger_source='TemplateName: TENANT_USER_PASSWORD_RESET',
                                force_django_smtp=True,
                            )
                            if not sent:
                                logger.warning(
                                    'No active notification template found for TENANT_USER_PASSWORD_RESET'
                                )
                    finally:
                        connection.set_schema_to_public()
            else:
                tenant, tenant_resolution = _resolve_tenant_by_email(email)
                if tenant and tenant.account_status == 'Active':
                    TenantAuthToken.objects.filter(
                        tenant_profile=tenant,
                        token_type=TenantAuthToken.TokenType.INVITE,
                        is_used=False,
                    ).update(is_used=True)
                    tenant_token = TenantAuthToken.objects.create(
                        tenant_profile=tenant,
                        token=secrets.token_urlsafe(32),
                        token_type=TenantAuthToken.TokenType.INVITE,
                        expires_at=timezone.now() + timedelta(hours=1),
                    )
                    reset_url = request.build_absolute_uri(
                        reverse('set_password', args=[tenant_token.token])
                    )
                    send_named_notification_email(
                        'TENANT_PASSWORD_RESET',
                        recipient_email=tenant.primary_email,
                        context_dict={
                            'admin_user': {'first_name': tenant.company_name},
                            'tenant': {'company_name': tenant.company_name},
                            'reset_url': reset_url,
                        },
                        language='en',
                        default_subject='Reset Your iRoad Password',
                        trigger_source='TemplateName: TENANT_PASSWORD_RESET',
                        force_django_smtp=True,
                    )
                elif tenant_resolution in ('ambiguous_active', 'ambiguous_inactive'):
                    logger.warning(
                        'Password reset skipped for duplicated tenant email: %s',
                        email,
                    )

        # Always show the same success_message text to avoid information leaks
        return render(
            request,
            self.template_request,
            {
                'form': ForgotPasswordForm(),
                'success_message': self.success_message,
                'reset_url': reset_url,
            },
        )



class ResetPasswordConfirmView(View):
    """Public: choose new password using password_reset token."""

    template_form = 'auth/new_password.html'
    template_error = 'auth/token_error.html'

    def _render_error(self, request, message):
        return render(
            request,
            self.template_error,
            {'error_message': message},
        )

    def _get_reset_token(self, raw_token):
        try:
            return AdminAuthToken.objects.select_related('admin_user').get(
                token=raw_token,
                token_type=AdminAuthToken.TokenType.PASSWORD_RESET,
            )
        except AdminAuthToken.DoesNotExist:
            return None

    def get(self, request, token):
        reset_tok = self._get_reset_token(token)
        if reset_tok is None:
            return self._render_error(request, 'Invalid reset link.')
        if reset_tok.is_used:
            return self._render_error(
                request,
                'This reset link has already been used.',
            )
        if reset_tok.is_expired:
            return self._render_error(
                request,
                'This reset link has expired.',
            )
        return render(
            request,
            self.template_form,
            {
                'form': SetPasswordForm(),
                'account_email': reset_tok.admin_user.email,
            },
        )

    def post(self, request, token):
        reset_tok = self._get_reset_token(token)
        if reset_tok is None:
            return self._render_error(request, 'Invalid reset link.')
        if reset_tok.is_used:
            return self._render_error(
                request,
                'This reset link has already been used.',
            )
        if reset_tok.is_expired:
            return self._render_error(
                request,
                'This reset link has expired.',
            )

        form = SetPasswordForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_form,
                {
                    'form': form,
                    'account_email': reset_tok.admin_user.email,
                },
            )

        user = reset_tok.admin_user
        user.set_password(form.cleaned_data['password'])
        user.save(update_fields=['password'])

        reset_tok.is_used = True
        reset_tok.save(update_fields=['is_used'])

        reset_failed_attempts(user.email)

        messages.success(
            request,
            'Password reset successful. Please login.',
        )
        return redirect(reverse('login'))


class TenantUserResetPasswordConfirmView(View):
    """Public: set a new password for tenant user via signed reset token."""

    template_form = 'auth/new_password.html'
    template_error = 'auth/token_error.html'

    def _render_error(self, request, message):
        return render(
            request,
            self.template_error,
            {'error_message': message},
        )

    def _resolve_token_user(self, token):
        payload = _read_tenant_user_password_reset_token(token)
        if not payload:
            return None, None
        tenant_id = (payload.get('tenant_id') or '').strip()
        email = (payload.get('email') or '').strip().lower()
        if not tenant_id or not email:
            return None, None
        registry = (
            TenantRegistry.objects.select_related('tenant_profile')
            .filter(tenant_profile_id=tenant_id)
            .first()
        )
        if not registry:
            return None, None
        from tenant_workspace.models import TenantUser

        connection.set_schema_to_public()
        try:
            connection.set_tenant(registry)
            tenant_user = TenantUser.objects.filter(
                email__iexact=email,
                status=TenantUser.Status.ACTIVE,
            ).first()
            if not tenant_user:
                return None, None
            return registry, tenant_user
        finally:
            connection.set_schema_to_public()

    def get(self, request, token):
        registry, tenant_user = self._resolve_token_user(token)
        if not registry or not tenant_user:
            return self._render_error(request, 'Invalid or expired reset link.')
        return render(
            request,
            self.template_form,
            {
                'form': SetPasswordForm(),
                'account_email': tenant_user.email,
            },
        )

    def post(self, request, token):
        registry, tenant_user = self._resolve_token_user(token)
        if not registry or not tenant_user:
            return self._render_error(request, 'Invalid or expired reset link.')

        form = SetPasswordForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_form,
                {
                    'form': form,
                    'account_email': tenant_user.email,
                },
            )

        from tenant_workspace.models import TenantUser

        connection.set_schema_to_public()
        try:
            connection.set_tenant(registry)
            fresh_user = TenantUser.objects.filter(pk=tenant_user.pk).first()
            if not fresh_user or fresh_user.status != TenantUser.Status.ACTIVE:
                return self._render_error(request, 'User account is not available.')
            fresh_user.password_hash = make_password(form.cleaned_data['password'])
            fresh_user.temp_password_expires_at = None
            fresh_user.save(update_fields=['password_hash', 'temp_password_expires_at', 'updated_at'])
        finally:
            connection.set_schema_to_public()

        messages.success(request, 'Password reset successful. Please login.')
        login_url = f"{reverse('login')}?tid={registry.tenant_profile.tenant_id}"
        return redirect(login_url)


class DashboardView(LoginRequiredMixin, View):
    template_name = 'dashboard/dashboard.html'

    def get(self, request):
        now = timezone.now()
        auth_tab_sync_bump = request.session.pop(
            'notify_auth_tabs',
            False,
        )

        # ── IMPORT ALL NEEDED MODELS ──
        from .models import (
            TenantProfile, SubscriptionOrder,
            Transaction, SupportTicket, AdminUser,
            StandardInvoice, AuditLog, ActiveSession,
            Role
        )

        # ── EXECUTIVE KPIs ──

        # MRR: current month paid/issued invoices
        month_start = now.replace(
            day=1, hour=0, minute=0,
            second=0, microsecond=0)
        mrr = StandardInvoice.objects.filter(
            status__in=['Issued', 'Paid'],
            issue_date__gte=month_start
        ).aggregate(
            total=Sum('base_currency_equivalent_amount')
        )['total'] or Decimal('0.00')

        period_30d_start = now - timedelta(days=30)
        period_60d_start = now - timedelta(days=60)
        revenue_30d = StandardInvoice.objects.filter(
            status__in=['Issued', 'Paid'],
            issue_date__gte=period_30d_start,
            issue_date__lte=now,
        ).aggregate(
            t=Sum('base_currency_equivalent_amount')
        )['t'] or Decimal('0.00')
        revenue_prev_30d = StandardInvoice.objects.filter(
            status__in=['Issued', 'Paid'],
            issue_date__gte=period_60d_start,
            issue_date__lt=period_30d_start,
        ).aggregate(
            t=Sum('base_currency_equivalent_amount')
        )['t'] or Decimal('0.00')
        if revenue_prev_30d > 0:
            revenue_30d_growth_pct = float(
                (revenue_30d - revenue_prev_30d) / revenue_prev_30d * 100)
        elif revenue_30d > 0:
            revenue_30d_growth_pct = 100.0
        else:
            revenue_30d_growth_pct = 0.0

        active_tenants = TenantProfile.objects.filter(
            account_status='Active').count()
        tenants_active_month_start = TenantProfile.objects.filter(
            account_status='Active',
            registered_at__lt=month_start,
        ).count()
        active_tenants_net_mtd = active_tenants - tenants_active_month_start
        new_signups_mtd = TenantProfile.objects.filter(
            registered_at__gte=month_start,
        ).count()

        fleet_totals = TenantProfile.objects.filter(
            account_status='Active',
        ).aggregate(
            trucks=Sum(
                F('active_max_internal_trucks')
                + F('active_max_external_trucks')
            ),
            drivers=Sum('active_max_drivers'),
        )
        total_trucks_licensed = int(fleet_totals['trucks'] or 0)
        total_drivers_licensed = int(fleet_totals['drivers'] or 0)

        failed_payment_count = Transaction.objects.filter(
            status__in=['Failed', 'Rejected'],
        ).count()

        pending_orders = SubscriptionOrder.objects.filter(
            order_status='Pending_Payment').count()

        pending_bank_txns = Transaction.objects.filter(
            status='Pending',
            payment_method__method_type='Offline_Bank'
        ).count()

        open_tickets = SupportTicket.objects.filter(
            status__in=['New', 'In_Progress']).count()

        overdue_cutoff = now - timedelta(hours=48)
        overdue_tickets = SupportTicket.objects.filter(
            ~Q(status='Closed'),
            created_at__lt=overdue_cutoff
        ).count()

        active_admin_sessions = count_active_admin_sessions()

        # ── ATTENTION CENTER ──

        pending_orders_list = SubscriptionOrder.objects.filter(
            order_status='Pending_Payment'
        ).select_related('tenant').order_by(
            'created_at')[:5]

        overdue_tickets_list = SupportTicket.objects.filter(
            ~Q(status='Closed'),
            created_at__lt=overdue_cutoff
        ).select_related('tenant', 'category').order_by(
            'created_at')[:5]

        failed_transactions = Transaction.objects.filter(
            status__in=['Failed', 'Rejected']
        ).select_related('tenant').order_by(
            '-created_at')[:5]

        suspended_tenants = TenantProfile.objects.filter(
            ~Q(account_status='Active')
        ).order_by('-updated_at')[:5]

        # ── STAFF ANALYTICS ──

        role_distribution = Role.objects.filter(
            status='Active'
        ).annotate(
            staff_count=Count('admin_users')
        ).values('role_name_en', 'staff_count')

        stale_cutoff = now - timedelta(days=30)
        stale_accounts = AdminUser.objects.filter(
            Q(last_login_at__lt=stale_cutoff) |
            Q(last_login_at__isnull=True),
            status='Active'
        ).count()

        suspended_admins = AdminUser.objects.filter(
            status='Suspended').count()

        total_staff = AdminUser.objects.exclude(
            status='Suspended').count()

        total_tenants = TenantProfile.objects.count()
        inactive_tenants = max(0, total_tenants - active_tenants)

        attention_chart = {
            'labels': [
                'Pending orders',
                'Pending bank transfers',
                'Open tickets',
                'Overdue (>48h)',
                'Failed / rejected',
            ],
            'data': [
                pending_orders,
                pending_bank_txns,
                open_tickets,
                overdue_tickets,
                failed_payment_count,
            ],
        }
        tenant_status_chart = {
            'labels': ['Active subscribers', 'Other statuses'],
            'data': [active_tenants, inactive_tenants],
        }

        # ── REVENUE + LICENSED FLEET TREND (last 6 months) ──
        revenue_data = []
        revenue_labels = []
        fleet_trucks_series = []
        fleet_drivers_series = []
        for i in range(5, -1, -1):
            month_date = (now.replace(day=1) -
                          timedelta(days=i * 30))
            m_start = month_date.replace(
                day=1, hour=0, minute=0,
                second=0, microsecond=0)
            if m_start.month == 12:
                m_end = m_start.replace(
                    year=m_start.year + 1, month=1)
            else:
                m_end = m_start.replace(
                    month=m_start.month + 1)

            total = StandardInvoice.objects.filter(
                status__in=['Issued', 'Paid'],
                issue_date__gte=m_start,
                issue_date__lt=m_end
            ).aggregate(
                t=Sum('base_currency_equivalent_amount')
            )['t'] or Decimal('0.00')

            revenue_data.append(float(total))
            revenue_labels.append(
                m_start.strftime('%b %Y'))

            ft = TenantProfile.objects.filter(
                account_status='Active',
                registered_at__lt=m_end,
            ).aggregate(
                trucks=Sum(
                    F('active_max_internal_trucks')
                    + F('active_max_external_trucks')
                ),
                drivers=Sum('active_max_drivers'),
            )
            fleet_trucks_series.append(int(ft['trucks'] or 0))
            fleet_drivers_series.append(int(ft['drivers'] or 0))

        fleet_chart = {
            'labels': revenue_labels,
            'trucks': fleet_trucks_series,
            'drivers': fleet_drivers_series,
        }

        # ── RECENT AUDIT LOG ──
        recent_audit = AuditLog.objects.select_related(
            'admin'
        ).order_by('-timestamp')[:10]

        # ── ACCESS HEATMAP (Successful Logins last 30 days) ──
        heatmap_qs = AccessLog.objects.filter(
            attempt_type='Login',
            status='Success',
            timestamp__gte=now - timedelta(days=30)
        ).annotate(
            hour=ExtractHour('timestamp'),
            weekday=ExtractWeekDay('timestamp') # 1=Sun, 7=Sat
        ).values('hour', 'weekday').annotate(count=Count('id'))

        # Prepare 7x24 matrix (Monday-Sunday)
        heatmap_matrix = [[0] * 24 for _ in range(7)]
        # Map Django Weekday (1=Sun...7=Sat) to Matrix Row (0=Mon...6=Sun)
        day_map = {2:0, 3:1, 4:2, 5:3, 6:4, 7:5, 1:6}
        for entry in heatmap_qs:
            d_idx = day_map.get(entry['weekday'])
            if d_idx is not None:
                heatmap_matrix[d_idx][entry['hour']] = entry['count']

        context = {
            # KPIs
            'mrr': mrr,
            'revenue_30d': revenue_30d,
            'revenue_30d_growth_pct': revenue_30d_growth_pct,
            'active_tenants': active_tenants,
            'active_tenants_net_mtd': active_tenants_net_mtd,
            'new_signups_mtd': new_signups_mtd,
            'tenants_active_month_start': tenants_active_month_start,
            'total_trucks_licensed': total_trucks_licensed,
            'total_drivers_licensed': total_drivers_licensed,
            'failed_payment_count': failed_payment_count,
            'pending_orders': pending_orders,
            'pending_bank_txns': pending_bank_txns,
            'open_tickets': open_tickets,
            'overdue_tickets': overdue_tickets,
            'active_admin_sessions': active_admin_sessions,
            # Attention
            'pending_orders_list': pending_orders_list,
            'overdue_tickets_list': overdue_tickets_list,
            'failed_transactions': failed_transactions,
            'suspended_tenants': suspended_tenants,
            # Staff
            'role_distribution': list(role_distribution),
            'stale_accounts': stale_accounts,
            'suspended_admins': suspended_admins,
            'total_staff': total_staff,
            # Heatmap
            'heatmap_data': heatmap_matrix,
            # Charts (Chart.js)
            'revenue_data': revenue_data,
            'revenue_labels': revenue_labels,
            'fleet_chart': fleet_chart,
            'attention_chart': attention_chart,
            'tenant_status_chart': tenant_status_chart,
            # Audit
            'recent_audit': recent_audit,
            # Meta
            'page_title': _('Super Admin Dashboard'),
            'auth_tab_sync_bump': auth_tab_sync_bump,
        }

        return render(
            request,
            self.template_name,
            context)


class AccessLogListView(LoginRequiredMixin, View):
    template_name = 'security/logs/access_log_list.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        sort_by = request.GET.get('sort', 'rank').strip()
        sort_dir = request.GET.get('dir', 'desc').strip().lower()
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'desc'

        qs = AccessLog.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('timestamp').desc(),
            )
        )

        attempt_filter = request.GET.get('attempt_type', 'All')
        if attempt_filter in [
            AccessLog.AttemptType.LOGIN,
            AccessLog.AttemptType.LOGOUT,
            AccessLog.AttemptType.TOKEN_REFRESH,
        ]:
            qs = qs.filter(attempt_type=attempt_filter)

        status_filter = request.GET.get('status', 'All')
        if status_filter in [
            AccessLog.Status.SUCCESS,
            AccessLog.Status.FAILED,
            AccessLog.Status.BLOCKED,
        ]:
            qs = qs.filter(status=status_filter)

        domain_filter = request.GET.get('user_domain', 'All')
        if domain_filter in ['Admin', 'Tenant_User', 'Driver']:
            qs = qs.filter(user_domain=domain_filter)

        search_query = request.GET.get('q', '').strip()
        if search_query:
            qs = qs.filter(
                Q(email_used__icontains=search_query)
                | Q(ip_address__icontains=search_query)
            )

        from_date = request.GET.get('from_date', '').strip()
        to_date = request.GET.get('to_date', '').strip()
        fd = parse_date(from_date) if from_date else None
        td = parse_date(to_date) if to_date else None
        if fd:
            qs = qs.filter(timestamp__date__gte=fd)
        if td:
            qs = qs.filter(timestamp__date__lte=td)

        sort_mapping = {
            'rank': ['default_rank'],
            'attempt_type': ['attempt_type'],
            'status': ['status'],
            'user_domain': ['user_domain'],
            'email': ['email_used'],
            'ip': ['ip_address'],
            'timestamp': ['timestamp'],
        }
        active_sort_fields = sort_mapping.get(sort_by, ['default_rank'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else f'-{f}')
        qs = qs.order_by(*ordering)
        total_count = qs.count()

        paginator = Paginator(qs, 10)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        start_index = page_obj.start_index()
        for offset, log in enumerate(page_obj.object_list):
            log.list_rank = total_count - (start_index + offset) + 1

        query_params = request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page', None)

        context = {
            'access_logs': page_obj,
            'total_count': total_count,
            'page_title': 'Authentication Access Log',
            'status_filter': status_filter,
            'attempt_type_filter': attempt_filter,
            'domain_filter': domain_filter,
            'search_query': search_query,
            'from_date': from_date,
            'to_date': to_date,
            'current_sort': sort_by,
            'current_dir': sort_dir,
            'filter_query': query_params.urlencode(),
        }
        return render(request, self.template_name, context)


class AuditLogListView(LoginRequiredMixin, View):
    template_name = 'security/logs/audit_log_list.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        sort_by = request.GET.get('sort', 'rank').strip()
        sort_dir = request.GET.get('dir', 'desc').strip().lower()
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'desc'

        qs = AuditLog.objects.select_related('admin').all()

        action_filter = request.GET.get('action_type', 'All')
        if action_filter in ['Create', 'Update', 'Delete', 'Status_Change']:
            qs = qs.filter(action_type=action_filter)

        module_filter = request.GET.get('module_name', '').strip()
        if module_filter:
            qs = qs.filter(module_name=module_filter)

        admin_filter = request.GET.get('admin', '').strip()
        if admin_filter:
            qs = qs.filter(admin_id=admin_filter)

        from_date = request.GET.get('from_date', '').strip()
        to_date = request.GET.get('to_date', '').strip()
        fd = parse_date(from_date) if from_date else None
        td = parse_date(to_date) if to_date else None
        if fd:
            qs = qs.filter(timestamp__date__gte=fd)
        if td:
            qs = qs.filter(timestamp__date__lte=td)

        search_query = request.GET.get('q', '').strip()
        if search_query:
            qs = qs.filter(
                Q(module_name__icontains=search_query)
                | Q(record_id__icontains=search_query)
            )

        sort_mapping = {
            'rank': ['timestamp'],
            'admin': ['admin__first_name', 'admin__last_name', 'admin__email'],
            'action_type': ['action_type'],
            'module_name': ['module_name'],
            'ip': ['ip_address'],
            'timestamp': ['timestamp'],
        }
        active_sort_fields = sort_mapping.get(sort_by, ['timestamp'])
        ordering = []
        for field_name in active_sort_fields:
            ordering.append(
                field_name if sort_dir == 'asc' else f'-{field_name}'
            )
        qs = qs.order_by(*ordering)
        total_count = qs.count()

        paginator = Paginator(qs, 10)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        start_index = page_obj.start_index()
        for offset, log in enumerate(page_obj.object_list):
            log.list_rank = total_count - (start_index + offset) + 1

        query_params = request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page', None)

        context = {
            'audit_logs': page_obj,
            'action_filter': action_filter,
            'module_filter': module_filter,
            'admin_filter': admin_filter,
            'from_date': from_date,
            'to_date': to_date,
            'search_query': search_query,
            'current_sort': sort_by,
            'current_dir': sort_dir,
            'admins': AdminUser.objects.order_by('first_name', 'last_name'),
            'modules': AuditLog.objects.values_list(
                'module_name', flat=True
            ).distinct().order_by('module_name'),
            'filter_query': query_params.urlencode(),
        }
        return render(request, self.template_name, context)


class AuditLogDetailView(LoginRequiredMixin, View):
    template_name = 'security/logs/audit_log_detail.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        audit_entry = get_object_or_404(AuditLog, pk=pk)
        old_payload_pretty = (
            json.dumps(audit_entry.old_payload, indent=2, ensure_ascii=False)
            if audit_entry.old_payload is not None
            else None
        )
        new_payload_pretty = (
            json.dumps(audit_entry.new_payload, indent=2, ensure_ascii=False)
            if audit_entry.new_payload is not None
            else None
        )
        context = {
            'audit_entry': audit_entry,
            'old_payload_pretty': old_payload_pretty,
            'new_payload_pretty': new_payload_pretty,
        }
        return render(request, self.template_name, context)


class AdminSecuritySettingsView(LoginRequiredMixin, View):
    template_name = 'security/admin_security_settings.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        obj, _created = AdminSecuritySettings.objects.get_or_create(
            setting_id='ADMIN-SEC-CONF',
            defaults={
                'session_timeout_minutes': 1440,
                'max_failed_logins': 3,
                'lockout_duration_minutes': 30,
                'otp_timeout_seconds': 300,
            },
        )
        form = AdminSecuritySettingsForm(instance=obj)
        return render(request, self.template_name, {'form': form, 'obj': obj})

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        obj, _created = AdminSecuritySettings.objects.get_or_create(
            setting_id='ADMIN-SEC-CONF',
            defaults={
                'session_timeout_minutes': 1440,
                'max_failed_logins': 3,
                'lockout_duration_minutes': 30,
                'otp_timeout_seconds': 300,
            },
        )
        old_obj = AdminSecuritySettings.objects.get(setting_id=obj.setting_id)
        form = AdminSecuritySettingsForm(request.POST, instance=obj)
        if not form.is_valid():
            messages.error(
                request,
                'Could not save admin security settings. Please correct the highlighted fields.',
            )
            return render(request, self.template_name, {'form': form, 'obj': obj})
        form.instance.updated_by = request.user
        form.save()
        log_audit_action(
            request,
            'Update',
            'Admin Security Settings',
            'ADMIN-SEC-CONF',
            old_instance=old_obj,
            new_instance=obj,
        )
        messages.success(request, 'Admin security settings saved successfully.')
        return redirect(reverse('security_settings'))


class TenantSecuritySettingsView(LoginRequiredMixin, View):
    template_name = 'security/tenant_security_settings.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        obj, _created = TenantSecuritySettings.objects.get_or_create(
            setting_id='TENANT-SEC-CONF',
            defaults={
                'tenant_web_timeout_hours': 12,
                'driver_app_timeout_days': 30,
                'max_failed_logins': 5,
                'lockout_duration_minutes': 15,
            },
        )
        form = TenantSecuritySettingsForm(instance=obj)
        return render(request, self.template_name, {'form': form, 'obj': obj})

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        obj, _created = TenantSecuritySettings.objects.get_or_create(
            setting_id='TENANT-SEC-CONF',
            defaults={
                'tenant_web_timeout_hours': 12,
                'driver_app_timeout_days': 30,
                'max_failed_logins': 5,
                'lockout_duration_minutes': 15,
            },
        )
        old_obj = TenantSecuritySettings.objects.get(setting_id=obj.setting_id)
        form = TenantSecuritySettingsForm(request.POST, instance=obj)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form, 'obj': obj})

        form.instance.updated_by = request.user
        form.save()
        log_audit_action(
            request,
            'Update',
            'Tenant Security Settings',
            'TENANT-SEC-CONF',
            old_instance=old_obj,
            new_instance=obj,
        )
        messages.success(request, 'Tenant security settings saved successfully.')
        return redirect(reverse('tenant_security_settings'))


class ActiveSessionListView(LoginRequiredMixin, View):
    template_name = 'security/sessions/session_list.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        domain_filter = request.GET.get('user_domain', 'All')
        search_query = request.GET.get('q', '').strip()

        # TODO Phase 11 Redis: Replace this DB query with
        # Redis scan for live JWT sessions when Redis
        # is implemented.
        qs = ActiveSession.objects.filter(is_active=True)
        if domain_filter in ['Admin', 'Tenant_User', 'Driver']:
            qs = qs.filter(user_domain=domain_filter)
        if search_query:
            qs = qs.filter(
                Q(reference_name__icontains=search_query)
                | Q(ip_address__icontains=search_query)
            )

        qs = qs.order_by('-started_at')
        paginator = Paginator(qs, 5)
        page_number = request.GET.get('page', 1)
        sessions_page = paginator.get_page(page_number)

        total_active = ActiveSession.objects.filter(is_active=True).count()
        total_admin = ActiveSession.objects.filter(
            is_active=True,
            user_domain='Admin',
        ).count()
        total_tenant = ActiveSession.objects.filter(
            is_active=True,
            user_domain='Tenant_User',
        ).count()
        total_driver = ActiveSession.objects.filter(
            is_active=True,
            user_domain='Driver',
        ).count()

        context = {
            'sessions': sessions_page,
            'domain_filter': domain_filter,
            'search_query': search_query,
            'total_active': total_active,
            'total_admin': total_admin,
            'total_tenant': total_tenant,
            'total_driver': total_driver,
        }
        return render(request, self.template_name, context)


from superadmin.redis_helpers import (
    get_all_active_admin_sessions,
    revoke_admin_session,
    get_all_active_tenant_sessions,
    revoke_tenant_session_by_jti,
)


class ActiveSessionsView(LoginRequiredMixin, View):
    def get(self, request):
        if not request.user.is_root:
            messages.error(request, 'Access denied.')
            return redirect('dashboard')

        admin_sessions = get_all_active_admin_sessions()
        tenant_sessions = get_all_active_tenant_sessions()
        sessions = list(admin_sessions)
        for t in tenant_sessions:
            display_name = (t.get('reference_name') or '').strip()
            name_parts = [p for p in display_name.split(' ') if p]
            first_name = name_parts[0] if name_parts else (t.get('user_domain') or 'Tenant')
            last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
            user_domain = (t.get('user_domain') or 'Tenant_User').strip()
            role_label = 'Tenant Admin' if user_domain == 'Tenant_User' else user_domain.replace('_', ' ')
            sessions.append({
                'jti': t.get('jti'),
                'admin_id': '',
                'email': '',
                'first_name': first_name,
                'last_name': last_name,
                'role': role_label,
                'ip_address': t.get('ip_address', ''),
                'user_agent': t.get('user_agent', ''),
                'user_domain': user_domain,
                'started_at': t.get('started_at', ''),
                'last_activity': t.get('last_activity', ''),
                'ttl_seconds': t.get('ttl_seconds', 0),
            })
        current_jti = request.session.get('jti')

        search_query = request.GET.get('q', '').strip()
        role_filter = request.GET.get('role', 'All').strip()
        scope_filter = request.GET.get('scope', 'All').strip()
        sort_by = request.GET.get('sort', 'rank').strip()
        sort_dir = request.GET.get('dir', 'desc').strip().lower()
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'desc'

        if search_query:
            q = search_query.lower()
            sessions = [
                s for s in sessions
                if q in (s.get('email', '') or '').lower()
                or q in (s.get('first_name', '') or '').lower()
                or q in (s.get('last_name', '') or '').lower()
                or q in (s.get('ip_address', '') or '').lower()
                or q in (s.get('user_domain', '') or '').lower()
            ]

        all_roles = sorted(
            {
                (s.get('role', '') or '').strip()
                for s in sessions
                if (s.get('role', '') or '').strip()
            }
        )
        if role_filter != 'All':
            sessions = [s for s in sessions if (s.get('role', '') or '') == role_filter]

        if scope_filter == 'Current':
            sessions = [s for s in sessions if s.get('jti') == current_jti]
        elif scope_filter == 'Others':
            sessions = [s for s in sessions if s.get('jti') != current_jti]

        def _dt(value):
            if not value:
                return datetime.min
            try:
                return datetime.fromisoformat(str(value))
            except Exception:
                return datetime.min

        sort_mapping = {
            'rank': lambda s: (
                s.get('jti') == current_jti,
                _dt(s.get('started_at')),
            ),
            'user': lambda s: (
                f"{s.get('first_name', '')} {s.get('last_name', '')}".strip().lower(),
                str(s.get('email', '')).lower(),
            ),
            'role': lambda s: str(s.get('role', '')).lower(),
            'ip': lambda s: str(s.get('ip_address', '')).lower(),
            'device': lambda s: str(s.get('user_agent', '')).lower(),
            'started_at': lambda s: _dt(s.get('started_at')),
            'last_activity': lambda s: _dt(s.get('last_activity')),
            'ttl': lambda s: int(s.get('ttl_seconds') or 0),
        }
        sort_key = sort_mapping.get(sort_by, sort_mapping['rank'])
        reverse = sort_dir == 'desc'
        sessions.sort(key=sort_key, reverse=reverse)

        paginator = Paginator(sessions, 10)
        page_number = request.GET.get('page', 1)
        sessions_page = paginator.get_page(page_number)
        total_count = len(sessions)
        start_index = sessions_page.start_index()
        for offset, session in enumerate(sessions_page.object_list):
            session['list_rank'] = total_count - (start_index + offset) + 1

        context = {
            'sessions': sessions_page,
            'current_jti': current_jti,
            'total_count': total_count,
            'search_query': search_query,
            'role_filter': role_filter,
            'scope_filter': scope_filter,
            'roles': all_roles,
            'current_sort': sort_by,
            'current_dir': sort_dir,
        }

        if request.GET.get('partial') == '1':
            return render(request, 'security/_active_sessions_fragments.html', context)

        return render(request, 'security/active_sessions.html', context)


class RevokeSessionView(LoginRequiredMixin, View):
    def post(self, request, jti):
        if not request.user.is_root:
            messages.error(request, 'Access denied.')
            return redirect('dashboard')

        current_jti = request.session.get('jti')

        if jti == current_jti:
            messages.error(
                request,
                'You cannot revoke your own active session.',
            )
            return redirect('active_sessions')

        revoke_admin_session(jti)
        revoke_tenant_session_by_jti(jti)

        from superadmin.auth_helpers import log_access

        log_access('Logout', 'Success', request.user.email, request.META.get('REMOTE_ADDR'))

        messages.success(
            request,
            f'Session {jti[:8]}... has been revoked.',
        )
        return redirect('active_sessions')


class SessionRevokeView(LoginRequiredMixin, View):
    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def post(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        session = get_object_or_404(ActiveSession, session_id=pk)
        old_obj = ActiveSession.objects.get(session_id=pk)
        if not session.is_active:
            messages.info(request, 'Session is already revoked.')
            return redirect(reverse('active_sessions'))

        session.is_active = False
        session.revoked_at = timezone.now()
        session.revoked_by = request.user
        session.save()

        log_audit_action(
            request,
            'Status_Change',
            'Active Sessions',
            str(session.session_id),
            old_instance=old_obj,
            new_instance=session,
        )
        from superadmin.redis_helpers import (
            revoke_admin_session,
            revoke_tenant_session_key,
        )

        rj = (session.redis_jti or str(session.session_id)).strip()
        if session.user_domain == 'Admin':
            revoke_admin_session(rj)
        elif session.user_domain in ('Tenant_User', 'Driver'):
            tid = str(session.tenant_id) if session.tenant_id else ''
            if tid:
                revoke_tenant_session_key(tid, rj)
        messages.success(request, 'Session revoked successfully.')
        return redirect(reverse('active_sessions'))


class MassRevokeView(LoginRequiredMixin, View):
    template_name = 'security/sessions/mass_revoke.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        selected_tenant_id = request.GET.get('tenant_id', '').strip()
        tenants = TenantProfile.objects.filter(
            account_status='Active'
        ).order_by('company_name')
        return render(
            request,
            self.template_name,
            {
                'tenants': tenants,
                'selected_tenant_id': selected_tenant_id,
            },
        )

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return JsonResponse({'error': 'Forbidden'}, status=403)

        tenant_id = request.POST.get('tenant_id', '').strip()
        if not tenant_id:
            return JsonResponse({'error': 'tenant_id is required'}, status=400)

        tenant = TenantProfile.objects.filter(tenant_id=tenant_id).first()
        if tenant is None:
            return JsonResponse({'error': 'Tenant not found'}, status=404)

        from django.db.models import Q

        active_qs = ActiveSession.objects.filter(
            is_active=True,
            user_domain__in=['Tenant_User', 'Driver'],
        ).filter(
            Q(tenant=tenant)
            | Q(
                tenant__isnull=True,
                reference_id=str(tenant.tenant_id),
            ),
        )
        revoked_count = active_qs.count()
        active_qs.update(
            is_active=False,
            revoked_at=timezone.now(),
            revoked_by=request.user,
        )

        log_audit_action(
            request,
            'Status_Change',
            'Active Sessions - Mass Revoke',
            str(tenant_id),
            new_instance=None,
        )
        from superadmin.redis_helpers import revoke_all_tenant_sessions

        redis_revoked_count = revoke_all_tenant_sessions(str(tenant.tenant_id))
        # Redis is the source of truth for active tenant JWT sessions.
        # Fall back to DB-marked count only when Redis deleted nothing.
        total_revoked_count = int(redis_revoked_count or 0) or int(revoked_count or 0)
        return JsonResponse(
            {'message': 'Mass revoke executed.', 'revoked_count': total_revoked_count},
            status=200,
        )


class RoleListView(LoginRequiredMixin, View):
    template_name = 'system_users/roles/role_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort_by = request.GET.get('sort', 'role_name_en')
        sort_dir = request.GET.get('dir', 'asc')

        # Whitelist sortable fields
        sortable_fields = {
            'rank': 'default_rank',
            'role_name_en': 'role_name_en',
            'role_name_ar': 'role_name_ar',
            'description': 'description',
            'status': 'status',
            'is_system_default': 'is_system_default'
        }
        db_field = sortable_fields.get(sort_by, 'role_name_en')
        ordering = db_field if sort_dir == 'asc' else '-' + db_field

        roles_qs = Role.objects.all()
        
        if search_query:
            roles_qs = roles_qs.filter(
                Q(role_name_en__icontains=search_query) |
                Q(role_name_ar__icontains=search_query)
            )
            
        if status_filter in ['Active', 'Inactive']:
            roles_qs = roles_qs.filter(status=status_filter)

        # Annotate with stable rank based on default order (role_name_en)
        # We do this after filtering but before final sorting
        roles_qs = roles_qs.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('role_name_en').asc()
            )
        )

        roles_qs = roles_qs.order_by(ordering)

        total_count = roles_qs.count()
        paginator = Paginator(roles_qs, 10)
        page_number = request.GET.get('page', 1)
        roles_page = paginator.get_page(page_number)

        context = {
            'roles': roles_page,
            'search_query': search_query,
            'status_filter': status_filter,
            'current_sort': sort_by,
            'current_dir': sort_dir,
            'total_count': total_count,
        }
        return render(request, self.template_name, context)


class RoleDetailView(LoginRequiredMixin, View):
    template_name = 'system_users/roles/role_detail.html'

    def get(self, request, pk):
        role = get_object_or_404(
            Role.objects.select_related('created_by', 'updated_by'),
            pk=pk,
        )
        return render(
            request,
            self.template_name,
            {'role': role, 'page_title': 'Role details'},
        )


class RoleCreateView(LoginRequiredMixin, View):
    template_name = 'system_users/roles/role_form.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('role_list'))
        return None

    def get(self, request):
        root_redirect = self._require_root(request)
        if root_redirect:
            return root_redirect
        return render(request, self.template_name, {'form': RoleForm(), 'is_edit': False})

    def post(self, request):
        root_redirect = self._require_root(request)
        if root_redirect:
            return root_redirect

        form = RoleForm(request.POST)
        if form.is_valid():
            role = form.save(commit=False)
            role.created_by = request.user
            role.save()
            messages.success(request, 'Role created successfully.')
            return redirect(reverse('role_list'))

        return render(request, self.template_name, {'form': form, 'is_edit': False})


class RoleUpdateView(LoginRequiredMixin, View):
    template_name = 'system_users/roles/role_form.html'

    def _require_root_or_redirect(self, request, redirect_to):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(redirect_to)
        return None

    def get(self, request, pk):
        redirect_resp = self._require_root_or_redirect(request, reverse('role_list'))
        if redirect_resp:
            return redirect_resp

        role = get_object_or_404(Role, pk=pk)
        if role.is_system_default:
            messages.error(request, 'System default roles cannot be modified')
            return redirect(reverse('role_list'))

        return render(request, self.template_name, {'form': RoleForm(instance=role), 'is_edit': True})

    def post(self, request, pk):
        redirect_resp = self._require_root_or_redirect(request, reverse('role_list'))
        if redirect_resp:
            return redirect_resp

        role = get_object_or_404(Role, pk=pk)
        if role.is_system_default:
            messages.error(request, 'System default roles cannot be modified')
            return redirect(reverse('role_list'))

        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            role = form.save(commit=False)
            role.updated_by = request.user
            role.save()
            messages.success(request, 'Role updated successfully.')
            return redirect(reverse('role_list'))

        return render(request, self.template_name, {'form': form, 'is_edit': True})


class RoleToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('role_list'))

        role = get_object_or_404(Role, pk=pk)
        if role.is_system_default:
            messages.error(request, 'System default roles cannot be modified')
            return redirect(reverse('role_list'))

        if role.status == 'Active':
            target_status = 'Inactive'
            active_users = AdminUser.objects.filter(role=role, status='Active')
            if active_users.exists():
                messages.error(
                    request,
                    f'Cannot deactivate role — {active_users.count()} users are currently assigned to it',
                )
                return redirect(reverse('role_list'))
        else:
            target_status = 'Active'

        role.status = target_status
        role.updated_by = request.user
        role.save(update_fields=['status', 'updated_by'])

        messages.success(request, 'Role status updated successfully.')
        return redirect(reverse('role_list'))


class RoleDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('role_list'))

        # Hard delete is forbidden; always redirect with error.
        messages.error(request, 'Roles cannot be deleted. Set to Inactive instead.')
        return redirect(reverse('role_list'))


def _revoke_user_sessions(user):
    """
    Phase 1 session invalidation for suspended users.
    This deletes DB-backed sessions that contain the suspended user id.
    """
    user_id = str(user.pk)
    for session in Session.objects.all():
        try:
            decoded = session.get_decoded()
        except Exception:
            continue
        auth_user_id = decoded.get('_auth_user_id')
        if auth_user_id is not None and str(auth_user_id) == user_id:
            session.delete()


class AdminUserListView(LoginRequiredMixin, View):
    template_name = 'system_users/admin_users/admin_user_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        role_filter = request.GET.get('role', 'All')
        sort_by = request.GET.get('sort', 'updated_at')
        sort_dir = request.GET.get('dir', 'desc')

        users_qs = AdminUser.objects.filter(is_deleted=False).select_related('role', 'created_by', 'updated_by')

        # Sortable fields whitelist and mapping
        sort_mapping = {
            'rank': ['default_rank'],
            'name': ['first_name', 'last_name'],
            'email': ['email'],
            'role': ['role__role_name_en'],
            'status': ['status'],
            'last_login_at': ['last_login_at'],
            'created_at': ['created_at'],
        }

        active_sort_fields = sort_mapping.get(sort_by, ['first_name', 'last_name'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else '-' + f)

        if search_query:
            users_qs = users_qs.filter(
                Q(first_name__icontains=search_query)
                | Q(last_name__icontains=search_query)
                | Q(email__icontains=search_query)
            )

        if status_filter in [choice[0] for choice in AdminUser.STATUS_CHOICES]:
            users_qs = users_qs.filter(status=status_filter)

        if role_filter != 'All' and role_filter:
            users_qs = users_qs.filter(role_id=role_filter)

        # Annotate with stable rank based on default order (-created_at)
        users_qs = users_qs.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc()
            )
        )

        users_qs = users_qs.order_by(*ordering)
        total_count = users_qs.count()

        paginator = Paginator(users_qs, 10)
        page_number = request.GET.get('page', 1)
        users_page = paginator.get_page(page_number)
        start_index = users_page.start_index()
        for offset, user in enumerate(users_page.object_list):
            # Keep displayed rank aligned with newest-first ordering.
            user.list_rank = total_count - (start_index + offset) + 1

        context = {
            'admin_users': users_page,
            'search_query': search_query,
            'status_filter': status_filter,
            'role_filter': role_filter,
            'current_sort': sort_by,
            'current_dir': sort_dir,
            'total_count': total_count,
            'roles': Role.objects.all().order_by('role_name_en'),
            'statuses': AdminUser.STATUS_CHOICES,
            'page_title': 'Admin Users Master',
        }
        return render(request, self.template_name, context)


class AdminUserCreateView(LoginRequiredMixin, View):
    template_name = 'system_users/admin_users/admin_user_form.html'

    @staticmethod
    def _is_root_role(role):
        role_name = (getattr(role, 'role_name_en', '') or '').strip().lower()
        return role_name == 'super admin'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('admin_user_list'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        form = AdminUserForm(initial={'status': 'Pending_Activation'})
        return render(request, self.template_name, {'form': form, 'is_edit': False})

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        form = AdminUserForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form, 'is_edit': False})

        user = form.save(commit=False)
        # Phase 1/2 rule: invite flow (no password yet)
        user.status = 'Pending_Activation'
        user.is_root = self._is_root_role(user.role)
        user.created_by = request.user
        user.updated_by = request.user
        user.save()

        auth_token = create_auth_token(user, 'invite')
        invite_url = request.build_absolute_uri(
            reverse('set_password', args=[auth_token.token])
        )

        sent = send_auth_email(
            user,
            'invite',
            {'admin_user': user, 'invite_url': invite_url},
        )
        if sent:
            messages.success(
                request,
                f'Admin user created. Invitation email sent to {user.email}.',
            )
        else:
            messages.warning(
                request,
                f'Admin user created, but invite email failed for {user.email}.',
            )
        return redirect(reverse('admin_user_list'))


class MyAccountView(LoginRequiredMixin, View):
    template_name = 'system_users/my_account.html'

    def get(self, request):
        form = MyAccountForm(instance=request.user)
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'target_user': request.user,
            },
        )

    def post(self, request):
        form = MyAccountForm(request.POST, instance=request.user)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'target_user': request.user,
                },
            )

        old_payload = {
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'phone_number': request.user.phone_number,
            'email': request.user.email,
        }
        
        updated = form.save(commit=False)
        updated.updated_by = request.user
        
        update_fields = [
            'first_name',
            'last_name',
            'phone_number',
            'updated_by',
            'updated_at',
        ]
        if getattr(request.user, 'is_root', False):
            update_fields.append('email')

        new_password = form.cleaned_data.get('new_password')
        if new_password:
            updated.set_password(new_password)
            update_fields.append('password')
            old_payload['_security_event'] = 'Password Changed'

        updated.save(update_fields=update_fields)

        log_audit_action(
            request=request,
            action_type='Update',
            module_name='My Account',
            record_id=str(getattr(request.user, 'id', '')),
            old_payload=old_payload,
            new_instance=updated,
        )

        if new_password:
            # Re-authenticate the user so they don't get logged out after password change
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, updated)
            messages.success(request, 'Your account and password were updated successfully.')
        else:
            messages.success(request, 'Your account details were updated successfully.')

        return redirect('my_account')


class SetPasswordView(View):
    """Public: activate invited admin/tenant via token."""

    template_form = 'auth/set_password.html'
    template_error = 'auth/token_error.html'

    def _render_error(self, request, message):
        return render(
            request,
            self.template_error,
            {'error_message': message},
        )

    def _get_invite_token(self, raw_token):
        try:
            return AdminAuthToken.objects.select_related('admin_user').get(
                token=raw_token,
                token_type=AdminAuthToken.TokenType.INVITE,
            )
        except AdminAuthToken.DoesNotExist:
            return None

    def _get_tenant_invite_token(self, raw_token):
        try:
            return TenantAuthToken.objects.select_related('tenant_profile').get(
                token=raw_token,
                token_type=TenantAuthToken.TokenType.INVITE,
            )
        except TenantAuthToken.DoesNotExist:
            return None

    def get(self, request, token):
        invite = self._get_invite_token(token)
        tenant_invite = self._get_tenant_invite_token(token) if invite is None else None
        active_invite = invite or tenant_invite
        if active_invite is None:
            return self._render_error(request, 'Invalid invite link.')
        if active_invite.is_used:
            return self._render_error(request, 'This invite link has already been used.')
        if active_invite.is_expired:
            return self._render_error(request, 'This invite link has expired.')
        invite_email = (
            invite.admin_user.email if invite is not None else tenant_invite.tenant_profile.primary_email
        )
        return render(
            request,
            self.template_form,
            {
                'form': SetPasswordForm(),
                'invite_email': invite_email,
            },
        )

    def post(self, request, token):
        invite = self._get_invite_token(token)
        tenant_invite = self._get_tenant_invite_token(token) if invite is None else None
        active_invite = invite or tenant_invite
        if active_invite is None:
            return self._render_error(request, 'Invalid invite link.')
        if active_invite.is_used:
            return self._render_error(request, 'This invite link has already been used.')
        if active_invite.is_expired:
            return self._render_error(request, 'This invite link has expired.')

        form = SetPasswordForm(request.POST)
        if not form.is_valid():
            invite_email = (
                invite.admin_user.email if invite is not None else tenant_invite.tenant_profile.primary_email
            )
            return render(
                request,
                self.template_form,
                {
                    'form': form,
                    'invite_email': invite_email,
                },
            )

        password = form.cleaned_data['password']
        ip = _client_ip(request)
        if invite is not None:
            user = invite.admin_user
            user.set_password(password)
            user.status = 'Active'
            user.two_factor_enabled = True
            # Keep root marker aligned to current role during activation.
            role_name = (getattr(user.role, 'role_name_en', '') or '').strip().lower()
            user.is_root = role_name == 'super admin'
            user.save(update_fields=['password', 'status', 'two_factor_enabled', 'is_root'])
            invite.is_used = True
            invite.save(update_fields=['is_used'])
            log_access('Login', 'Success', user.email, ip)
        else:
            tenant = tenant_invite.tenant_profile
            tenant.portal_bootstrap_password_hash = make_password(password)
            tenant.save(update_fields=['portal_bootstrap_password_hash'])
            tenant_invite.is_used = True
            tenant_invite.save(update_fields=['is_used'])
            reset_failed_attempts(tenant.primary_email)
            log_access('Login', 'Success', tenant.primary_email, ip)

        messages.success(
            request,
            'Password set successfully. Please login.',
        )

        return redirect(reverse('login'))


class AdminUserResendInviteView(LoginRequiredMixin, View):
    """Resend the activation link to a Pending user."""

    def _require_root(self, request):
        if not getattr(request.user, "is_root", False):
            messages.error(request, "Access denied: root admin only.")
            return redirect(reverse("admin_user_list"))
        return None

    def post(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        user = get_object_or_404(AdminUser, pk=pk)

        if user.status != "Pending_Activation":
            messages.error(
                request,
                f"User {user.email} is already active or suspended.",
            )
            return redirect(reverse("admin_user_list"))

        auth_token = create_auth_token(user, "invite")
        invite_url = request.build_absolute_uri(
            reverse("set_password", args=[auth_token.token])
        )

        sent = send_auth_email(
            user, "invite", {"admin_user": user, "invite_url": invite_url}
        )

        if sent:
            messages.success(
                request,
                f"Invitation email resent successfully to {user.email}.",
            )
        else:
            messages.error(
                request,
                f"Failed to send invitation email to {user.email}. Check system logs.",
            )

        return redirect(reverse("admin_user_list"))


class AdminUserUpdateView(LoginRequiredMixin, View):
    template_name = 'system_users/admin_users/admin_user_form.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('admin_user_list'))
        return None

    def get(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        target_user = get_object_or_404(
            AdminUser.objects.select_related('role', 'created_by', 'updated_by'),
            pk=pk,
        )
        form = AdminUserForm(instance=target_user)
        return render(request, self.template_name, {'form': form, 'is_edit': True})

    def post(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        target_user = get_object_or_404(
            AdminUser.objects.select_related('role', 'created_by', 'updated_by'),
            pk=pk,
        )
        original_role = target_user.role
        original_status = target_user.status

        form = AdminUserForm(request.POST, instance=target_user)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form, 'is_edit': True})

        # Block root admin role changes.
        new_role = form.cleaned_data.get('role')
        # NOTE: ModelForm mutates the instance during validation, so we must
        # compare against the original role captured before form.is_valid().
        if target_user.is_root and new_role != original_role:
            messages.error(request, 'Root admin role can NEVER be changed')
            return redirect(reverse('admin_user_edit', args=[pk]))

        # Extra safety: root admin cannot be suspended.
        if target_user.is_root and form.cleaned_data.get('status') == 'Suspended':
            messages.error(request, 'Root admin cannot be suspended')
            return redirect(reverse('admin_user_edit', args=[pk]))

        user = form.save(commit=False)
        user.updated_by = request.user
        
        status_changed = (user.status == 'Suspended' and original_status != 'Suspended')
        user.save()

        if status_changed:
            from superadmin.redis_helpers import revoke_all_sessions_for_admin
            revoke_all_sessions_for_admin(user.id)
            _revoke_user_sessions(user)
            messages.info(request, f'Admin sessions for {user.email} have been revoked due to account suspension.')

        messages.success(request, 'Admin user updated successfully.')
        return redirect(reverse('admin_user_list'))


class AdminUserDetailView(LoginRequiredMixin, View):
    template_name = 'system_users/admin_users/admin_user_detail.html'

    def get(self, request, pk):
        target_user = get_object_or_404(
            AdminUser.objects.select_related('role', 'created_by', 'updated_by'),
            pk=pk,
        )
        return render(request, self.template_name, {'target_user': target_user, 'page_title': 'Admin User Details'})


class AdminUserToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('admin_user_list'))

        target_user = get_object_or_404(AdminUser, pk=pk)

        if target_user.is_root:
            messages.error(request, 'Root admin cannot be suspended')
            return redirect(reverse('admin_user_list'))

        # If target is currently active/pending => suspend, else => activate.
        is_suspending = target_user.status != 'Suspended'
        if is_suspending and target_user.pk == request.user.pk:
            messages.error(request, 'You cannot suspend your own account')
            return redirect(reverse('admin_user_list'))

        if is_suspending:
            # Phase 1: suspend + invalidate active sessions
            target_user.status = 'Suspended'
            target_user.updated_by = request.user
            target_user.save(update_fields=['status', 'updated_by'])

            from superadmin.redis_helpers import revoke_all_sessions_for_admin

            revoke_all_sessions_for_admin(target_user.id)
            # Kill Switch: all Redis sessions for this admin destroyed.
            # (Tenant suspend uses ``revoke_all_tenant_sessions`` in TenantUpdateView.)
            _revoke_user_sessions(target_user)
            messages.success(request, 'Admin user suspended successfully.')
        else:
            target_user.status = 'Active'
            target_user.two_factor_enabled = True
            target_user.updated_by = request.user
            target_user.save(update_fields=['status', 'two_factor_enabled', 'updated_by'])
            messages.success(request, 'Admin user activated successfully.')

        return redirect(reverse('admin_user_list'))


class AdminUserDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('admin_user_list'))

        target_user = get_object_or_404(AdminUser, pk=pk)
        if target_user.is_root:
            messages.error(request, 'Root admin cannot be deleted')
            return redirect(reverse('admin_user_list'))
        if target_user.pk == request.user.pk:
            messages.error(request, 'You cannot delete your own account')
            return redirect(reverse('admin_user_list'))

        if not target_user.is_deleted:
            target_user.is_deleted = True
            target_user.updated_by = request.user
            target_user.save(update_fields=['is_deleted', 'updated_by'])
            from superadmin.redis_helpers import revoke_all_sessions_for_admin

            revoke_all_sessions_for_admin(target_user.id)
            _revoke_user_sessions(target_user)
            messages.success(request, 'Admin user removed successfully.')
        else:
            messages.info(request, 'Admin user is already removed.')

        return redirect(reverse('admin_user_list'))

    def get(self, request, pk):
        return self.post(request, pk)


class SystemUsersAnalyticsView(LoginRequiredMixin, View):
    template_name = 'system_users/analytics/users_analytics.html'

    def get(self, request):
        from django.utils.translation import gettext as _
        from django.utils import translation
        from datetime import timedelta

        is_ar = translation.get_language() == 'ar'

        total_staff = AdminUser.objects.exclude(status='Suspended').count()
        suspended_count = AdminUser.objects.filter(status='Suspended').count()
        pending_count = AdminUser.objects.filter(status='Pending_Activation').count()

        total_active = AdminUser.objects.filter(status='Active').count()
        two_fa_enabled_count = AdminUser.objects.filter(
            two_factor_enabled=True,
            status='Active',
        ).count()
        two_fa_rate = (two_fa_enabled_count / total_active * 100) if total_active > 0 else 0

        stale_threshold = timezone.now() - timedelta(days=30)
        stale_accounts_qs = (
            AdminUser.objects.filter(last_login_at__lt=stale_threshold, status='Active')
            .select_related('role')
            .order_by('-last_login_at')
        )

        stale_accounts = []
        now = timezone.now()
        for u in stale_accounts_qs:
            days_since = (now.date() - u.last_login_at.date()).days if u.last_login_at else None
            role_display = None
            if u.role:
                role_display = u.role.role_name_ar if is_ar else u.role.role_name_en
            
            stale_accounts.append(
                {
                    'name': f'{u.first_name} {u.last_name}',
                    'email': u.email,
                    'role': role_display,
                    'last_login_at': u.last_login_at,
                    'days_since_login': days_since,
                }
            )

        # Role distribution: Active users per role (include Unassigned if role is null)
        role_distribution = []
        active_users_by_role = (
            AdminUser.objects.filter(status='Active')
            .values('role')
            .annotate(count=Count('id'))
        )

        counts_by_role_id = {row['role']: row['count'] for row in active_users_by_role}
        for role in Role.objects.all().order_by('role_name_en'):
            role_name = role.role_name_ar if is_ar else role.role_name_en
            role_distribution.append(
                {
                    'role_name': role_name,
                    'count': counts_by_role_id.get(role.pk, 0),
                }
            )
        if None in counts_by_role_id:
            role_distribution.append({'role_name': _('Unassigned'), 'count': counts_by_role_id.get(None, 0)})

        recently_created = []
        for u in AdminUser.objects.select_related('role').order_by('-created_at')[:5]:
            # Add a dynamic attribute for the role name to be used in template
            if u.role:
                u.display_role = u.role.role_name_ar if is_ar else u.role.role_name_en
            else:
                u.display_role = '-'
            recently_created.append(u)

        context = {
            'total_staff': total_staff,
            'suspended_count': suspended_count,
            'pending_count': pending_count,
            'two_fa_enabled_count': two_fa_enabled_count,
            'two_fa_rate': round(two_fa_rate, 2),
            'stale_accounts': stale_accounts,
            'role_distribution': role_distribution,
            'recently_created': recently_created,
            'page_title': _('System Users Analytics'),
        }
        return render(request, self.template_name, context)


class CountryListView(LoginRequiredMixin, View):
    template_name = 'master_data/countries/country_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'desc')

        countries_qs = Country.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('name_en').asc(),
            ),
            status_sort=Case(
                When(is_active=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
        )

        if search_query:
            countries_qs = countries_qs.filter(
                Q(name_en__icontains=search_query)
                | Q(country_code__icontains=search_query)
                | Q(name_ar__icontains=search_query)
            )

        if status_filter == 'Active':
            countries_qs = countries_qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            countries_qs = countries_qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'code': 'country_code',
            'name_en': 'name_en',
            'name_ar': 'name_ar',
            'status': 'status_sort',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            countries_qs = countries_qs.order_by(
                F(order_by_field).desc(nulls_last=True),
                '-default_rank',
            )
        else:
            countries_qs = countries_qs.order_by(
                F(order_by_field).asc(nulls_first=True),
                'default_rank',
            )
        total_count = countries_qs.count()

        paginator = Paginator(countries_qs, 10)
        page_number = request.GET.get('page', 1)
        countries_page = paginator.get_page(page_number)

        context = {
            'countries': countries_page,
            'search_query': search_query,
            'status_filter': status_filter,
            'current_sort': sort,
            'current_dir': direction,
            'total_count': total_count,
            'page_title': 'Countries Master',
        }
        return render(request, self.template_name, context)


class CountryDetailView(LoginRequiredMixin, View):
    template_name = 'master_data/countries/country_detail.html'

    def get(self, request, pk):
        country = get_object_or_404(Country, pk=pk)
        return render(
            request,
            self.template_name,
            {'country': country, 'page_title': 'Country Details'},
        )


class CountryCreateView(LoginRequiredMixin, View):
    template_name = 'master_data/countries/country_form.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('country_list'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        form = CountryForm(is_edit=False)
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': False,
            },
        )

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        form = CountryForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': False,
                },
            )

        country = form.save(commit=False)
        country.country_code = country.country_code.upper().strip()
        country.created_by = request.user
        country.save()
        log_audit_action(
            request,
            'Create',
            'Countries Master',
            str(country.country_code),
            new_instance=country,
        )

        messages.success(request, 'Country created successfully.')
        # TODO Phase 10: Invalidate country cache here
        return redirect(reverse('country_list'))


class CountryUpdateView(LoginRequiredMixin, View):
    template_name = 'master_data/countries/country_form.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('country_list'))
        return None

    def get(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        country = get_object_or_404(Country, pk=pk)
        form = CountryForm(instance=country, is_edit=True)
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': True,
                'country': country,
            },
        )

    def post(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        country = get_object_or_404(Country, pk=pk)
        old_obj = Country.objects.get(country_code=country.country_code)
        form = CountryForm(request.POST, instance=country, is_edit=True)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': True,
                    'country': country,
                },
            )

        form.save()
        log_audit_action(
            request,
            'Update',
            'Countries Master',
            str(country.country_code),
            old_instance=old_obj,
            new_instance=country,
        )
        messages.success(request, 'Country updated successfully.')
        # TODO Phase 10: Invalidate country cache here
        return redirect(reverse('country_list'))


class CountryToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('country_list'))

        country = get_object_or_404(Country, pk=pk)
        old_obj = Country.objects.get(country_code=country.country_code)

        if country.is_active:
            # TODO Phase 5: Check if country is linked to active Tenants
            #               before deactivating — implement when Tenant
            #               model exists
            country.is_active = False
            messages.success(request, 'Country deactivated successfully.')
        else:
            country.is_active = True
            messages.success(request, 'Country activated successfully.')

        country.save(update_fields=['is_active'])
        log_audit_action(
            request,
            'Status_Change',
            'Countries Master',
            str(country.country_code),
            old_instance=old_obj,
            new_instance=country,
        )
        # TODO Phase 10: Invalidate country cache here
        return redirect(reverse('country_list'))


class CountryDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        messages.error(request, 'Countries cannot be deleted. Deactivate instead.')
        return redirect(reverse('country_list'))

    def get(self, request, pk):
        messages.error(request, 'Countries cannot be deleted. Deactivate instead.')
        return redirect(reverse('country_list'))


class CountryImportExcelView(LoginRequiredMixin, View):
    def post(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('country_list'))

        if load_workbook is None:
            messages.error(
                request,
                'Excel import dependency is unavailable. Please install openpyxl.',
            )
            return redirect(reverse('country_list'))

        upload = request.FILES.get('excel_file')
        if not upload:
            messages.error(request, 'Please choose an Excel file to import.')
            return redirect(reverse('country_list'))

        filename = (upload.name or '').lower()
        if not (
            filename.endswith('.xlsx')
            or filename.endswith('.xlsm')
            or filename.endswith('.xltx')
            or filename.endswith('.xltm')
        ):
            messages.error(request, 'Only Excel .xlsx files are supported.')
            return redirect(reverse('country_list'))

        try:
            wb = load_workbook(upload, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, 'Invalid Excel file. Please use the provided sample.')
            return redirect(reverse('country_list'))

        header_row = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[1]]
        required_headers = {'country_code', 'name_en', 'name_ar'}
        if not required_headers.issubset(set(header_row)):
            messages.error(
                request,
                'Invalid template headers. Required: country_code, name_en, name_ar.',
            )
            return redirect(reverse('country_list'))

        idx = {name: header_row.index(name) for name in required_headers}
        is_active_idx = header_row.index('is_active') if 'is_active' in header_row else None

        created_count = 0
        skipped_count = 0
        row_errors = []

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            country_code = str(row[idx['country_code']] or '').strip().upper()
            name_en = str(row[idx['name_en']] or '').strip()
            name_ar = str(row[idx['name_ar']] or '').strip()

            if not country_code and not name_en and not name_ar:
                continue

            if not country_code or not name_en or not name_ar:
                skipped_count += 1
                row_errors.append(f'Row {row_no}: missing required value(s).')
                continue

            if Country.objects.filter(country_code=country_code).exists():
                skipped_count += 1
                continue

            is_active = True
            if is_active_idx is not None:
                raw_active = row[is_active_idx]
                if raw_active is not None and str(raw_active).strip() != '':
                    norm = str(raw_active).strip().lower()
                    is_active = norm in ('1', 'true', 'yes', 'y', 'active')

            country = Country(
                country_code=country_code,
                name_en=name_en,
                name_ar=name_ar,
                is_active=is_active,
                created_by=request.user,
            )
            try:
                country.full_clean()
                country.save()
                log_audit_action(
                    request,
                    'Create',
                    'Countries Master',
                    str(country.country_code),
                    new_instance=country,
                )
                created_count += 1
            except Exception:
                skipped_count += 1
                row_errors.append(f'Row {row_no}: invalid data or duplicate.')

        if created_count:
            messages.success(
                request,
                f'Country import completed. Created {created_count}, skipped {skipped_count}.',
            )
        else:
            messages.error(
                request,
                f'No countries imported. Skipped {skipped_count}.',
            )
        if row_errors:
            messages.warning(request, ' | '.join(row_errors[:3]))
        return redirect(reverse('country_list'))


class CountrySampleExcelView(LoginRequiredMixin, View):
    def get(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('country_list'))

        if Workbook is None:
            messages.error(
                request,
                'Excel sample dependency is unavailable. Please install openpyxl.',
            )
            return redirect(reverse('country_list'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Countries'
        ws.append(['country_code', 'name_en', 'name_ar', 'is_active'])
        ws.append(['SA', 'Saudi Arabia', 'المملكة العربية السعودية', 'TRUE'])
        ws.append(['AE', 'United Arab Emirates', 'الإمارات العربية المتحدة', 'TRUE'])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="country_import_sample.xlsx"'
        return response


class CurrencyListView(LoginRequiredMixin, View):
    template_name = 'master_data/currencies/currency_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'desc')

        currencies_qs = Currency.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('name_en').asc(),
            ),
            status_sort=Case(
                When(is_active=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
        )

        if search_query:
            currencies_qs = currencies_qs.filter(
                Q(name_en__icontains=search_query)
                | Q(currency_code__icontains=search_query)
                | Q(name_ar__icontains=search_query)
            )

        if status_filter == 'Active':
            currencies_qs = currencies_qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            currencies_qs = currencies_qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'code': 'currency_code',
            'symbol': 'currency_symbol',
            'name_en': 'name_en',
            'name_ar': 'name_ar',
            'decimal': 'decimal_places',
            'status': 'status_sort',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            currencies_qs = currencies_qs.order_by(
                F(order_by_field).desc(nulls_last=True),
                '-default_rank',
            )
        else:
            currencies_qs = currencies_qs.order_by(
                F(order_by_field).asc(nulls_first=True),
                'default_rank',
            )
        total_count = currencies_qs.count()

        paginator = Paginator(currencies_qs, 10)
        page_number = request.GET.get('page', 1)
        currencies_page = paginator.get_page(page_number)
        start_index = currencies_page.start_index()
        for offset, currency in enumerate(currencies_page.object_list):
            # Show descending list ID so top rows have higher numbers.
            currency.list_rank = total_count - (start_index + offset) + 1

        context = {
            'currencies': currencies_page,
            'search_query': search_query,
            'status_filter': status_filter,
            'current_sort': sort,
            'current_dir': direction,
            'total_count': total_count,
            'page_title': 'Currencies Master',
        }
        return render(request, self.template_name, context)


class CurrencyDetailView(LoginRequiredMixin, View):
    template_name = 'master_data/currencies/currency_detail.html'

    def get(self, request, pk):
        currency = get_object_or_404(Currency, pk=pk)
        return render(
            request,
            self.template_name,
            {'currency': currency, 'page_title': 'Currency Details'},
        )


class CurrencyCreateView(LoginRequiredMixin, View):
    template_name = 'master_data/currencies/currency_form.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('currency_list'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        form = CurrencyForm(is_edit=False)
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': False,
            },
        )

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        form = CurrencyForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': False,
                },
            )

        currency = form.save(commit=False)
        currency.currency_code = currency.currency_code.upper().strip()
        currency.created_by = request.user
        currency.save()
        log_audit_action(
            request,
            'Create',
            'Currencies Master',
            str(currency.currency_code),
            new_instance=currency,
        )

        messages.success(request, 'Currency created successfully.')
        return redirect(reverse('currency_list'))


class CurrencyUpdateView(LoginRequiredMixin, View):
    template_name = 'master_data/currencies/currency_form.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('currency_list'))
        return None

    def get(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        currency = get_object_or_404(Currency, pk=pk)
        form = CurrencyForm(instance=currency, is_edit=True)
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': True,
                'currency': currency,
            },
        )

    def post(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        currency = get_object_or_404(Currency, pk=pk)
        old_obj = Currency.objects.get(currency_code=currency.currency_code)
        form = CurrencyForm(request.POST, instance=currency, is_edit=True)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': True,
                    'currency': currency,
                },
            )

        # CRITICAL: Enforce immutable PK even if a client bypasses disabled field.
        form.instance.currency_code = currency.currency_code
        form.save()
        log_audit_action(
            request,
            'Update',
            'Currencies Master',
            str(currency.currency_code),
            old_instance=old_obj,
            new_instance=currency,
        )

        messages.success(request, 'Currency updated successfully.')
        return redirect(reverse('currency_list'))


class CurrencyToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('currency_list'))

        currency = get_object_or_404(Currency, pk=pk)
        old_obj = Currency.objects.get(currency_code=currency.currency_code)

        if currency.is_active:
            # TODO Phase 6: Check if currency is linked to active
            #               Subscription Plans or Payment Methods
            #               before deactivating
            currency.is_active = False
            messages.success(request, 'Currency deactivated successfully.')
        else:
            currency.is_active = True
            messages.success(request, 'Currency activated successfully.')

        currency.save(update_fields=['is_active'])
        log_audit_action(
            request,
            'Status_Change',
            'Currencies Master',
            str(currency.currency_code),
            old_instance=old_obj,
            new_instance=currency,
        )
        return redirect(reverse('currency_list'))


class CurrencyDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        messages.error(request, 'Currencies cannot be deleted. Deactivate instead.')
        return redirect(reverse('currency_list'))

    def get(self, request, pk):
        messages.error(request, 'Currencies cannot be deleted. Deactivate instead.')
        return redirect(reverse('currency_list'))


class CurrencyImportExcelView(LoginRequiredMixin, View):
    def post(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('currency_list'))

        if load_workbook is None:
            messages.error(
                request,
                'Excel import dependency is unavailable. Please install openpyxl.',
            )
            return redirect(reverse('currency_list'))

        upload = request.FILES.get('excel_file')
        if not upload:
            messages.error(request, 'Please choose an Excel file to import.')
            return redirect(reverse('currency_list'))

        filename = (upload.name or '').lower()
        if not (
            filename.endswith('.xlsx')
            or filename.endswith('.xlsm')
            or filename.endswith('.xltx')
            or filename.endswith('.xltm')
        ):
            messages.error(request, 'Only Excel .xlsx files are supported.')
            return redirect(reverse('currency_list'))

        try:
            wb = load_workbook(upload, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, 'Invalid Excel file. Please use the provided sample.')
            return redirect(reverse('currency_list'))

        header_row = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[1]]
        required_headers = {'currency_code', 'name_en', 'name_ar', 'currency_symbol'}
        if not required_headers.issubset(set(header_row)):
            messages.error(
                request,
                'Invalid template headers. Required: currency_code, name_en, name_ar, currency_symbol.',
            )
            return redirect(reverse('currency_list'))

        idx = {name: header_row.index(name) for name in required_headers}
        decimal_idx = header_row.index('decimal_places') if 'decimal_places' in header_row else None
        is_active_idx = header_row.index('is_active') if 'is_active' in header_row else None

        created_count = 0
        skipped_count = 0
        row_errors = []

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            currency_code = str(row[idx['currency_code']] or '').strip().upper()
            name_en = str(row[idx['name_en']] or '').strip()
            name_ar = str(row[idx['name_ar']] or '').strip()
            currency_symbol = str(row[idx['currency_symbol']] or '').strip()

            if not currency_code and not name_en and not name_ar and not currency_symbol:
                continue

            if not currency_code or not name_en or not name_ar or not currency_symbol:
                skipped_count += 1
                row_errors.append(f'Row {row_no}: missing required value(s).')
                continue

            if Currency.objects.filter(currency_code=currency_code).exists():
                skipped_count += 1
                continue

            decimal_places = 2
            if decimal_idx is not None and row[decimal_idx] not in (None, ''):
                try:
                    decimal_places = int(row[decimal_idx])
                except Exception:
                    skipped_count += 1
                    row_errors.append(f'Row {row_no}: decimal_places must be an integer.')
                    continue

            is_active = True
            if is_active_idx is not None:
                raw_active = row[is_active_idx]
                if raw_active is not None and str(raw_active).strip() != '':
                    norm = str(raw_active).strip().lower()
                    is_active = norm in ('1', 'true', 'yes', 'y', 'active')

            currency = Currency(
                currency_code=currency_code,
                name_en=name_en,
                name_ar=name_ar,
                currency_symbol=currency_symbol,
                decimal_places=decimal_places,
                is_active=is_active,
                created_by=request.user,
            )
            try:
                currency.full_clean()
                currency.save()
                log_audit_action(
                    request,
                    'Create',
                    'Currencies Master',
                    str(currency.currency_code),
                    new_instance=currency,
                )
                created_count += 1
            except Exception:
                skipped_count += 1
                row_errors.append(f'Row {row_no}: invalid data or duplicate.')

        if created_count:
            messages.success(
                request,
                f'Currency import completed. Created {created_count}, skipped {skipped_count}.',
            )
        else:
            messages.error(
                request,
                f'No currencies imported. Skipped {skipped_count}.',
            )
        if row_errors:
            messages.warning(request, ' | '.join(row_errors[:3]))
        return redirect(reverse('currency_list'))


class CurrencySampleExcelView(LoginRequiredMixin, View):
    def get(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('currency_list'))

        if Workbook is None:
            messages.error(
                request,
                'Excel sample dependency is unavailable. Please install openpyxl.',
            )
            return redirect(reverse('currency_list'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Currencies'
        ws.append([
            'currency_code', 'name_en', 'name_ar', 'currency_symbol',
            'decimal_places', 'is_active',
        ])
        ws.append(['SAR', 'Saudi Riyal', 'ريال سعودي', 'ر.س', 2, 'TRUE'])
        ws.append(['USD', 'US Dollar', 'دولار أمريكي', '$', 2, 'TRUE'])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=\"currency_import_sample.xlsx\"'
        return response


class GeneralTaxSettingsView(LoginRequiredMixin, View):
    template_name = 'system_config/general_tax_settings.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        obj, _created = GeneralTaxSettings.objects.get_or_create(
            setting_id='GLOBAL-TAX-SETTING',
            defaults={
                'prices_include_tax': False,
                'location_verification': 'Profile_Only',
            },
        )
        form = GeneralTaxSettingsForm(instance=obj)
        return render(request, self.template_name, {'form': form, 'obj': obj})

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        obj, _created = GeneralTaxSettings.objects.get_or_create(
            setting_id='GLOBAL-TAX-SETTING',
            defaults={
                'prices_include_tax': False,
                'location_verification': 'Profile_Only',
            },
        )
        form = GeneralTaxSettingsForm(request.POST, instance=obj)
        old_obj = GeneralTaxSettings.objects.get(setting_id=obj.setting_id)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'obj': obj},
            )

        form.instance.updated_by = request.user
        form.instance.save(update_fields=['prices_include_tax', 'location_verification', 'updated_by', 'updated_at'])
        log_audit_action(
            request,
            'Update',
            'General Tax Settings',
            str(obj.setting_id),
            old_instance=old_obj,
            new_instance=obj,
        )
        messages.success(request, 'General tax settings saved successfully.')
        return redirect(reverse('general_tax_settings'))


class LegalIdentityView(LoginRequiredMixin, View):
    template_name = 'system_config/legal_identity.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        obj, _created = LegalIdentity.objects.get_or_create(
            identity_id='GLOBAL-LEGAL-IDENTITY',
            defaults={
                'company_logo': None,
                'company_name_en': 'IRoad',
                'company_name_ar': 'IRoad',
                'company_country_code': None,
                'commercial_register': 'N/A',
                'tax_number': 'N/A',
                'registered_address': 'N/A',
                'support_email': 'admin@example.com',
                'support_phone': '',
            },
        )
        form = LegalIdentityForm(instance=obj)
        return render(request, self.template_name, {'form': form, 'obj': obj})

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        obj, _created = LegalIdentity.objects.get_or_create(
            identity_id='GLOBAL-LEGAL-IDENTITY',
            defaults={
                'company_logo': None,
                'company_name_en': 'IRoad',
                'company_name_ar': 'IRoad',
                'company_country_code': None,
                'commercial_register': 'N/A',
                'tax_number': 'N/A',
                'registered_address': 'N/A',
                'support_email': 'admin@example.com',
                'support_phone': '',
            },
        )

        form = LegalIdentityForm(request.POST, request.FILES, instance=obj)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'obj': obj},
            )

        company_logo = request.FILES.get('company_logo')
        if company_logo:
            ext = os.path.splitext(company_logo.name or '')[1].lower() or '.png'
            company_logo.name = f'legal_{obj.identity_id}_{uuid.uuid4().hex[:10]}{ext}'

        form.instance.updated_by = request.user
        form.instance.save(update_fields=[
            'company_logo',
            'company_name_en',
            'company_name_ar',
            'company_country_code',
            'commercial_register',
            'tax_number',
            'registered_address',
            'support_email',
            'support_phone',
            'updated_by',
            'updated_at',
        ])
        messages.success(request, 'IRoad legal identity saved successfully.')
        return redirect(reverse('legal_identity'))


class GlobalSystemRulesView(LoginRequiredMixin, View):
    template_name = 'system_config/global_system_rules.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        obj, _created = GlobalSystemRules.objects.get_or_create(
            rule_id='GLOBAL-SYSTEM-RULES',
            defaults={
                'system_timezone': 'Asia/Riyadh',
                'default_date_format': 'DD/MM/YYYY',
                'grace_period_days': 3,
                'standard_billing_cycle': 30,
            },
        )
        form = GlobalSystemRulesForm(instance=obj)
        return render(request, self.template_name, {'form': form, 'obj': obj})

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        obj, _created = GlobalSystemRules.objects.get_or_create(
            rule_id='GLOBAL-SYSTEM-RULES',
            defaults={
                'system_timezone': 'Asia/Riyadh',
                'default_date_format': 'DD/MM/YYYY',
                'grace_period_days': 3,
                'standard_billing_cycle': 30,
            },
        )

        form = GlobalSystemRulesForm(request.POST, instance=obj)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'obj': obj},
            )

        form.instance.updated_by = request.user
        form.instance.save(update_fields=[
            'system_timezone',
            'default_date_format',
            'grace_period_days',
            'standard_billing_cycle',
            'updated_by',
            'updated_at',
        ])
        messages.success(request, 'Global system rules saved successfully.')
        return redirect(reverse('global_system_rules'))


class BaseCurrencyView(LoginRequiredMixin, View):
    template_name = 'system_config/base_currency.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('dashboard'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        # Phase 4: create once; Phase 5 will enforce immutability based on transactions.
        sar = Currency.objects.filter(currency_code='SAR').first()
        obj, _created = BaseCurrencyConfig.objects.get_or_create(
            setting_id='GLOBAL-BASE-CURRENCY',
            defaults={'base_currency': sar},
        )
        form = BaseCurrencyForm(instance=obj)
        return render(request, self.template_name, {'form': form, 'obj': obj})

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp

        sar = Currency.objects.filter(currency_code='SAR').first()
        obj, _created = BaseCurrencyConfig.objects.get_or_create(
            setting_id='GLOBAL-BASE-CURRENCY',
            defaults={'base_currency': sar},
        )

        form = BaseCurrencyForm(request.POST, instance=obj)
        old_obj = BaseCurrencyConfig.objects.get(setting_id=obj.setting_id)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'obj': obj},
            )

        # TODO Phase 5: Check if any financial transactions
        #               exist before allowing change.
        #               If yes, block change entirely.
        form.instance.updated_by = request.user
        form.instance.save(update_fields=['base_currency', 'updated_by', 'updated_at'])
        log_audit_action(
            request,
            'Update',
            'Base Currency Config',
            str(obj.setting_id),
            old_instance=old_obj,
            new_instance=obj,
        )
        messages.success(request, 'Base currency saved successfully.')
        return redirect(reverse('base_currency'))


def _get_base_currency_code():
    """
    Helper to fetch the current base currency code (used to exclude it from FX rates).
    """
    obj, _created = BaseCurrencyConfig.objects.get_or_create(
        setting_id='GLOBAL-BASE-CURRENCY',
        defaults={
            'base_currency': Currency.objects.filter(currency_code='SAR').first(),
        },
    )
    if obj.base_currency_id:
        return obj.base_currency.currency_code
    return None


def _require_root_or_redirect(request):
    if not getattr(request.user, 'is_root', False):
        messages.error(request, 'Access denied: root admin only.')
        return redirect(reverse('dashboard'))
    return None


class ExchangeRateListView(LoginRequiredMixin, View):
    template_name = 'system_config/exchange_rates/fx_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'desc')

        base_code = _get_base_currency_code()
        base_config = BaseCurrencyConfig.objects.get_or_create(
            setting_id='GLOBAL-BASE-CURRENCY',
            defaults={'base_currency': Currency.objects.filter(currency_code='SAR').first()},
        )[0]
        base_currency = base_config.base_currency

        qs = (
            ExchangeRate.objects.select_related('currency')
            .annotate(
                default_rank=Window(
                    expression=RowNumber(),
                    order_by=F('updated_at').desc(),
                ),
                status_sort=Case(
                    When(is_active=True, then=Value(1)),
                    default=Value(2),
                    output_field=IntegerField(),
                ),
            )
        )

        if search_query:
            qs = qs.filter(
                Q(currency__currency_code__icontains=search_query)
                | Q(currency__name_en__icontains=search_query)
            )

        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'code': 'currency__currency_code',
            'name': 'currency__name_en',
            'symbol': 'currency__currency_symbol',
            'rate': 'exchange_rate',
            'status': 'status_sort',
            'updated': 'updated_at',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-default_rank')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'default_rank')

        total_count = qs.count()
        paginator = Paginator(qs, 10)
        page_number = request.GET.get('page', 1)
        rates_page = paginator.get_page(page_number)

        context = {
            'exchange_rates': rates_page,
            'search_query': search_query,
            'status_filter': status_filter,
            'current_sort': sort,
            'current_dir': direction,
            'total_count': total_count,
            'base_currency': base_currency,
            'base_currency_code': base_code,
        }
        return render(request, self.template_name, context)


class ExchangeRateCreateView(LoginRequiredMixin, View):
    template_name = 'system_config/exchange_rates/fx_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        base_code = _get_base_currency_code()
        form = ExchangeRateForm(base_currency_code=base_code)
        return render(
            request,
            self.template_name,
            {'form': form, 'is_edit': False, 'base_currency_code': base_code},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        base_code = _get_base_currency_code()
        form = ExchangeRateForm(request.POST, base_currency_code=base_code)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False, 'base_currency_code': base_code},
            )

        currency = form.cleaned_data.get('currency')
        if currency and base_code and currency.currency_code == base_code:
            form.add_error('currency', 'Currency must not be the base currency.')
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False, 'base_currency_code': base_code},
            )

        if currency and ExchangeRate.objects.filter(currency=currency, is_active=True).exists():
            form.add_error(
                'currency',
                'An active rate already exists for this currency. Edit it instead.',
            )
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False, 'base_currency_code': base_code},
            )

        rate = form.save(commit=False)
        rate.updated_by = request.user
        rate.save()
        log_audit_action(
            request,
            'Create',
            'Exchange Rates',
            str(rate.fx_id),
            new_instance=rate,
        )

        FXRateChangeLog.objects.create(
            currency=rate.currency,
            old_rate=Decimal('0.000000'),
            new_rate=rate.exchange_rate,
            notes='Initial rate set',
            changed_by=request.user,
        )

        messages.success(request, 'Exchange rate created successfully.')
        return redirect(reverse('fx_rate_list'))


class ExchangeRateUpdateView(LoginRequiredMixin, View):
    template_name = 'system_config/exchange_rates/fx_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        base_code = _get_base_currency_code()
        rate = get_object_or_404(ExchangeRate, pk=pk)
        old_obj = ExchangeRate.objects.get(pk=pk)
        form = ExchangeRateForm(instance=rate, base_currency_code=base_code)
        form.fields['currency'].disabled = True

        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': True,
                'base_currency_code': base_code,
                'rate_obj': rate,
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        rate = get_object_or_404(ExchangeRate, pk=pk)

        # Optional action: toggle active status without touching the FX value.
        if request.POST.get('action') == 'toggle_status':
            new_active = not rate.is_active
            rate.is_active = new_active
            rate.updated_by = request.user
            rate.save(update_fields=['is_active', 'updated_by', 'updated_at'])
            log_audit_action(
                request,
                'Status_Change',
                'Exchange Rates',
                str(rate.fx_id),
                old_instance=old_obj,
                new_instance=rate,
            )
            messages.success(
                request,
                'Exchange rate activated successfully.' if new_active else 'Exchange rate deactivated successfully.',
            )
            return redirect(reverse('fx_rate_list'))

        base_code = _get_base_currency_code()
        form = ExchangeRateForm(request.POST, instance=rate, base_currency_code=base_code)
        # The currency field is disabled on the edit UI, so browsers won't submit it.
        # Disable it here too, so validation doesn't fail due to missing data.
        if 'currency' in form.fields:
            form.fields['currency'].disabled = True
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'base_currency_code': base_code, 'rate_obj': rate},
            )

        # Backend safeguard: currency cannot be changed.
        form.instance.currency = rate.currency
        form.instance.updated_by = request.user

        old_rate = ExchangeRate.objects.get(pk=rate.fx_id).exchange_rate
        form.save()
        log_audit_action(
            request,
            'Update',
            'Exchange Rates',
            str(rate.fx_id),
            old_instance=old_obj,
            new_instance=rate,
        )

        FXRateChangeLog.objects.create(
            currency=rate.currency,
            old_rate=old_rate,
            new_rate=rate.exchange_rate,
            notes=request.POST.get('change_notes', ''),
            changed_by=request.user,
        )

        messages.success(request, 'Exchange rate updated successfully.')
        return redirect(reverse('fx_rate_list'))


class FXRateChangeLogView(LoginRequiredMixin, View):
    template_name = 'system_config/exchange_rates/fx_log.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        currency_code = request.GET.get('currency', '').strip()
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'desc')

        qs = (
            FXRateChangeLog.objects.select_related('currency', 'changed_by')
            .annotate(
                default_rank=Window(
                    expression=RowNumber(),
                    order_by=F('changed_at').desc(),
                ),
            )
        )

        if currency_code:
            qs = qs.filter(currency__currency_code=currency_code)

        if search_query:
            qs = qs.filter(
                Q(currency__currency_code__icontains=search_query)
                | Q(currency__name_en__icontains=search_query)
                | Q(notes__icontains=search_query)
                | Q(changed_by__email__icontains=search_query)
            )

        # Annotate delta so templates can display +/- with color.
        delta_expr = ExpressionWrapper(
            F('new_rate') - F('old_rate'),
            output_field=DecimalField(max_digits=12, decimal_places=6),
        )
        qs = qs.annotate(delta=delta_expr, delta_abs=Abs(delta_expr))

        sort_mapping = {
            'rank': 'default_rank',
            'currency': 'currency__currency_code',
            'old_rate': 'old_rate',
            'new_rate': 'new_rate',
            'change': 'delta_abs',
            'notes': 'notes',
            'changed_by': 'changed_by__email',
            'changed_at': 'changed_at',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-default_rank')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'default_rank')

        paginator = Paginator(qs, 10)
        page_number = request.GET.get('page', 1)
        log_page = paginator.get_page(page_number)
        total_count = qs.count()
        start_index = log_page.start_index()
        for offset, fx_log in enumerate(log_page.object_list):
            # Show descending list ID so top rows have higher numbers.
            fx_log.list_rank = total_count - (start_index + offset) + 1

        currencies = Currency.objects.all().order_by('name_en')

        context = {
            'fx_logs': log_page,
            'currencies': currencies,
            'search_query': search_query,
            'selected_currency_code': currency_code,
            'current_sort': sort,
            'current_dir': direction,
        }
        return render(request, self.template_name, context)


class ExchangeRateImportExcelView(LoginRequiredMixin, View):
    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        if load_workbook is None:
            messages.error(
                request,
                'Excel import dependency is unavailable. Please install openpyxl.',
            )
            return redirect(reverse('fx_rate_list'))

        upload = request.FILES.get('excel_file')
        if not upload:
            messages.error(request, 'Please choose an Excel file to import.')
            return redirect(reverse('fx_rate_list'))

        filename = (upload.name or '').lower()
        if not (
            filename.endswith('.xlsx')
            or filename.endswith('.xlsm')
            or filename.endswith('.xltx')
            or filename.endswith('.xltm')
        ):
            messages.error(request, 'Only Excel .xlsx files are supported.')
            return redirect(reverse('fx_rate_list'))

        try:
            wb = load_workbook(upload, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, 'Invalid Excel file. Please use the provided sample.')
            return redirect(reverse('fx_rate_list'))

        header_row = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[1]]
        required_headers = {'currency_code', 'exchange_rate'}
        if not required_headers.issubset(set(header_row)):
            messages.error(
                request,
                'Invalid template headers. Required: currency_code, exchange_rate.',
            )
            return redirect(reverse('fx_rate_list'))

        idx = {name: header_row.index(name) for name in required_headers}
        is_active_idx = header_row.index('is_active') if 'is_active' in header_row else None
        notes_idx = header_row.index('change_notes') if 'change_notes' in header_row else None

        created_count = 0
        skipped_count = 0
        row_errors = []
        base_code = _get_base_currency_code()

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            currency_code = str(row[idx['currency_code']] or '').strip().upper()
            raw_rate = row[idx['exchange_rate']]

            if not currency_code and raw_rate in (None, ''):
                continue
            if not currency_code or raw_rate in (None, ''):
                skipped_count += 1
                row_errors.append(f'Row {row_no}: missing required value(s).')
                continue

            if base_code and currency_code == base_code:
                skipped_count += 1
                row_errors.append(f'Row {row_no}: base currency is not allowed.')
                continue

            currency = Currency.objects.filter(currency_code=currency_code, is_active=True).first()
            if not currency:
                skipped_count += 1
                row_errors.append(f'Row {row_no}: active currency not found.')
                continue

            try:
                exchange_rate = Decimal(str(raw_rate))
                if exchange_rate <= 0:
                    raise ValueError('non-positive rate')
            except Exception:
                skipped_count += 1
                row_errors.append(f'Row {row_no}: exchange_rate must be a positive number.')
                continue

            if ExchangeRate.objects.filter(currency=currency, is_active=True).exists():
                skipped_count += 1
                continue

            is_active = True
            if is_active_idx is not None:
                raw_active = row[is_active_idx]
                if raw_active is not None and str(raw_active).strip() != '':
                    norm = str(raw_active).strip().lower()
                    is_active = norm in ('1', 'true', 'yes', 'y', 'active')

            rate = ExchangeRate(
                currency=currency,
                exchange_rate=exchange_rate,
                is_active=is_active,
                updated_by=request.user,
            )
            try:
                rate.full_clean()
                rate.save()
                log_audit_action(
                    request,
                    'Create',
                    'Exchange Rates',
                    str(rate.fx_id),
                    new_instance=rate,
                )
                notes = ''
                if notes_idx is not None:
                    notes = str(row[notes_idx] or '').strip()
                FXRateChangeLog.objects.create(
                    currency=rate.currency,
                    old_rate=Decimal('0.000000'),
                    new_rate=rate.exchange_rate,
                    notes=notes or 'Imported from Excel',
                    changed_by=request.user,
                )
                created_count += 1
            except Exception:
                skipped_count += 1
                row_errors.append(f'Row {row_no}: invalid data or duplicate.')

        if created_count:
            messages.success(
                request,
                f'Exchange-rate import completed. Created {created_count}, skipped {skipped_count}.',
            )
        else:
            messages.error(
                request,
                f'No exchange rates imported. Skipped {skipped_count}.',
            )
        if row_errors:
            messages.warning(request, ' | '.join(row_errors[:3]))
        return redirect(reverse('fx_rate_list'))


class ExchangeRateSampleExcelView(LoginRequiredMixin, View):
    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        if Workbook is None:
            messages.error(
                request,
                'Excel sample dependency is unavailable. Please install openpyxl.',
            )
            return redirect(reverse('fx_rate_list'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'ExchangeRates'
        ws.append(['currency_code', 'exchange_rate', 'is_active', 'change_notes'])
        ws.append(['USD', '3.750000', 'TRUE', 'Initial import'])
        ws.append(['EUR', '4.090000', 'TRUE', 'Initial import'])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="exchange_rate_import_sample.xlsx"'
        return response


class TaxCodeListView(LoginRequiredMixin, View):
    template_name = 'system_config/tax_codes/tax_code_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'asc')

        tax_codes_qs = TaxCode.objects.filter(is_deleted=False).select_related(
            'applicable_country_code'
        ).annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('tax_code').asc(),
            ),
            country_default_sort=Case(
                When(is_default_for_country=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
            intl_default_sort=Case(
                When(is_international_default=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
            status_sort=Case(
                When(is_active=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
        )

        if search_query:
            tax_codes_qs = tax_codes_qs.filter(
                Q(name_en__icontains=search_query)
                | Q(tax_code__icontains=search_query)
                | Q(applicable_country_code__name_en__icontains=search_query)
            )

        if status_filter == 'Active':
            tax_codes_qs = tax_codes_qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            tax_codes_qs = tax_codes_qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'code': 'tax_code',
            'rate': 'rate_percent',
            'country': 'applicable_country_code__name_en',
            'country_default': 'country_default_sort',
            'intl_default': 'intl_default_sort',
            'status': 'status_sort',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            tax_codes_qs = tax_codes_qs.order_by(
                F(order_by_field).desc(nulls_last=True),
                '-default_rank',
            )
        else:
            tax_codes_qs = tax_codes_qs.order_by(
                F(order_by_field).asc(nulls_first=True),
                'default_rank',
            )

        total_count = tax_codes_qs.count()
        paginator = Paginator(tax_codes_qs, 10)
        page_number = request.GET.get('page', 1)
        tax_codes_page = paginator.get_page(page_number)
        start_index = tax_codes_page.start_index()
        for offset, tax_code in enumerate(tax_codes_page.object_list):
            # Show descending list ID so top rows have higher numbers.
            tax_code.list_rank = total_count - (start_index + offset) + 1

        context = {
            'tax_codes': tax_codes_page,
            'search_query': search_query,
            'status_filter': status_filter,
            'current_sort': sort,
            'current_dir': direction,
            'total_count': total_count,
            'page_title': 'Tax Codes Master',
        }
        return render(request, self.template_name, context)


class TaxCodeDetailView(LoginRequiredMixin, View):
    template_name = 'system_config/tax_codes/tax_code_detail.html'

    def get(self, request, pk):
        tax_code = get_object_or_404(
            TaxCode.objects.select_related(
                'applicable_country_code', 'created_by', 'updated_by'
            ),
            pk=pk,
        )
        return render(
            request,
            self.template_name,
            {'tax_code': tax_code, 'page_title': 'Tax Code Details'},
        )


class TaxCodeCreateView(LoginRequiredMixin, View):
    template_name = 'system_config/tax_codes/tax_code_form.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('tax_code_list'))
        return None

    def get(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        form = TaxCodeForm(is_edit=False)
        return render(
            request,
            self.template_name,
            {'form': form, 'is_edit': False},
        )

    def post(self, request):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        form = TaxCodeForm(request.POST, is_edit=False)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False},
            )
        tax_code = form.save(commit=False)
        tax_code.updated_by = request.user
        tax_code.created_by = request.user
        tax_code.save()
        messages.success(request, 'Tax code created successfully.')
        return redirect(reverse('tax_code_list'))


class TaxCodeUpdateView(LoginRequiredMixin, View):
    template_name = 'system_config/tax_codes/tax_code_form.html'

    def _require_root(self, request):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('tax_code_list'))
        return None

    def get(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        tax_code = get_object_or_404(TaxCode, pk=pk)
        form = TaxCodeForm(instance=tax_code, is_edit=True)
        return render(
            request,
            self.template_name,
            {'form': form, 'is_edit': True, 'tax_code_obj': tax_code},
        )

    def post(self, request, pk):
        redirect_resp = self._require_root(request)
        if redirect_resp:
            return redirect_resp
        tax_code = get_object_or_404(TaxCode, pk=pk)
        form = TaxCodeForm(request.POST, instance=tax_code, is_edit=True)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'tax_code_obj': tax_code},
            )
        tax_code_obj = form.save(commit=False)
        tax_code_obj.tax_code = tax_code.tax_code
        tax_code_obj.updated_by = request.user
        tax_code_obj.save()
        messages.success(request, 'Tax code updated successfully.')
        return redirect(reverse('tax_code_list'))


class TaxCodeToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('tax_code_list'))
        tax_code = get_object_or_404(TaxCode, pk=pk)
        deactivating_default = (
            tax_code.is_active
            and (tax_code.is_default_for_country or tax_code.is_international_default)
        )
        tax_code.is_active = not tax_code.is_active
        tax_code.updated_by = request.user
        tax_code.save(update_fields=['is_active', 'updated_by'])
        if deactivating_default:
            messages.warning(
                request,
                'You deactivated a default tax code. '
                'Review country/international defaults.',
            )
        messages.success(request, 'Tax code status updated successfully.')
        return redirect(reverse('tax_code_list'))


class TaxCodeDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('tax_code_list'))
        tax_code = get_object_or_404(TaxCode, pk=pk)
        if tax_code.is_deleted:
            messages.info(request, 'Tax code is already removed.')
            return redirect(reverse('tax_code_list'))
        tax_code.is_deleted = True
        tax_code.updated_by = request.user
        tax_code.save(update_fields=['is_deleted', 'updated_by'])
        messages.success(request, 'Tax code removed successfully.')
        return redirect(reverse('tax_code_list'))

    def get(self, request, pk):
        return self.post(request, pk)


class PlanListView(LoginRequiredMixin, View):
    template_name = 'subscription/plans/plan_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort_by = request.GET.get('sort', 'updated_at')
        sort_dir = request.GET.get('dir', 'desc')

        plans_qs = SubscriptionPlan.objects.filter(is_deleted=False).annotate(
            pricing_rows_count=Count('pricing_cycles')
        )

        # Sortable fields whitelist and mapping
        sort_mapping = {
            'rank': ['default_rank'],
            'name_en': ['plan_name_en'],
            'name_ar': ['plan_name_ar'],
            'cycle': ['base_cycle_days'],
            'pricing': ['pricing_rows_count'],
            'status': ['is_active'],
            'updated_at': ['updated_at'],
        }

        active_sort_fields = sort_mapping.get(sort_by, ['updated_at'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else '-' + f)

        if search_query:
            plans_qs = plans_qs.filter(
                Q(plan_name_en__icontains=search_query)
                | Q(plan_name_ar__icontains=search_query)
            )

        if status_filter == 'Active':
            plans_qs = plans_qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            plans_qs = plans_qs.filter(is_active=False)

        # Annotate with stable rank based on default order (newest first).
        plans_qs = plans_qs.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc()
            )
        )

        plans_qs = plans_qs.order_by(*ordering)
        total_count = plans_qs.count()

        paginator = Paginator(plans_qs, 10)
        page_number = request.GET.get('page', 1)
        plans_page = paginator.get_page(page_number)
        start_index = plans_page.start_index()
        for offset, plan in enumerate(plans_page.object_list):
            # Show descending list ID so newest appears with highest number.
            plan.list_rank = total_count - (start_index + offset) + 1

        return render(
            request,
            self.template_name,
            {
                'plans': plans_page,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort_by,
                'current_dir': sort_dir,
                'total_count': total_count,
            },
        )


class PlanCreateView(LoginRequiredMixin, View):
    template_name = 'subscription/plans/plan_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        return render(
            request,
            self.template_name,
            {
                'form': SubscriptionPlanForm(),
                'is_edit': False,
                'pricing_rows': [self._empty_row(0)],
                'currencies': Currency.objects.filter(is_active=True).order_by('name_en'),
                'global_currency_code': '',
            },
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        form = SubscriptionPlanForm(request.POST)
        rows = self._extract_rows(request.POST)
        global_currency_code = request.POST.get('global_currency', '').strip()
        rows = self._apply_global_currency(rows, global_currency_code)
        valid_rows, row_errors, duplicate_error = self._validate_rows(rows)

        has_errors = False
        if not form.is_valid():
            has_errors = True
        if not global_currency_code:
            has_errors = True
            messages.error(request, 'Please select a currency for pricing cycles.')
        if not valid_rows:
            has_errors = True
            messages.error(request, 'At least one pricing cycle is required.')
        if row_errors:
            has_errors = True
        if duplicate_error:
            has_errors = True
            messages.error(request, duplicate_error)

        if has_errors:
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': False,
                    'pricing_rows': self._rows_with_errors(rows, row_errors),
                    'currencies': Currency.objects.filter(is_active=True).order_by('name_en'),
                    'global_currency_code': global_currency_code,
                },
            )

        plan = form.save(commit=False)
        plan.created_by = request.user
        plan.save()
        log_audit_action(
            request,
            'Create',
            'Subscription Plans',
            str(plan.plan_id),
            new_instance=plan,
        )

        for row in valid_rows:
            PlanPricingCycle.objects.create(
                plan=plan,
                number_of_cycles=row['cleaned_data']['number_of_cycles'],
                currency=row['cleaned_data']['currency'],
                price=row['cleaned_data']['price'],
                is_admin_only_cycle=row['cleaned_data']['is_admin_only_cycle'],
            )

        messages.success(request, 'Subscription plan created successfully.')
        return redirect(reverse('plan_detail', kwargs={'pk': plan.plan_id}))

    def _empty_row(self, index):
        return {
            'row_index': index,
            'pricing_id': '',
            'number_of_cycles': '',
            'currency': '',
            'price': '',
            'is_admin_only_cycle': False,
            'delete': False,
            'errors': [],
        }

    def _extract_rows(self, post_data):
        row_indices = set()
        for key in post_data.keys():
            if key.startswith('pricing-'):
                parts = key.split('-')
                if len(parts) >= 3 and parts[1].isdigit():
                    row_indices.add(int(parts[1]))

        rows = []
        for index in sorted(row_indices):
            prefix = f'pricing-{index}-'
            row = {
                'row_index': index,
                'pricing_id': post_data.get(prefix + 'pricing_id', '').strip(),
                'number_of_cycles': post_data.get(prefix + 'number_of_cycles', '').strip(),
                'currency': post_data.get(prefix + 'currency', '').strip(),
                'price': post_data.get(prefix + 'price', '').strip(),
                'is_admin_only_cycle': (
                    post_data.get(prefix + 'is_admin_only_cycle', '').strip() in ('1', 'on', 'true', 'True')
                ),
                'delete': post_data.get(prefix + 'delete', '').strip() == '1',
            }
            rows.append(row)
        return rows

    def _apply_global_currency(self, rows, global_currency_code):
        normalized_currency = (global_currency_code or '').strip()
        for row in rows:
            if row.get('delete'):
                continue
            if not any([
                row.get('number_of_cycles'),
                row.get('currency'),
                row.get('price'),
            ]):
                continue
            row['currency'] = normalized_currency
        return rows

    def _validate_rows(self, rows):
        valid_rows = []
        row_errors = {}
        seen = set()
        duplicate_error = None

        for row in rows:
            if row.get('delete'):
                continue
            if not any([
                row.get('number_of_cycles'),
                row.get('currency'),
                row.get('price'),
            ]):
                continue

            form = PlanPricingCycleForm(
                {
                    'number_of_cycles': row.get('number_of_cycles'),
                    'currency': row.get('currency'),
                    'price': row.get('price'),
                    'is_admin_only_cycle': row.get('is_admin_only_cycle', False),
                }
            )
            if not form.is_valid():
                row_errors[row['row_index']] = form.errors
                continue

            combo = (
                form.cleaned_data['number_of_cycles'],
                form.cleaned_data['currency'].currency_code,
            )
            if combo in seen:
                duplicate_error = (
                    'Duplicate pricing row found for same cycles and currency.'
                )
                continue
            seen.add(combo)
            valid_rows.append({'row_index': row['row_index'], 'cleaned_data': form.cleaned_data})

        return valid_rows, row_errors, duplicate_error

    def _rows_with_errors(self, rows, row_errors):
        merged = []
        if not rows:
            return [self._empty_row(0)]
        for row in rows:
            row_copy = dict(row)
            errors = row_errors.get(row['row_index'])
            row_copy['errors'] = (
                [f"{k}: {', '.join(v)}" for k, v in errors.items()]
                if errors else []
            )
            merged.append(row_copy)
        return merged


class PlanDetailView(LoginRequiredMixin, View):
    template_name = 'subscription/plans/plan_detail.html'

    def get(self, request, pk):
        plan = get_object_or_404(
            SubscriptionPlan.objects.prefetch_related('pricing_cycles__currency'),
            pk=pk,
        )
        return render(
            request,
            self.template_name,
            {
                'plan': plan,
                'pricing_cycles': plan.pricing_cycles.all().order_by('number_of_cycles'),
            },
        )


class PlanUpdateView(PlanCreateView):
    template_name = 'subscription/plans/plan_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        plan = get_object_or_404(SubscriptionPlan, pk=pk)
        form = SubscriptionPlanForm(instance=plan)
        pricing_rows = []
        for idx, pricing in enumerate(
            plan.pricing_cycles.select_related('currency').all().order_by('number_of_cycles'),
            start=0,
        ):
            pricing_rows.append({
                'row_index': idx,
                'pricing_id': str(pricing.pricing_id),
                'number_of_cycles': pricing.number_of_cycles,
                'currency': pricing.currency_id,
                'price': pricing.price,
                'is_admin_only_cycle': pricing.is_admin_only_cycle,
                'delete': False,
                'errors': [],
            })
        if not pricing_rows:
            pricing_rows = [self._empty_row(0)]
        global_currency_code = ''
        for row in pricing_rows:
            if row.get('currency'):
                global_currency_code = row['currency']
                break

        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': True,
                'plan': plan,
                'pricing_rows': pricing_rows,
                'currencies': Currency.objects.filter(is_active=True).order_by('name_en'),
                'global_currency_code': global_currency_code,
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        plan = get_object_or_404(SubscriptionPlan, pk=pk)
        old_obj = SubscriptionPlan.objects.get(pk=pk)
        form = SubscriptionPlanForm(request.POST, instance=plan)
        rows = self._extract_rows(request.POST)
        global_currency_code = request.POST.get('global_currency', '').strip()
        rows = self._apply_global_currency(rows, global_currency_code)
        valid_rows, row_errors, duplicate_error = self._validate_rows(rows)

        has_errors = False
        if not form.is_valid():
            has_errors = True
        if not global_currency_code:
            has_errors = True
            messages.error(request, 'Please select a currency for pricing cycles.')
        if not valid_rows:
            has_errors = True
            messages.error(request, 'At least one pricing cycle is required.')
        if row_errors:
            has_errors = True
        if duplicate_error:
            has_errors = True
            messages.error(request, duplicate_error)

        if has_errors:
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': True,
                    'plan': plan,
                    'pricing_rows': self._rows_with_errors(rows, row_errors),
                    'currencies': Currency.objects.filter(is_active=True).order_by('name_en'),
                    'global_currency_code': global_currency_code,
                },
            )

        plan = form.save()
        log_audit_action(
            request,
            'Update',
            'Subscription Plans',
            str(plan.plan_id),
            old_instance=old_obj,
            new_instance=plan,
        )
        existing_map = {
            str(item.pricing_id): item
            for item in plan.pricing_cycles.all()
        }
        keep_ids = set()

        for row in rows:
            if row.get('delete') and row.get('pricing_id'):
                existing = existing_map.get(row['pricing_id'])
                if existing:
                    existing.delete()
                continue
            if row.get('delete'):
                continue
            if not any([row.get('number_of_cycles'), row.get('currency'), row.get('price')]):
                continue

            cleaned = next(
                (vr['cleaned_data'] for vr in valid_rows if vr['row_index'] == row['row_index']),
                None,
            )
            if cleaned is None:
                continue

            if row.get('pricing_id'):
                existing = existing_map.get(row['pricing_id'])
                if existing:
                    existing.number_of_cycles = cleaned['number_of_cycles']
                    existing.currency = cleaned['currency']
                    existing.price = cleaned['price']
                    existing.is_admin_only_cycle = cleaned['is_admin_only_cycle']
                    existing.save()
                    keep_ids.add(str(existing.pricing_id))
                    continue

            new_obj = PlanPricingCycle.objects.create(
                plan=plan,
                number_of_cycles=cleaned['number_of_cycles'],
                currency=cleaned['currency'],
                price=cleaned['price'],
                is_admin_only_cycle=cleaned['is_admin_only_cycle'],
            )
            keep_ids.add(str(new_obj.pricing_id))

        for pricing_id, pricing_obj in existing_map.items():
            if pricing_id not in keep_ids:
                if not any(r.get('pricing_id') == pricing_id and r.get('delete') for r in rows):
                    pricing_obj.delete()

        if not plan.pricing_cycles.exists():
            messages.error(request, 'At least one pricing cycle is required.')
            return redirect(reverse('plan_edit', kwargs={'pk': plan.plan_id}))

        messages.success(request, 'Subscription plan updated successfully.')
        return redirect(reverse('plan_detail', kwargs={'pk': plan.plan_id}))


class PlanToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        plan = get_object_or_404(SubscriptionPlan, pk=pk)
        old_obj = SubscriptionPlan.objects.get(pk=pk)
        if plan.is_active:
            # TODO Phase 6: Check active tenant subscriptions
            #               before deactivating this plan
            plan.is_active = False
            messages.success(request, 'Plan deactivated successfully.')
        else:
            plan.is_active = True
            messages.success(request, 'Plan activated successfully.')
        plan.save(update_fields=['is_active', 'updated_at'])
        log_audit_action(
            request,
            'Status_Change',
            'Subscription Plans',
            str(plan.plan_id),
            old_instance=old_obj,
            new_instance=plan,
        )
        return redirect(reverse('plan_list'))


class PlanDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('plan_list'))
        plan = get_object_or_404(SubscriptionPlan, pk=pk)
        if plan.is_deleted:
            messages.info(request, 'Plan is already removed.')
            return redirect(reverse('plan_list'))
        plan.is_deleted = True
        plan.save(update_fields=['is_deleted'])
        messages.success(request, 'Plan removed successfully.')
        return redirect(reverse('plan_list'))

    def get(self, request, pk):
        return self.post(request, pk)


class AddOnsPolicyListView(LoginRequiredMixin, View):
    template_name = 'subscription/addons/policy_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort_by = request.GET.get('sort', 'created_at')
        sort_dir = request.GET.get('dir', 'desc')

        policies_qs = AddOnsPricingPolicy.objects.filter(is_deleted=False)

        # Sortable fields whitelist and mapping
        sort_mapping = {
            'rank': ['default_rank'],
            'policy_name': ['policy_name'],
            'user_price': ['extra_internal_user_price'],
            'truck_price': ['extra_internal_truck_price'],
            'ext_truck_price': ['extra_external_truck_price'],
            'driver_price': ['extra_driver_price'],
            'shipment_price': ['extra_shipment_price'],
            'storage_price': ['extra_storage_gb_price'],
            'status': ['is_active'],
            'updated_at': ['updated_at'],
        }

        active_sort_fields = sort_mapping.get(sort_by, ['updated_at'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else '-' + f)

        if search_query:
            policies_qs = policies_qs.filter(
                policy_name__icontains=search_query,
            )

        if status_filter == 'Active':
            policies_qs = policies_qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            policies_qs = policies_qs.filter(is_active=False)

        # Annotate with stable rank based on default order (newest first).
        policies_qs = policies_qs.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('updated_at').desc()
            )
        )

        policies_qs = policies_qs.order_by(*ordering)
        total_count = policies_qs.count()

        paginator = Paginator(policies_qs, 10)
        page_number = request.GET.get('page', 1)
        policies_page = paginator.get_page(page_number)
        start_index = policies_page.start_index()
        for offset, policy in enumerate(policies_page.object_list):
            # Show descending list ID so newest appears with highest number.
            policy.list_rank = total_count - (start_index + offset) + 1

        return render(
            request,
            self.template_name,
            {
                'policies': policies_page,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort_by,
                'current_dir': sort_dir,
                'total_count': total_count,
            },
        )


class AddOnsPolicyCreateView(LoginRequiredMixin, View):
    template_name = 'subscription/addons/policy_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {'form': AddOnsPricingPolicyForm(), 'is_edit': False},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        form = AddOnsPricingPolicyForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False},
            )

        policy = form.save(commit=False)
        policy.updated_by = request.user
        policy.save()

        if policy.is_active:
            AddOnsPricingPolicy.objects.exclude(
                policy_id=policy.policy_id
            ).update(is_active=False)

        messages.success(request, 'Add-ons policy saved successfully.')
        return redirect(reverse('addons_policy_list'))


class AddOnsPolicyUpdateView(LoginRequiredMixin, View):
    template_name = 'subscription/addons/policy_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        policy = get_object_or_404(AddOnsPricingPolicy, pk=pk)
        return render(
            request,
            self.template_name,
            {
                'form': AddOnsPricingPolicyForm(instance=policy),
                'is_edit': True,
                'policy': policy,
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        policy = get_object_or_404(AddOnsPricingPolicy, pk=pk)
        form = AddOnsPricingPolicyForm(request.POST, instance=policy)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'policy': policy},
            )
        policy_obj = form.save(commit=False)
        policy_obj.updated_by = request.user
        policy_obj.save()

        if policy_obj.is_active:
            AddOnsPricingPolicy.objects.exclude(
                policy_id=policy_obj.policy_id
            ).update(is_active=False)

        messages.success(request, 'Add-ons policy updated successfully.')
        return redirect(reverse('addons_policy_list'))


class AddOnsPolicyDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('addons_policy_list'))
        policy = get_object_or_404(AddOnsPricingPolicy, pk=pk)
        if policy.is_deleted:
            messages.info(request, 'Policy is already removed.')
            return redirect(reverse('addons_policy_list'))
        policy.is_deleted = True
        policy.save(update_fields=['is_deleted'])
        messages.success(request, 'Policy removed successfully.')
        return redirect(reverse('addons_policy_list'))

    def get(self, request, pk):
        return self.post(request, pk)


class PromoCodeListView(LoginRequiredMixin, View):
    template_name = 'subscription/promo/promo_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort_by = request.GET.get('sort', 'created_at')
        sort_dir = request.GET.get('dir', 'desc')

        qs = PromoCode.objects.filter(is_deleted=False).prefetch_related('applicable_plans')

        # Sortable fields whitelist and mapping
        sort_mapping = {
            'rank': ['default_rank'],
            'code': ['code'],
            'type': ['discount_type'],
            'value': ['discount_value'],
            'duration': ['discount_duration'],
            'valid_from': ['valid_from'],
            'valid_until': ['valid_until'],
            'uses': ['current_uses'],
            'status': ['is_active'],
            'created_at': ['created_at'],
        }

        active_sort_fields = sort_mapping.get(sort_by, ['created_at'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else '-' + f)

        if search_query:
            qs = qs.filter(code__icontains=search_query)

        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        # Annotate with stable rank based on default order (-created_at)
        qs = qs.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc()
            )
        )

        qs = qs.order_by(*ordering)
        total_count = qs.count()

        paginator = Paginator(qs, 10)
        page_number = request.GET.get('page', 1)
        promo_page = paginator.get_page(page_number)
        start_index = promo_page.start_index()
        for offset, promo in enumerate(promo_page.object_list):
            # Show descending list ID so newest appears with highest number.
            promo.list_rank = total_count - (start_index + offset) + 1

        now = timezone.now()
        return render(
            request,
            self.template_name,
            {
                'promo_codes': promo_page,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort_by,
                'current_dir': sort_dir,
                'now': now,
            },
        )


class PromoCodeDetailView(LoginRequiredMixin, View):
    template_name = 'subscription/promo/promo_detail.html'

    def get(self, request, pk):
        promo = get_object_or_404(
            PromoCode.objects.prefetch_related('applicable_plans'),
            pk=pk,
        )
        selected_plan_ids = list(
            promo.applicable_plans.values_list('plan_id', flat=True),
        )
        applies_all_plans = len(selected_plan_ids) == 0
        all_active_plans = SubscriptionPlan.objects.filter(
            is_active=True,
        ).order_by('plan_name_en')
        now = timezone.now()
        return render(
            request,
            self.template_name,
            {
                'promo': promo,
                'now': now,
                'all_active_plans': all_active_plans,
                'selected_plan_ids': selected_plan_ids,
                'applies_all_plans': applies_all_plans,
            },
        )


class PromoCodeCreateView(LoginRequiredMixin, View):
    template_name = 'subscription/promo/promo_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {'form': PromoCodeForm(), 'is_edit': False},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = PromoCodeForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False},
            )
        promo = form.save(commit=False)
        promo.code = promo.code.upper().strip()
        promo.created_by = request.user
        promo.current_uses = 0
        promo.save()
        form.save_m2m()
        messages.success(request, 'Promo code created successfully.')
        return redirect(reverse('promo_code_list'))


class PromoCodeUpdateView(LoginRequiredMixin, View):
    template_name = 'subscription/promo/promo_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        promo = get_object_or_404(PromoCode, pk=pk)
        form = PromoCodeForm(instance=promo)
        form.fields['code'].disabled = True
        return render(
            request,
            self.template_name,
            {'form': form, 'is_edit': True, 'promo': promo},
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        promo = get_object_or_404(PromoCode, pk=pk)
        form = PromoCodeForm(request.POST, instance=promo)
        form.fields['code'].disabled = True
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'promo': promo},
            )
        promo_obj = form.save(commit=False)
        promo_obj.code = promo.code
        # Save only the fields that were actually in the form to avoid overwriting current_uses
        save_fields = [
            'discount_type', 'discount_value', 'discount_duration',
            'valid_from', 'valid_until', 'max_uses', 'is_active',
        ]
        promo_obj.save(update_fields=save_fields)
        form.save_m2m()
        messages.success(request, 'Promo code updated successfully.')
        return redirect(reverse('promo_code_list'))


class PromoCodeToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        promo = get_object_or_404(PromoCode, pk=pk)
        promo.is_active = not promo.is_active
        promo.save(update_fields=['is_active'])
        messages.success(request, 'Promo code status updated successfully.')
        return redirect(reverse('promo_code_list'))


class PromoCodeDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('promo_code_list'))
        promo = get_object_or_404(PromoCode, pk=pk)
        if promo.is_deleted:
            messages.info(request, 'Promo code is already removed.')
            return redirect(reverse('promo_code_list'))
        promo.is_deleted = True
        promo.save(update_fields=['is_deleted'])
        messages.success(request, 'Promo code removed successfully.')
        return redirect(reverse('promo_code_list'))

    def get(self, request, pk):
        return self.post(request, pk)


class BankAccountListView(LoginRequiredMixin, View):
    template_name = 'payment/bank_accounts/account_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        currency_filter = request.GET.get('currency', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'desc')

        qs = BankAccount.objects.select_related('currency').annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('bank_name').asc(),
            ),
        )
        if search_query:
            qs = qs.filter(
                Q(bank_name__icontains=search_query)
                | Q(iban_number__icontains=search_query)
                | Q(account_holder_name__icontains=search_query)
                | Q(account_number__icontains=search_query)
            )
        if currency_filter:
            qs = qs.filter(currency_id=currency_filter)
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'bank': 'bank_name',
            'holder': 'account_holder_name',
            'iban': 'iban_number',
            'account_no': 'account_number',
            'currency': 'currency_id',
            'swift': 'swift_code',
            'status': 'is_active',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-default_rank')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'default_rank')

        paginator = Paginator(qs, 10)
        accounts = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = accounts.start_index()
        for offset, account in enumerate(accounts.object_list):
            # Show descending list ID so top rows have higher numbers.
            account.list_rank = total_count - (start_index + offset) + 1

        return render(
            request,
            self.template_name,
            {
                'accounts': accounts,
                'search_query': search_query,
                'currency_filter': currency_filter,
                'status_filter': status_filter,
                'currencies': Currency.objects.filter(is_active=True).order_by('name_en'),
                'current_sort': sort,
                'current_dir': direction,
            },
        )


class BankAccountCreateView(LoginRequiredMixin, View):
    template_name = 'payment/bank_accounts/account_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {'form': BankAccountForm(), 'is_edit': False},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = BankAccountForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False},
            )
        account = form.save(commit=False)
        account.iban_number = account.iban_number.upper().replace(' ', '').strip()
        account.created_by = request.user
        account.save()
        messages.success(request, 'Bank account created successfully.')
        return redirect(reverse('bank_account_list'))


class BankAccountUpdateView(LoginRequiredMixin, View):
    template_name = 'payment/bank_accounts/account_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        account = get_object_or_404(BankAccount, pk=pk)
        return render(
            request,
            self.template_name,
            {'form': BankAccountForm(instance=account), 'is_edit': True, 'account': account},
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        account = get_object_or_404(BankAccount, pk=pk)
        form = BankAccountForm(request.POST, instance=account)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'account': account},
            )
        obj = form.save(commit=False)
        obj.iban_number = obj.iban_number.upper().replace(' ', '').strip()
        obj.updated_by = request.user
        obj.save()
        messages.success(request, 'Bank account updated successfully.')
        return redirect(reverse('bank_account_list'))


class BankAccountToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        account = get_object_or_404(BankAccount, pk=pk)
        if account.is_active:
            # TODO Phase 8: Check if account is linked to active
            #               Payment Methods before deactivating
            account.is_active = False
            messages.success(request, 'Bank account deactivated successfully.')
        else:
            account.is_active = True
            messages.success(request, 'Bank account activated successfully.')
        account.save(update_fields=['is_active'])
        return redirect(reverse('bank_account_list'))


class BankAccountDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        messages.error(request, 'Bank accounts cannot be deleted. Deactivate instead.')
        return redirect(reverse('bank_account_list'))

    def get(self, request, pk):
        messages.error(request, 'Bank accounts cannot be deleted. Deactivate instead.')
        return redirect(reverse('bank_account_list'))


GATEWAY_MASKED_CREDENTIALS_PLACEHOLDER = '{"masked":"********"}'


class PaymentGatewayListView(LoginRequiredMixin, View):
    template_name = 'payment/gateways/gateway_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        environment_filter = request.GET.get('environment', 'All')
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'created_at')
        direction = request.GET.get('dir', 'desc')

        qs = PaymentGateway.objects.filter(is_deleted=False).annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc(),
            ),
        )
        if search_query:
            qs = qs.filter(gateway_name__icontains=search_query)
        if environment_filter in ['Test', 'Live']:
            qs = qs.filter(environment=environment_filter)
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'name': 'gateway_name',
            'environment': 'environment',
            'status': 'is_active',
            'created_at': 'created_at',
        }
        order_by_field = sort_mapping.get(sort, 'created_at')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-default_rank')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'default_rank')

        paginator = Paginator(qs, 10)
        gateways = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = gateways.start_index()
        for offset, gateway in enumerate(gateways.object_list):
            # Show descending list ID so top rows have higher numbers.
            gateway.list_rank = total_count - (start_index + offset) + 1

        return render(
            request,
            self.template_name,
            {
                'gateways': gateways,
                'search_query': search_query,
                'environment_filter': environment_filter,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
            },
        )


class PaymentGatewayCreateView(LoginRequiredMixin, View):
    template_name = 'payment/gateways/gateway_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {'form': PaymentGatewayForm(), 'is_edit': False},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = PaymentGatewayForm(request.POST)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False},
            )
        gateway = form.save(commit=False)
        gateway.created_by = request.user
        gateway.save()
        messages.success(request, 'Payment gateway created successfully.')
        return redirect(reverse('gateway_list'))


class PaymentGatewayUpdateView(LoginRequiredMixin, View):
    template_name = 'payment/gateways/gateway_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        gateway = get_object_or_404(PaymentGateway, pk=pk)
        form = PaymentGatewayForm(instance=gateway)
        form.initial['credentials_payload'] = GATEWAY_MASKED_CREDENTIALS_PLACEHOLDER
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': True,
                'gateway': gateway,
                'masked_placeholder': GATEWAY_MASKED_CREDENTIALS_PLACEHOLDER,
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        gateway = get_object_or_404(PaymentGateway, pk=pk)
        post_data = request.POST.copy()
        raw_payload = (post_data.get('credentials_payload') or '').strip()

        # Keep existing credentials when user keeps masked placeholder.
        if raw_payload == GATEWAY_MASKED_CREDENTIALS_PLACEHOLDER:
            post_data['credentials_payload'] = json.dumps(gateway.credentials_payload)

        form = PaymentGatewayForm(post_data, instance=gateway)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': True,
                    'gateway': gateway,
                    'masked_placeholder': GATEWAY_MASKED_CREDENTIALS_PLACEHOLDER,
                },
            )
        obj = form.save(commit=False)
        obj.updated_by = request.user
        obj.save()
        messages.success(request, 'Payment gateway updated successfully.')
        return redirect(reverse('gateway_list'))


class PaymentGatewayToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        gateway = get_object_or_404(PaymentGateway, pk=pk)
        if gateway.is_active:
            # TODO Phase 8: Check if gateway is linked to active
            #               Payment Methods before deactivating
            gateway.is_active = False
            messages.success(request, 'Payment gateway deactivated successfully.')
        else:
            gateway.is_active = True
            messages.success(request, 'Payment gateway activated successfully.')
        gateway.save(update_fields=['is_active'])
        return redirect(reverse('gateway_list'))


class PaymentGatewayDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('gateway_list'))
        gateway = get_object_or_404(PaymentGateway, pk=pk)
        if gateway.is_deleted:
            messages.info(request, 'Gateway is already removed.')
            return redirect(reverse('gateway_list'))
        gateway.is_deleted = True
        gateway.save(update_fields=['is_deleted'])
        messages.success(request, 'Gateway removed successfully.')
        return redirect(reverse('gateway_list'))

    def get(self, request, pk):
        return self.post(request, pk)


class PaymentMethodListView(LoginRequiredMixin, View):
    template_name = 'payment/methods/method_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        type_filter = request.GET.get('method_type', 'All')
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'created_at')
        direction = request.GET.get('dir', 'desc')
        qs = PaymentMethod.objects.select_related(
            'gateway',
            'dedicated_bank_account',
        ).annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc(),
            ),
            source_name=Case(
                When(method_type='Online_Gateway', then=F('gateway__gateway_name')),
                default=F('dedicated_bank_account__bank_name'),
                output_field=CharField(),
            ),
            status_sort=Case(
                When(is_active=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
        )
        if search_query:
            qs = qs.filter(
                Q(method_name_en__icontains=search_query)
                | Q(gateway__gateway_name__icontains=search_query)
                | Q(dedicated_bank_account__bank_name__icontains=search_query)
            )
        if type_filter in ['Online_Gateway', 'Offline_Bank']:
            qs = qs.filter(method_type=type_filter)
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'display_order': 'display_order',
            'name': 'method_name_en',
            'type': 'method_type',
            'source': 'source_name',
            'status': 'status_sort',
            'created_at': 'created_at',
        }
        order_by_field = sort_mapping.get(sort, 'created_at')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-method_id')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'method_id')

        paginator = Paginator(qs, 10)
        methods = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = methods.start_index()
        for offset, method in enumerate(methods.object_list):
            # Show descending list ID so top rows have higher numbers.
            method.list_rank = total_count - (start_index + offset) + 1

        return render(
            request,
            self.template_name,
            {
                'methods': methods,
                'search_query': search_query,
                'type_filter': type_filter,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
            },
        )


class PaymentMethodCreateView(LoginRequiredMixin, View):
    template_name = 'payment/methods/method_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {'form': PaymentMethodForm(), 'is_edit': False},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = PaymentMethodForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': False},
            )
        payment_method = form.save(commit=False)
        payment_method.created_by = request.user
        payment_method.save()
        messages.success(request, 'Payment method created successfully.')
        return redirect(reverse('payment_method_list'))


class PaymentMethodUpdateView(LoginRequiredMixin, View):
    template_name = 'payment/methods/method_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        method = get_object_or_404(PaymentMethod, pk=pk)
        return render(
            request,
            self.template_name,
            {'form': PaymentMethodForm(instance=method), 'is_edit': True, 'method': method},
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        method = get_object_or_404(PaymentMethod, pk=pk)
        form = PaymentMethodForm(request.POST, request.FILES, instance=method)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'method': method},
            )
        obj = form.save(commit=False)
        obj.updated_by = request.user
        obj.save()
        messages.success(request, 'Payment method updated successfully.')
        return redirect(reverse('payment_method_list'))


class PaymentMethodToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        method = get_object_or_404(PaymentMethod, pk=pk)
        method.is_active = not method.is_active
        method.save(update_fields=['is_active'])
        messages.success(request, 'Payment method status updated successfully.')
        return redirect(reverse('payment_method_list'))


class PaymentMethodDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        messages.error(
            request,
            'Payment methods cannot be deleted. Deactivate instead.',
        )
        return redirect(reverse('payment_method_list'))

    def get(self, request, pk):
        messages.error(
            request,
            'Payment methods cannot be deleted. Deactivate instead.',
        )
        return redirect(reverse('payment_method_list'))


class CommGatewayListView(LoginRequiredMixin, View):
    template_name = 'comm/gateways/gateway_list.html'

    def get(self, request):
        qs = CommGateway.objects.filter(is_deleted=False)

        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(provider_name__icontains=q)
                | Q(host_url__icontains=q)
                | Q(sender_id__icontains=q)
            )

        type_filter = request.GET.get('gateway_type', 'All').strip() or 'All'
        if type_filter != 'All':
            qs = qs.filter(gateway_type=type_filter)

        status_filter = request.GET.get('status', 'All').strip() or 'All'
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort = request.GET.get('sort', 'updated')
        direction = request.GET.get('dir', 'desc')
        sort_mapping = {
            'rank': 'default_rank',
            'type': 'gateway_type',
            'provider': 'provider_name',
            'host': 'host_url',
            'port': 'port',
            'sender': 'sender_id',
            'encryption': 'encryption_type',
            'status': 'is_active',
            'updated': 'updated_at',
        }
        qs = qs.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('updated_at').desc(),
            )
        )
        order_by_field = sort_mapping.get(sort, 'updated_at')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), 'provider_name')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'provider_name')

        paginator = Paginator(qs, 10)
        gateways = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = gateways.start_index()
        for offset, gateway in enumerate(gateways.object_list):
            # Show descending list ID so newest appears with highest number.
            gateway.list_rank = total_count - (start_index + offset) + 1
        return render(
            request,
            self.template_name,
            {
                'gateways': gateways,
                'search_query': q,
                'type_filter': type_filter,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'active_email_id': CommGateway.objects.filter(
                    is_deleted=False, gateway_type='Email', is_active=True
                ).values_list('gateway_id', flat=True).first(),
                'active_sms_id': CommGateway.objects.filter(
                    is_deleted=False, gateway_type='SMS', is_active=True
                ).values_list('gateway_id', flat=True).first(),
            },
        )


class CommGatewayCreateView(LoginRequiredMixin, View):
    template_name = 'comm/gateways/gateway_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(request, self.template_name, {'form': CommGatewayForm(), 'is_edit': False})

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = CommGatewayForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form, 'is_edit': False})
        instance = form.save(commit=False)
        instance.updated_by = request.user
        instance.save() # Model now handles singularity
        messages.success(request, 'Communication gateway saved successfully.')
        return redirect(reverse('comm_gateway_list'))


class CommGatewayUpdateView(LoginRequiredMixin, View):
    template_name = 'comm/gateways/gateway_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        gateway = get_object_or_404(CommGateway, pk=pk)
        form = CommGatewayForm(instance=gateway)
        form.initial['password_secret'] = '********'
        return render(
            request,
            self.template_name,
            {'form': form, 'is_edit': True, 'gateway': gateway},
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        gateway = get_object_or_404(CommGateway, pk=pk)
        post_data = request.POST.copy()
        if (post_data.get('password_secret') or '').strip() == '********':
            post_data['password_secret'] = gateway.password_secret
        form = CommGatewayForm(post_data, instance=gateway)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'gateway': gateway},
            )
        instance = form.save(commit=False)
        instance.updated_by = request.user
        instance.save() # Model now handles singularity
        messages.success(request, 'Communication gateway updated successfully.')
        return redirect(reverse('comm_gateway_list'))


class CommGatewayToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        gateway = get_object_or_404(CommGateway, pk=pk)
        gateway.is_active = not gateway.is_active
        gateway.updated_by = request.user
        gateway.save() # Model now handles singularity
        messages.success(request, 'Gateway status updated successfully.')
        return redirect(reverse('comm_gateway_list'))


class CommGatewayDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('comm_gateway_list'))
        gateway = get_object_or_404(CommGateway, pk=pk)
        if gateway.is_deleted:
            messages.info(request, 'Gateway is already removed.')
            return redirect(reverse('comm_gateway_list'))
        gateway.is_deleted = True
        gateway.updated_by = request.user
        gateway.save(update_fields=['is_deleted', 'updated_by'])
        messages.success(request, 'Gateway removed successfully.')
        return redirect(reverse('comm_gateway_list'))

    def get(self, request, pk):
        return self.post(request, pk)


class CommGatewayTestConnectionView(LoginRequiredMixin, View):
    """
    Spec CP-PCS-P6 §5.1.3: UI must include a "Test Connection" button 
    that attempts to send a dummy message using the entered credentials 
    before saving.
    """
    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return JsonResponse({'success': False, 'message': 'Permission denied.'})

        try:
            g_type = request.POST.get('gateway_type')
            host = request.POST.get('host_url')
            port = request.POST.get('port')
            user = request.POST.get('username_key')
            password = request.POST.get('password_secret')
            enc_type = request.POST.get('encryption_type')
            sender = request.POST.get('sender_id')
            provider_name = request.POST.get('provider_name')
            test_recipient_phone = (request.POST.get('test_recipient_phone') or '').strip()

            # If password is masked '********', use the existing password from record (if ID supplied)
            record_id = request.POST.get('gateway_id')
            if record_id and password == '********':
                existing = CommGateway.objects.filter(pk=record_id).first()
                if existing:
                    password = existing.password_secret

            if g_type == 'Email':
                if not all([host, user, password]):
                    return JsonResponse({'success': False, 'message': 'Incomplete SMTP credentials (host, username, password required).'})
                enc = (enc_type or 'TLS').strip() or 'TLS'
                try:
                    p = int(port) if str(port).strip() else None
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'Port must be a number.'})
                if p is None:
                    p = 465 if enc == 'SSL' else 587

                host_clean = (host or '').strip()
                test_recipient = (getattr(request.user, 'email', '') or '').strip()
                if not test_recipient:
                    return JsonResponse({
                        'success': False,
                        'message': 'Connection check requires your admin account to have an email address for dummy test delivery.',
                    })
                from_email = _normalize_from_email_header(
                    (sender or '').strip() or user.strip(),
                    user.strip(),
                )
                envelope_from = _extract_sender_address(from_email, user.strip())
                try:
                    # Match runtime send path in communication_helpers / DatabaseEmailBackend
                    if enc == 'SSL':
                        server = smtplib.SMTP_SSL(host_clean, p, timeout=15)
                    else:
                        server = smtplib.SMTP(host_clean, p, timeout=15)
                        if enc == 'TLS':
                            server.ehlo()
                            server.starttls()
                            server.ehlo()
                    server.login(user, password)
                    test_subject = 'iRoad Gateway Test Message'
                    test_context = {
                        'provider_name': (request.POST.get('provider_name') or '').strip() or 'N/A',
                        'host': host_clean,
                        'recipient_email': test_recipient,
                    }
                    html_body = render_to_string('comm/emails/gateway_test_message.html', test_context)
                    text_body = strip_tags(html_body)
                    mime_msg = MIMEMultipart('alternative')
                    mime_msg['Subject'] = test_subject
                    mime_msg['From'] = from_email
                    mime_msg['To'] = test_recipient
                    mime_msg.attach(MIMEText(text_body, 'plain', 'utf-8'))
                    mime_msg.attach(MIMEText(html_body, 'html', 'utf-8'))
                    server.sendmail(envelope_from, [test_recipient], mime_msg.as_string())
                    try:
                        server.quit()
                    except Exception:
                        pass
                    return JsonResponse({
                        'success': True,
                        'message': f'SMTP connection successful and testing email sent to {test_recipient}.',
                    })
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Connection failed: {str(e)}'})

            elif g_type == 'SMS':
                if not all([host, user, password, sender]):
                    return JsonResponse({
                        'success': False,
                        'message': (
                            'Incomplete SMS credentials (Host/API URL, API Key, '
                            'Secret, and Sender ID are required).'
                        ),
                    })

                if not test_recipient_phone:
                    test_recipient_phone = (getattr(request.user, 'phone_number', '') or '').strip()
                if not test_recipient_phone:
                    return JsonResponse({
                        'success': False,
                        'message': (
                            'Enter a Test Recipient Phone, or set your admin phone number '
                            'to run SMS test delivery.'
                        ),
                    })

                from types import SimpleNamespace
                from superadmin.communication_helpers import send_sms_http_gateway

                # Build unsaved in-memory gateway for testing entered credentials.
                test_gateway = SimpleNamespace(
                    provider_name=(provider_name or '').strip(),
                    host_url=(host or '').strip(),
                    username_key=(user or '').strip(),
                    password_secret=password,
                    sender_id=(sender or '').strip(),
                )
                test_message = 'iRoad SMS gateway test: configuration validated.'
                try:
                    send_sms_http_gateway(
                        test_gateway,
                        test_recipient_phone,
                        test_message,
                        trigger_source='Gateway Test: SMS',
                    )
                except Exception as exc:
                    return JsonResponse({
                        'success': False,
                        'message': f'SMS test failed: {str(exc)}',
                    })
                return JsonResponse({
                    'success': True,
                    'message': (
                        f'SMS test successful. Dummy message sent to '
                        f'{test_recipient_phone}.'
                    ),
                })

            return JsonResponse({'success': False, 'message': 'Invalid gateway type.'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Server error: {str(e)}'})


class NotificationTemplateListView(LoginRequiredMixin, View):
    template_name = 'comm/templates/template_list.html'

    def get(self, request):
        ensure_default_notification_templates(
            created_by=request.user if getattr(request.user, 'is_authenticated', False) else None
        )
        search_query = request.GET.get('q', '').strip()
        channel_filter = request.GET.get('channel', 'All')
        category_filter = request.GET.get('category', 'All')
        status_filter = request.GET.get('status', 'All')

        qs = NotificationTemplate.objects.filter(is_deleted=False)
        if search_query:
            qs = qs.filter(template_name__icontains=search_query)
        if channel_filter in ['Email', 'SMS']:
            qs = qs.filter(channel_type=channel_filter)
        if category_filter in ['Transactional', 'Promotional']:
            qs = qs.filter(category=category_filter)
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        qs = qs.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc(),
            )
        )

        sort = request.GET.get('sort', 'created_at')
        direction = request.GET.get('dir', 'desc')
        sort_mapping = {
            'rank': 'default_rank',
            'name': 'template_name',
            'channel': 'channel_type',
            'category': 'category',
            'subject': 'subject_en',
            'status': 'is_active',
            'created_at': 'created_at',
        }
        order_by_field = sort_mapping.get(sort, 'created_at')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-template_name')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'template_name')

        paginator = Paginator(qs, 10)
        templates = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = templates.start_index()
        for offset, template_obj in enumerate(templates.object_list):
            # Show descending list ID so newest appears with highest number.
            template_obj.list_rank = total_count - (start_index + offset) + 1
        return render(
            request,
            self.template_name,
            {
                'templates_page': templates,
                'search_query': search_query,
                'channel_filter': channel_filter,
                'category_filter': category_filter,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
            },
        )


class NotificationTemplateCreateView(LoginRequiredMixin, View):
    template_name = 'comm/templates/template_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {'form': NotificationTemplateForm(), 'is_edit': False},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = NotificationTemplateForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form, 'is_edit': False})
        obj = form.save(commit=False)
        obj.created_by = request.user
        obj.save()
        messages.success(request, 'Notification template created successfully.')
        return redirect(reverse('notif_template_list'))


class NotificationTemplateUpdateView(LoginRequiredMixin, View):
    template_name = 'comm/templates/template_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        template_obj = get_object_or_404(NotificationTemplate, pk=pk)
        form = NotificationTemplateForm(instance=template_obj)
        return render(
            request,
            self.template_name,
            {'form': form, 'is_edit': True, 'template_obj': template_obj},
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        template_obj = get_object_or_404(NotificationTemplate, pk=pk)
        form = NotificationTemplateForm(request.POST, instance=template_obj)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'template_obj': template_obj},
            )
        form.save()
        messages.success(request, 'Notification template updated successfully.')
        return redirect(reverse('notif_template_list'))


class NotificationTemplateToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        template_obj = get_object_or_404(NotificationTemplate, pk=pk)
        is_in_active_mapping = EventMapping.objects.filter(
            is_active=True,
        ).filter(
            Q(primary_template=template_obj) | Q(fallback_template=template_obj)
        ).exists()
        # TODO: EventMapping may break if template deactivated
        template_obj.is_active = not template_obj.is_active
        template_obj.save(update_fields=['is_active'])
        if is_in_active_mapping and not template_obj.is_active:
            messages.warning(
                request,
                'Template is used in active event mapping. Deactivated with caution.',
            )
        else:
            messages.success(request, 'Template status updated successfully.')
        return redirect(reverse('notif_template_list'))


class NotificationTemplateDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'Access denied: root admin only.')
            return redirect(reverse('notif_template_list'))
        template_obj = get_object_or_404(NotificationTemplate, pk=pk)
        if template_obj.is_deleted:
            messages.info(request, 'Template is already removed.')
            return redirect(reverse('notif_template_list'))
        template_obj.is_deleted = True
        template_obj.save(update_fields=['is_deleted'])
        messages.success(request, 'Template removed from successfully.')
        return redirect(reverse('notif_template_list'))

    def get(self, request, pk):
        return self.post(request, pk)


def get_mock_preview_context():
    """Returns a dictionary of mock data for notification template preview."""
    return {
        'admin_user': {'first_name': 'John', 'last_name': 'Doe'},
        'company_name': 'Global Logistics Corp',
        'tenant': {
            'primary_email': 'admin@globallogistics.com',
            'tenant_id': 'TNT-98721-XQ',
        },
        'portal_bootstrap_password': 'S3cure!Password@2026',
        'api_bridge_key': 'br_live_51P2kL9H2j8mN4v6xYzQ1w2e3r4t5y6u',
        'reset_url': 'http://127.0.0.1:8000/new-password/mock-token/',
        'invite_url': 'http://127.0.0.1:8000/set-password/mock-token/',
        'portal_login_url': 'http://subdomain.iroad.com/login/',
        'name': 'Sarah Smith',
        'email': 'sarah@example.com',
        'password': 'TemporaryPassword123!',
        'login_url': 'http://127.0.0.1:8000/login/',
        'recipient_email': 'test@example.com',
        'sent_at': '2026-04-16 11:15:00',
        'gateway_name': 'Twilio SMS Gateway',
        'otp_code': '483920',
        'otp': '483920',
        'user_name': 'John Doe',
    }


@method_decorator(xframe_options_sameorigin, name='get')
class NotificationTemplatePreviewView(LoginRequiredMixin, View):
    def get(self, request, pk):
        """
        Renders a full HTML preview of the notification template.
        Supports 'lang' query param (en|ar) and device simulation context (todo).
        """
        template_obj = get_object_or_404(NotificationTemplate, pk=pk)
        lang = request.GET.get('lang', 'en').lower()

        subject = template_obj.subject_ar if lang == 'ar' else template_obj.subject_en
        body_html = template_obj.body_ar if lang == 'ar' else template_obj.body_en

        # Render with mock data + dynamic branding context
        def _render_preview_html(raw_html, context):
            """Best-effort renderer: render Django template syntax, and if it fails,
            degrade gracefully by stripping template tags so raw braces are not shown.
            """
            source = raw_html or ""
            try:
                return Template(source).render(Context(context))
            except Exception as err:
                logger.warning(f"Preview template render fallback triggered: {err}")

                without_blocks = re.sub(r"\{%\s*.*?%\}", "", source, flags=re.DOTALL)

                def _replace_var(match):
                    expr = (match.group(1) or "").strip()
                    key = expr.split("|", 1)[0].strip()
                    value = context.get(key, "")
                    return "" if value is None else str(value)

                return re.sub(r"\{\{\s*(.*?)\s*\}\}", _replace_var, without_blocks, flags=re.DOTALL)

        mock_ctx = get_mock_preview_context()
        try:
            from superadmin.communication_helpers import _merge_template_context
            mock_ctx = _merge_template_context(get_mock_preview_context())
            body_html = _render_preview_html(body_html, mock_ctx)
            subject = _render_preview_html(subject, mock_ctx)

            # Dynamic fix for legacy broken logo URL in hardcoded template bodies
            old_logo = "https://iroad-assets.s3.amazonaws.com/logo.png"
            new_logo = "https://ui-avatars.com/api/?name=iR&background=4f46e5&color=fff&rounded=true&size=128&bold=true"
            if old_logo in body_html:
                body_html = body_html.replace(old_logo, new_logo)

        except Exception as e:
            logger.error(f"Error rendering mock preview: {e}")

        if template_obj.channel_type == 'Email':
            from superadmin.communication_helpers import (
                _merge_template_context,
                _wrap_email_body,
            )

            body_lower = (body_html or '').lower()
            is_full_email_document = '<html' in body_lower and '<body' in body_lower

            # Check if the body already contains the system wrapper.
            # If not (e.g. user manually edited and stripped it), re-wrap it.
            if is_full_email_document:
                wrapped_content = body_html or ''
            elif 'email-wrapper' not in (body_html or ''):
                wrapped_content = _wrap_email_body(
                    inner_html=body_html or '<p>No content provided.</p>',
                    email_title=subject or 'iRoad Logistics',
                    use_rtl=(lang == 'ar'),
                )
            else:
                # Still need to handle RTL if the body is already wrapped but doesn't have dir="rtl"
                wrapped_content = body_html or ''
                if lang == 'ar' and 'dir="rtl"' not in wrapped_content:
                    wrapped_content = wrapped_content.replace('<html lang="en">', '<html lang="ar" dir="rtl">')

            # Resolve wrapper branding placeholders (company logo/name initials).
            wrapped_content = _render_preview_html(
                wrapped_content,
                _merge_template_context(mock_ctx),
            )

            return HttpResponse(wrapped_content)
        else:
            # SMS Preview: Simple styled box
            from django.template.loader import render_to_string
            context = {
                'body_text': body_html,
                'lang': lang,
                'is_sms': True,
            }
            # Create a simple container for SMS preview
            sms_html = render_to_string('comm/templates/sms_preview_wrapper.html', context)
            return HttpResponse(sms_html)


class EventMappingListView(LoginRequiredMixin, View):
    template_name = 'comm/events/event_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All').strip() or 'All'
        channel_filter = request.GET.get('channel', 'All').strip() or 'All'
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'desc')

        qs = EventMapping.objects.select_related(
            'primary_template', 'fallback_template'
        ).order_by('system_event')

        # Construct rows of both mapped and unmapped events for the template
        mappings_dict = {m.system_event: m for m in qs}
        all_rows = []
        for base_id, (code, label) in enumerate(EventMapping.SYSTEM_EVENT_CHOICES, start=1):
            if code in mappings_dict:
                mapping = mappings_dict[code]
                all_rows.append({
                    'row_id': base_id,
                    'row_type': 'mapped',
                    'mapping': mapping,
                    'event_label': mapping.get_system_event_display(),
                    'primary_channel_value': mapping.primary_channel or '',
                    'primary_template_name': (
                        mapping.primary_template.template_name
                        if mapping.primary_template else ''
                    ),
                    'fallback_channel_value': mapping.fallback_channel or '',
                    'fallback_template_name': (
                        mapping.fallback_template.template_name
                        if mapping.fallback_template else ''
                    ),
                    'status_label': 'Active' if mapping.is_active else 'Inactive',
                })
            else:
                all_rows.append({
                    'row_id': base_id,
                    'row_type': 'unmapped',
                    'event': {'code': code, 'label': label},
                    'event_label': label,
                    'primary_channel_value': '',
                    'primary_template_name': '',
                    'fallback_channel_value': '',
                    'fallback_template_name': '',
                    'status_label': 'Not Configured',
                })

        if search_query:
            search_l = search_query.lower()
            all_rows = [
                r for r in all_rows
                if search_l in (r.get('event_label') or '').lower()
                or search_l in (r.get('primary_template_name') or '').lower()
                or search_l in (r.get('fallback_template_name') or '').lower()
            ]

        if channel_filter in ['Email', 'SMS', 'Not Configured']:
            if channel_filter == 'Not Configured':
                all_rows = [
                    r for r in all_rows if r.get('primary_channel_value') == ''
                ]
            else:
                all_rows = [
                    r for r in all_rows
                    if r.get('primary_channel_value') == channel_filter
                ]

        if status_filter in ['Active', 'Inactive']:
            all_rows = [
                r for r in all_rows if r.get('status_label') == status_filter
            ]

        sort_mapping = {
            'rank': lambda r: r.get('row_id') or 0,
            'event': lambda r: (r.get('event_label') or '').lower(),
            'primary_channel': lambda r: (r.get('primary_channel_value') or '').lower(),
            'primary_template': lambda r: (r.get('primary_template_name') or '').lower(),
            'fallback_channel': lambda r: (r.get('fallback_channel_value') or '').lower(),
            'fallback_template': lambda r: (r.get('fallback_template_name') or '').lower(),
            'status': lambda r: (r.get('status_label') or '').lower(),
        }
        key_fn = sort_mapping.get(sort, sort_mapping['event'])
        all_rows = sorted(all_rows, key=key_fn, reverse=(direction == 'desc'))

        paginator = Paginator(all_rows, 10)
        page_number = request.GET.get('page', 1)
        page_rows = paginator.get_page(page_number)
        total_count = len(all_rows)
        start_index = page_rows.start_index()
        for offset, row in enumerate(page_rows.object_list):
            # Show descending list ID so newest/highest rows appear first.
            row['list_rank'] = total_count - (start_index + offset) + 1

        return render(
            request,
            self.template_name,
            {
                'page_rows': page_rows,
                'search_query': search_query,
                'status_filter': status_filter,
                'channel_filter': channel_filter,
                'current_sort': sort,
                'current_dir': direction,
            },
        )


class EventMappingCreateView(LoginRequiredMixin, View):
    template_name = 'comm/events/event_form.html'

    @staticmethod
    def _channel_templates():
        return list(
            NotificationTemplate.objects.filter(
                is_active=True,
                channel_type__in=['Email', 'SMS'],
            )
            .order_by('template_name')
            .values('template_id', 'template_name', 'channel_type')
        )

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        
        initial = {}
        target_event = request.GET.get('system_event')
        if target_event:
            initial['system_event'] = target_event

        return render(request, self.template_name, {
            'form': EventMappingForm(initial=initial),
            'is_edit': False,
            'channel_templates': self._channel_templates(),
        })


    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = EventMappingForm(request.POST)
        if form.is_valid():
            event_code = form.cleaned_data.get('system_event')
            if EventMapping.objects.filter(system_event=event_code).exists():
                form.add_error(
                    'system_event',
                    'A mapping already exists for this event. Edit it instead.',
                )
            else:
                obj = form.save(commit=False)
                obj.updated_by = request.user
                obj.save()
                messages.success(request, 'Event mapping created successfully.')
                return redirect(reverse('event_mapping_list'))
        return render(request, self.template_name, {
            'form': form,
            'is_edit': False,
            'channel_templates': self._channel_templates(),
        })


class EventMappingUpdateView(LoginRequiredMixin, View):
    template_name = 'comm/events/event_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        mapping = get_object_or_404(EventMapping, pk=pk)
        return render(
            request,
            self.template_name,
            {
                'form': EventMappingForm(instance=mapping),
                'is_edit': True,
                'mapping': mapping,
                'channel_templates': EventMappingCreateView._channel_templates(),
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        mapping = get_object_or_404(EventMapping, pk=pk)
        form = EventMappingForm(request.POST, instance=mapping)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': True,
                    'mapping': mapping,
                    'channel_templates': EventMappingCreateView._channel_templates(),
                },
            )
        obj = form.save(commit=False)
        obj.updated_by = request.user
        obj.save()
        messages.success(request, 'Event mapping updated successfully.')
        return redirect(reverse('event_mapping_list'))


class EventMappingToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        mapping = get_object_or_404(EventMapping, pk=pk)
        mapping.is_active = not mapping.is_active
        mapping.updated_by = request.user
        mapping.save(update_fields=['is_active', 'updated_by', 'updated_at'])
        messages.success(request, 'Event mapping status updated successfully.')
        return redirect(reverse('event_mapping_list'))


class PushNotificationListView(LoginRequiredMixin, View):
    template_name = 'comm/push/push_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        trigger_mode = request.GET.get('trigger_mode', 'All')
        dispatch_status = request.GET.get('dispatch_status', 'All')
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'asc')

        qs = PushNotification.objects.all()

        if search_query:
            qs = qs.filter(
                Q(internal_name__icontains=search_query)
                | Q(linked_event__icontains=search_query)
                | Q(target_audience__icontains=search_query)
            )

        if trigger_mode in ['Manual_Broadcast', 'System_Event']:
            qs = qs.filter(trigger_mode=trigger_mode)
        if dispatch_status in ['Draft', 'Scheduled', 'Completed']:
            qs = qs.filter(dispatch_status=dispatch_status)

        # Keep a lightweight stable row-id source (created_at descending), but
        # avoid costly window functions on every request.
        base_rank_ids = list(
            qs.order_by('-created_at').values_list('notification_id', flat=True)
        )
        rank_map = {nid: idx for idx, nid in enumerate(base_rank_ids, start=1)}

        if sort == 'audience':
            qs = qs.annotate(
                audience_event_sort=Case(
                    When(trigger_mode='System_Event', then=F('linked_event')),
                    default=F('target_audience'),
                    output_field=CharField(),
                ),
            )
            order_by_field = 'audience_event_sort'
        elif sort == 'name':
            order_by_field = 'internal_name'
        elif sort == 'trigger':
            order_by_field = 'trigger_mode'
        elif sort == 'scheduled':
            order_by_field = 'scheduled_at'
        elif sort == 'status':
            order_by_field = 'dispatch_status'
        else:
            # rank: base created ordering
            order_by_field = 'created_at'
            direction = 'desc' if direction == 'desc' else 'asc'

        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-created_at')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), '-created_at')

        paginator = Paginator(qs, 10)
        push_items = paginator.get_page(request.GET.get('page', 1))
        for row in push_items:
            row.default_rank = rank_map.get(row.notification_id, 0)
        return render(
            request,
            self.template_name,
            {
                'push_items': push_items,
                'search_query': search_query,
                'trigger_mode_filter': trigger_mode,
                'dispatch_status_filter': dispatch_status,
                'current_sort': sort,
                'current_dir': direction,
            },
        )


class PushNotificationCreateView(LoginRequiredMixin, View):
    template_name = 'comm/push/push_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(request, self.template_name, {'form': PushNotificationForm(), 'is_edit': False})

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = PushNotificationForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form, 'is_edit': False})
        obj = form.save(commit=False)
        obj.created_by = request.user
        obj.save()
        from superadmin.push_helpers import queue_push_notification

        if obj.trigger_mode == 'Manual_Broadcast' and obj.dispatch_status == 'Scheduled':
            queue_push_notification(obj)
        messages.success(request, 'Push notification created successfully.')
        return redirect(reverse('push_notif_list'))


class PushNotificationUpdateView(LoginRequiredMixin, View):
    template_name = 'comm/push/push_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        push_item = get_object_or_404(PushNotification, pk=pk)
        if push_item.dispatch_status == 'Completed':
            messages.error(request, 'Completed push notifications cannot be edited.')
            return redirect(reverse('push_notif_list'))
        return render(
            request,
            self.template_name,
            {'form': PushNotificationForm(instance=push_item), 'is_edit': True, 'push_item': push_item},
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        push_item = get_object_or_404(PushNotification, pk=pk)
        if push_item.dispatch_status == 'Completed':
            messages.error(request, 'Completed push notifications cannot be edited.')
            return redirect(reverse('push_notif_list'))
        form = PushNotificationForm(request.POST, instance=push_item)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'push_item': push_item},
            )
        updated = form.save()
        from superadmin.push_helpers import queue_push_notification

        if updated.trigger_mode == 'Manual_Broadcast' and updated.dispatch_status == 'Scheduled':
            queue_push_notification(updated)
        messages.success(request, 'Push notification updated successfully.')
        return redirect(reverse('push_notif_list'))


class SystemBannerListView(LoginRequiredMixin, View):
    template_name = 'comm/banners/banner_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        severity_filter = request.GET.get('severity', 'All')
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'valid_from')
        direction = request.GET.get('dir', 'desc')

        qs = SystemBanner.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('valid_from').desc(),
            ),
        )

        if search_query:
            qs = qs.filter(
                Q(title_en__icontains=search_query)
                | Q(message_en__icontains=search_query)
            )
        if severity_filter in ['Info', 'Warning', 'Critical']:
            qs = qs.filter(severity=severity_filter)
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'title': 'title_en',
            'severity': 'severity',
            'dismissible': 'is_dismissible',
            'valid_from': 'valid_from',
            'valid_until': 'valid_until',
            'status': 'is_active',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-valid_from')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), '-valid_from')

        paginator = Paginator(qs, 10)
        banners = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = banners.start_index()
        for offset, banner in enumerate(banners.object_list):
            # Show descending list ID so newest appears with highest number.
            banner.list_rank = total_count - (start_index + offset) + 1
        return render(
            request,
            self.template_name,
            {
                'banners': banners,
                'search_query': search_query,
                'severity_filter': severity_filter,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
            },
        )


class SystemBannerCreateView(LoginRequiredMixin, View):
    template_name = 'comm/banners/banner_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(request, self.template_name, {'form': SystemBannerForm(), 'is_edit': False})

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = SystemBannerForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form, 'is_edit': False})
        form.save()
        messages.success(request, 'System banner created successfully.')
        return redirect(reverse('banner_list'))


class SystemBannerUpdateView(LoginRequiredMixin, View):
    template_name = 'comm/banners/banner_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        banner = get_object_or_404(SystemBanner, pk=pk)
        return render(
            request,
            self.template_name,
            {'form': SystemBannerForm(instance=banner), 'is_edit': True, 'banner': banner},
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        banner = get_object_or_404(SystemBanner, pk=pk)
        form = SystemBannerForm(request.POST, instance=banner)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'banner': banner},
            )
        form.save()
        messages.success(request, 'System banner updated successfully.')
        return redirect(reverse('banner_list'))


class SystemBannerToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        banner = get_object_or_404(SystemBanner, pk=pk)
        banner.is_active = not banner.is_active
        banner.save(update_fields=['is_active'])
        messages.success(request, 'System banner status updated successfully.')
        return redirect(reverse('banner_list'))


class InternalAlertRouteListView(LoginRequiredMixin, View):
    template_name = 'comm/alerts/alert_list.html'

    def get(self, request):
        search_query = request.GET.get('q', '').strip()
        trigger_filter = request.GET.get('trigger_event', 'All')
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'asc')

        qs = InternalAlertRoute.objects.select_related('notify_role').annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('trigger_event').asc(),
            ),
            status_sort=Case(
                When(is_active=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
        )

        if search_query:
            qs = qs.filter(
                Q(notify_role__role_name_en__icontains=search_query)
                | Q(notify_custom_email__icontains=search_query)
                | Q(trigger_event__icontains=search_query)
            )

        trigger_codes = [code for code, _ in InternalAlertRoute.TRIGGER_EVENT_CHOICES]
        if trigger_filter in trigger_codes:
            qs = qs.filter(trigger_event=trigger_filter)

        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': 'default_rank',
            'trigger': 'trigger_event',
            'role': 'notify_role__role_name_en',
            'email': 'notify_custom_email',
            'status': 'status_sort',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-default_rank')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'default_rank')

        paginator = Paginator(qs, 10)
        routes = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = routes.start_index()
        for offset, route in enumerate(routes.object_list):
            # Show descending list ID so top rows have higher numbers.
            route.list_rank = total_count - (start_index + offset) + 1
        return render(
            request,
            self.template_name,
            {
                'routes': routes,
                'search_query': search_query,
                'trigger_filter': trigger_filter,
                'status_filter': status_filter,
                'trigger_choices': InternalAlertRoute.TRIGGER_EVENT_CHOICES,
                'current_sort': sort,
                'current_dir': direction,
            },
        )


class InternalAlertRouteCreateView(LoginRequiredMixin, View):
    template_name = 'comm/alerts/alert_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(request, self.template_name, {'form': InternalAlertRouteForm(), 'is_edit': False})

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = InternalAlertRouteForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form, 'is_edit': False})
        form.save()
        messages.success(request, 'Alert route created successfully.')
        return redirect(reverse('alert_route_list'))


class InternalAlertRouteUpdateView(LoginRequiredMixin, View):
    template_name = 'comm/alerts/alert_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        route = get_object_or_404(InternalAlertRoute, pk=pk)
        return render(
            request,
            self.template_name,
            {'form': InternalAlertRouteForm(instance=route), 'is_edit': True, 'route': route},
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        route = get_object_or_404(InternalAlertRoute, pk=pk)
        form = InternalAlertRouteForm(request.POST, instance=route)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {'form': form, 'is_edit': True, 'route': route},
            )
        form.save()
        messages.success(request, 'Alert route updated successfully.')
        return redirect(reverse('alert_route_list'))


class InternalAlertRouteToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        route = get_object_or_404(InternalAlertRoute, pk=pk)
        route.is_active = not route.is_active
        route.save(update_fields=['is_active'])
        messages.success(request, 'Alert route status updated successfully.')
        return redirect(reverse('alert_route_list'))

class InternalAlertNotificationReadView(LoginRequiredMixin, View):
    def post(self, request, pk):
        notif = get_object_or_404(
            InternalAlertNotification,
            notification_id=pk,
            admin_user=request.user,
        )
        if not notif.is_read:
            notif.is_read = True
            notif.read_at = timezone.now()
            notif.save(update_fields=['is_read', 'read_at'])
        next_url = (request.POST.get('next') or '').strip()
        if next_url:
            return redirect(next_url)
        return redirect('dashboard')


class InternalAlertNotificationReadAllView(LoginRequiredMixin, View):
    def post(self, request):
        now = timezone.now()
        InternalAlertNotification.objects.filter(
            admin_user=request.user,
            is_read=False,
        ).update(is_read=True, read_at=now)
        next_url = (request.POST.get('next') or '').strip()
        if next_url:
            return redirect(next_url)
        return redirect('dashboard')


class InternalAlertNotificationClearAllView(LoginRequiredMixin, View):
    def post(self, request):
        InternalAlertNotification.objects.filter(
            admin_user=request.user,
        ).delete()
        next_url = (request.POST.get('next') or '').strip()
        if next_url:
            return redirect(next_url)
        return redirect('dashboard')


class CommLogListView(LoginRequiredMixin, View):
    template_name = 'comm/logs/comm_log_list.html'

    def get(self, request):
        query = request.GET.get('q', '').strip()
        channel_filter = request.GET.get('channel', 'All')
        status_filter = request.GET.get('status', 'All')
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        sort = request.GET.get('sort', 'rank')
        direction = request.GET.get('dir', 'desc')

        qs = CommLog.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('dispatched_at').desc(),
            ),
            status_sort=Case(
                When(delivery_status='Sent', then=Value(1)),
                When(delivery_status='Failed', then=Value(2)),
                default=Value(99),
                output_field=IntegerField(),
            ),
        )
        qs = qs.exclude(delivery_status='Bounced')
        if query:
            qs = qs.filter(
                Q(recipient__icontains=query)
                | Q(trigger_source__icontains=query)
                | Q(client_id__icontains=query)
                | Q(error_details__icontains=query)
            )
        if channel_filter in ['Email', 'SMS', 'Push']:
            qs = qs.filter(channel_type=channel_filter)
        if status_filter in ['Sent', 'Failed']:
            qs = qs.filter(delivery_status=status_filter)
        if date_from:
            parsed_from = parse_date(date_from)
            if parsed_from:
                qs = qs.filter(dispatched_at__date__gte=parsed_from)
        if date_to:
            parsed_to = parse_date(date_to)
            if parsed_to:
                qs = qs.filter(dispatched_at__date__lte=parsed_to)

        sort_mapping = {
            'rank': 'default_rank',
            'recipient': 'recipient',
            'channel': 'channel_type',
            'trigger': 'trigger_source',
            'status': 'status_sort',
            'error': 'error_details',
            'dispatched': 'dispatched_at',
        }
        order_by_field = sort_mapping.get(sort, 'default_rank')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc(nulls_last=True), '-default_rank')
        else:
            qs = qs.order_by(F(order_by_field).asc(nulls_first=True), 'default_rank')

        paginator = Paginator(qs, 10)
        logs = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = logs.start_index()
        for offset, log in enumerate(logs.object_list):
            # Show descending list ID so top rows have higher numbers.
            log.list_rank = total_count - (start_index + offset) + 1
        return render(
            request,
            self.template_name,
            {
                'logs': logs,
                'search_query': query,
                'channel_filter': channel_filter,
                'status_filter': status_filter,
                'date_from': date_from,
                'date_to': date_to,
                'current_sort': sort,
                'current_dir': direction,
            },
        )


_SUSPEND_ACCOUNT_STATUSES = (
    'Suspended_Billing',
    'Suspended_Violation',
)


class TenantListView(LoginRequiredMixin, View):
    template_name = 'crm/tenants/tenant_list.html'

    def get(self, request):
        # Annotate with a stable rank based on registration date
        qs = TenantProfile.objects.filter(is_deleted=False).annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('registered_at').desc()
            )
        ).select_related(
            'country', 'current_plan', 'assigned_sales_rep'
        )

        search = request.GET.get('q', '').strip()
        status_filter = request.GET.get('account_status', 'All')
        rep_filter = request.GET.get('assigned_sales_rep', '').strip()
        plan_filter = request.GET.get('current_plan', '').strip()

        if search:
            qs = qs.filter(
                Q(company_name__icontains=search)
                | Q(primary_email__icontains=search)
            )
        codes = [c[0] for c in TenantProfile.STATUS_CHOICES]
        if status_filter == 'Suspended':
            qs = qs.filter(account_status__in=['Suspended_Billing', 'Suspended_Violation'])
        elif status_filter in codes:
            qs = qs.filter(account_status=status_filter)
        if rep_filter:
            try:
                uuid.UUID(str(rep_filter))
                qs = qs.filter(assigned_sales_rep_id=rep_filter)
            except ValueError:
                pass
        if plan_filter:
            try:
                uuid.UUID(str(plan_filter))
                qs = qs.filter(current_plan_id=plan_filter)
            except ValueError:
                pass

        # Advanced Sorting Logic
        sort = request.GET.get('sort', 'updated')
        direction = request.GET.get('dir', 'desc')
        
        sort_mapping = {
            'rank': 'default_rank',
            'company': 'company_name',
            'email': 'primary_email',
            'country': 'country__name_en',
            'plan': 'current_plan__plan_name_en',
            'expiry': 'subscription_expiry_date',
            'updated': 'updated_at',
            'status': 'account_status',
            'wallet': 'wallet_balance',
        }
        
        order_by_field = sort_mapping.get(sort, 'updated_at')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc())
        else:
            qs = qs.order_by(F(order_by_field).asc())

        # Pagination density: 10 per page
        paginator = Paginator(qs, 10)
        tenants = paginator.get_page(request.GET.get('page', 1))
        total_count = qs.count()
        start_index = tenants.start_index()
        for offset, tenant in enumerate(tenants.object_list):
            # Show descending list ID so newest appears with highest number.
            tenant.list_rank = total_count - (start_index + offset) + 1
        
        sales_reps = AdminUser.objects.filter(status='Active', is_deleted=False).order_by(
            'first_name', 'last_name'
        )
        plans = SubscriptionPlan.objects.filter(is_active=True, is_deleted=False).order_by(
            'plan_name_en'
        )
        
        return render(
            request,
            self.template_name,
            {
                'tenants': tenants,
                'search_query': search,
                'status_filter': status_filter,
                'rep_filter': rep_filter,
                'plan_filter': plan_filter,
                'sales_reps': sales_reps,
                'plans': plans,
                'today': date.today(),
                'status_choices': TenantProfile.STATUS_CHOICES,
                'sort': sort,
                'dir': direction,
            },
        )


class TenantCreateView(LoginRequiredMixin, View):
    template_name = 'crm/tenants/tenant_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {'form': TenantProfileCreateForm(), 'is_edit': False},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = TenantProfileCreateForm(request.POST)
        if form.is_valid():
            tenant = form.save(commit=False)
            tenant.wallet_balance = Decimal('0.00')
            tenant.total_ltv = Decimal('0.00')
            tenant.current_plan = None
            tenant.subscription_start_date = None
            tenant.subscription_expiry_date = None
            tenant.active_max_users = 0
            tenant.active_max_internal_trucks = 0
            tenant.active_max_external_trucks = 0
            tenant.active_max_drivers = 0
            tenant.save()
            import secrets
            from django.contrib.auth.hashers import make_password

            from superadmin.provisioning import (
                schedule_tenant_workspace_provisioning,
            )

            plain_key = secrets.token_urlsafe(32)
            plain_portal = secrets.token_urlsafe(14)
            tenant.api_bridge_secret_hash = make_password(plain_key)
            tenant.portal_bootstrap_password_hash = make_password(plain_portal)
            tenant.save(
                update_fields=[
                    'api_bridge_secret_hash',
                    'portal_bootstrap_password_hash',
                ],
            )
            schedule_tenant_workspace_provisioning(tenant)
            welcome_email_sent = False
            try:
                from superadmin.communication_helpers import send_tenant_welcome_email

                TenantAuthToken.objects.filter(
                    tenant_profile=tenant,
                    token_type=TenantAuthToken.TokenType.INVITE,
                    is_used=False,
                ).update(is_used=True)
                tenant_token = TenantAuthToken.objects.create(
                    tenant_profile=tenant,
                    token=secrets.token_urlsafe(32),
                    token_type=TenantAuthToken.TokenType.INVITE,
                    expires_at=timezone.now() + timedelta(hours=24),
                )
                invite_url = request.build_absolute_uri(
                    reverse('set_password', args=[tenant_token.token])
                )
                welcome_email_sent = bool(send_tenant_welcome_email(
                    tenant,
                    plain_key,
                    plain_portal,
                    invite_url=invite_url,
                ))
            except Exception:
                logger.exception(
                    'Welcome email failed for tenant %s',
                    tenant.tenant_id,
                )
            if welcome_email_sent:
                messages.success(
                    request,
                    f'Subscriber profile created for {tenant.primary_email}. '
                    
                )
            else:
                messages.warning(
                    request,
                    f'Subscriber profile created for {tenant.primary_email}, '
                    'but welcome email was not sent. Please verify Notification Template is Active and SMTP settings are valid.'
                )
            return redirect(reverse('tenant_detail', kwargs={'pk': tenant.pk}))
        return render(
            request,
            self.template_name,
            {'form': form, 'is_edit': False},
        )


class TenantUpdateView(LoginRequiredMixin, View):
    template_name = 'crm/tenants/tenant_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        tenant = get_object_or_404(TenantProfile, pk=pk)
        return render(
            request,
            self.template_name,
            {
                'form': TenantProfileUpdateForm(instance=tenant),
                'is_edit': True,
                'tenant': tenant,
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        tenant = get_object_or_404(TenantProfile, pk=pk)
        old_obj = TenantProfile.objects.get(pk=pk)
        old_status = tenant.account_status
        form = TenantProfileUpdateForm(request.POST, instance=tenant)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    'form': form,
                    'is_edit': True,
                    'tenant': tenant,
                },
            )
        new_status = form.cleaned_data['account_status']
        inst = form.save()
        if old_status != new_status:
            log_audit_action(
                request,
                'Status_Change',
                'Tenant Profile',
                str(inst.tenant_id),
                old_instance=old_obj,
                new_instance=inst,
            )
        messages.success(
            request,
            'Subscriber profile updated successfully.',
        )
        return redirect(reverse('tenant_detail', kwargs={'pk': tenant.pk}))


class TenantHistoryPartialView(LoginRequiredMixin, View):
    template_name = 'crm/tenants/partials/tenant_history_partial.html'

    def get(self, request, pk):
        tenant = get_object_or_404(TenantProfile, tenant_id=pk)
        # Fetch last 15 orders with their primary invoice pre-fetched
        orders = SubscriptionOrder.objects.filter(
            tenant=tenant
        ).select_related(
            'currency', 'created_by'
        ).prefetch_related(
            'invoices'
        ).order_by('-created_at')[:15]
        
        return render(request, self.template_name, {
            'tenant': tenant,
            'orders': orders,
        })


class TenantDetailView(LoginRequiredMixin, View):
    template_name = 'crm/tenants/tenant_detail.html'

    def get(self, request, pk):
        tenant = get_object_or_404(
            TenantProfile.objects.select_related(
                'country',
                'current_plan',
                'scheduled_downgrade_plan',
                'assigned_sales_rep',
            ),
            pk=pk,
        )
        notes = (
            tenant.crm_notes.select_related('admin')
            .order_by('-created_at')[:10]
        )
        return render(
            request,
            self.template_name,
            {
                'tenant': tenant,
                'notes': notes,
                'today': date.today(),
                'note_type_choices': CRMNote.NOTE_TYPE_CHOICES,
            },
        )


class CRMNoteCreateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        tenant = get_object_or_404(TenantProfile, pk=pk)
        note_type = request.POST.get('note_type', 'General')
        note_content = (request.POST.get('note_content') or '').strip()
        valid_types = {c[0] for c in CRMNote.NOTE_TYPE_CHOICES}
        if note_type not in valid_types:
            note_type = 'General'
        if not note_content:
            messages.error(request, 'Note content is required.')
            return redirect(reverse('tenant_detail', kwargs={'pk': pk}))
        created_note = CRMNote.objects.create(
            tenant=tenant,
            admin=request.user,
            note_type=note_type,
            note_content=note_content,
        )
        log_audit_action(
            request,
            'Create',
            'CRM Notes',
            str(tenant.tenant_id),
            new_instance=created_note,
        )
        messages.success(request, 'CRM note added successfully.')
        return redirect(reverse('tenant_detail', kwargs={'pk': pk}))

    def get(self, request, pk):
        return redirect(reverse('tenant_detail', kwargs={'pk': pk}))


# --- CRM: subscription orders, transactions, standard invoices ---


class RootRequiredMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if not getattr(request.user, 'is_root', False):
            messages.error(
                request,
                'You do not have permission to perform this action.',
            )
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)


class TenantImpersonationView(RootRequiredMixin, View):
    """CP-PCS-P5 short-lived JWT for root 'Login As' workspace handoff."""

    http_method_names = ['post']

    def post(self, request, pk):
        from django.conf import settings as dj_settings

        from superadmin.tenant_jwt import sign_cp_impersonation_jwt

        tenant = get_object_or_404(TenantProfile, pk=pk)
        if tenant.account_status != 'Active':
            messages.error(
                request,
                'Only active subscribers can be opened in the workspace.',
            )
            return redirect(reverse('tenant_detail', kwargs={'pk': pk}))

        token = sign_cp_impersonation_jwt(tenant, request.user, ttl_minutes=15)
        admin_label = (
            f'{request.user.first_name} {request.user.last_name}'.strip()
            or request.user.email
        )
        CRMNote.objects.create(
            tenant=tenant,
            admin=request.user,
            note_type='General',
            note_content=(
                f'Control Panel "Login As": {admin_label} issued a 15-minute '
                'workspace impersonation token (tenant portal handoff).'
            ),
        )
        log_audit_action(
            request,
            'Create',
            'Tenant Impersonation',
            str(tenant.tenant_id),
            new_instance=tenant,
        )
        target = (
            getattr(dj_settings, 'TENANT_IMPERSONATION_REDIRECT_URL', '') or ''
        ).strip()
        if target:
            join_char = '&' if ('?' in target) else '?'
            return HttpResponseRedirect(
                f'{target}{join_char}cp_impersonation_token={token}',
            )
        return render(
            request,
            'auth/tenant_impersonation_handoff.html',
            {
                'tenant': tenant,
                'impersonation_token': token,
                'expires_minutes': 15,
            },
        )


class TenantDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not getattr(request.user, 'is_root', False):
            messages.error(request, 'You do not have permission to perform this action.')
            return redirect(reverse('tenant_list'))
        tenant = get_object_or_404(TenantProfile, pk=pk)
        if tenant.is_deleted:
            messages.info(request, 'Tenant is already removed.')
            return redirect(reverse('tenant_list'))
        tenant.is_deleted = True
        tenant.save(update_fields=['is_deleted'])
        messages.success(request, 'Tenant removed successfully.')
        return redirect(reverse('tenant_list'))

    def get(self, request, pk):
        return self.post(request, pk)


_PLAN_CLASSIFICATIONS = {
    'New_Subscription',
    'Renewal',
    'Upgrade',
    'Downgrade',
}


def _billing_addon_unit_price(policy, add_on_type):
    mapping = {
        'Extra_User': policy.extra_internal_user_price,
        'Extra_Internal_Truck': policy.extra_internal_truck_price,
        'Extra_External_Truck': policy.extra_external_truck_price,
        'Extra_Driver': policy.extra_driver_price,
        'Extra_Shipment': policy.extra_shipment_price,
        'Extra_Storage_GB': policy.extra_storage_gb_price,
    }
    return mapping.get(add_on_type, Decimal('0.00'))


class OrderListView(LoginRequiredMixin, View):
    template_name = 'crm/orders/order_list.html'

    def get(self, request):
        qs = SubscriptionOrder.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc()
            )
        ).select_related(
            'tenant', 'currency', 'created_by',
        ).prefetch_related(
            'invoices'
        )

        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(tenant__company_name__icontains=q)
        
        tenant_id = request.GET.get('tenant', '').strip()
        if tenant_id:
            qs = qs.filter(tenant_id=tenant_id)

        oc = request.GET.get('order_classification', '').strip()
        if oc:
            qs = qs.filter(order_classification=oc)
        st = request.GET.get('order_status', '').strip()
        if st:
            qs = qs.filter(order_status=st)

        # Advanced Sorting Logic
        sort = request.GET.get('sort', 'date')
        direction = request.GET.get('dir', 'desc')
        
        sort_mapping = {
            'rank': 'default_rank',
            'tenant': 'tenant__company_name',
            'classification': 'order_classification',
            'currency': 'currency_id',
            'total': 'grand_total',
            'status': 'order_status',
            'created_by': 'created_by__first_name',
            'date': 'created_at',
        }
        
        order_by_field = sort_mapping.get(sort, 'created_at')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc())
        else:
            qs = qs.order_by(F(order_by_field).asc())

        # Increased pagination: 10 per page
        paginator = Paginator(qs, 10)
        page = paginator.get_page(request.GET.get('page'))
        total_count = qs.count()
        start_index = page.start_index()
        for offset, order in enumerate(page.object_list):
            # Show descending list ID so newest appears with highest number.
            order.list_rank = total_count - (start_index + offset) + 1
        
        return render(request, self.template_name, {
            'orders': page,
            'search_query': q,
            'tenant_filter': tenant_id,
            'classification_filter': oc,
            'status_filter': st,
            'classification_choices': SubscriptionOrder.CLASSIFICATION_CHOICES,
            'status_choices': SubscriptionOrder.STATUS_CHOICES,
            'sort': sort,
            'dir': direction,
        })


class OrderCreateView(RootRequiredMixin, View):
    template_name = 'crm/orders/order_form.html'

    def get(self, request):
        tenant_pre = request.GET.get('tenant', '').strip()
        tenant_obj = None
        if tenant_pre:
            tenant_obj = TenantProfile.objects.filter(
                tenant_id=tenant_pre,
            ).first()

        pricing_qs = PlanPricingCycle.objects.filter(
            plan__is_active=True,
            plan__is_deleted=False,
            currency__is_active=True,
        ).select_related('plan', 'currency')
        plan_qs = SubscriptionPlan.objects.filter(
            is_active=True,
            is_deleted=False,
            pricing_cycles__in=pricing_qs,
        ).distinct().order_by('plan_name_en')
        pricing_json = [
            {
                'plan_id': str(r.plan_id),
                'currency': r.currency_id,
                'cycles': r.number_of_cycles,
                'price': str(r.price),
            }
            for r in pricing_qs
        ]
        policy = AddOnsPricingPolicy.objects.filter(is_active=True).first()
        addons_json = {}
        if policy:
            for choice_code, _label in OrderAddonLine.ADDON_TYPE_CHOICES:
                addons_json[choice_code] = str(
                    _billing_addon_unit_price(policy, choice_code))

        upgrade_credits_by_currency = {}
        active_currencies = Currency.objects.filter(
            is_active=True,
            plan_pricing__in=pricing_qs,
        ).distinct().order_by('currency_code')
        if tenant_obj and tenant_obj.current_plan:
            for cur in active_currencies:
                op = resolve_upgrade_credit_basis_price(
                    tenant_obj.current_plan, cur.currency_code)
                cr = calculate_pro_rata_credit(tenant_obj, op)
                upgrade_credits_by_currency[cur.currency_code] = str(cr)

        return render(request, self.template_name, {
            'tenants': TenantProfile.objects.order_by('company_name'),
            'currencies': active_currencies,
            'plans': plan_qs,
            'payment_methods': PaymentMethod.objects.filter(
                is_active=True).order_by('display_order'),
            'classification_choices': SubscriptionOrder.CLASSIFICATION_CHOICES,
            'addon_action_choices': OrderAddonLine.ACTION_TYPE_CHOICES,
            'addon_type_choices': OrderAddonLine.ADDON_TYPE_CHOICES,
            'tenant_preselect': str(tenant_obj.tenant_id) if tenant_obj else '',
            'tenant_preselect_name': tenant_obj.company_name if tenant_obj else '',
            'tenant_locked': bool(tenant_obj),
            'pricing_data': pricing_json,
            'addons_data': addons_json,
            'upgrade_credits_by_currency': upgrade_credits_by_currency,
        })

    def post(self, request):
        tenant_id = request.POST.get('tenant', '').strip()
        classification = request.POST.get('order_classification', '').strip()
        currency_id = request.POST.get('currency', '').strip()
        payment_method_id = request.POST.get('payment_method', '').strip()
        promo_input = request.POST.get('promo_code', '').strip()

        tenant = TenantProfile.objects.filter(tenant_id=tenant_id).first()
        if not tenant:
            messages.error(request, 'Select a valid tenant.')
            return redirect('order_create')

        currency = Currency.objects.filter(
            currency_code=currency_id, is_active=True).first()
        if not currency:
            messages.error(request, 'Select a valid currency.')
            return redirect('order_create')

        pm = None
        if payment_method_id:
            pm = PaymentMethod.objects.filter(
                method_id=payment_method_id,
                is_active=True,
            ).first()
        if not pm:
            messages.error(request, 'Select a valid payment method.')
            return redirect('order_create')

        if classification not in dict(SubscriptionOrder.CLASSIFICATION_CHOICES):
            messages.error(request, 'Invalid order classification.')
            return redirect('order_create')

        tax = get_tax_code_for_tenant(tenant, client_ip=get_client_ip(request))
        if tax is None:
            messages.error(
                request,
                'No active tax setting found. Configure Global Tax Settings '
                'before creating orders.',
            )
            return redirect('order_create')
        fx = get_fx_snapshot(currency_id, strict=True)
        if fx is None:
            messages.error(
                request,
                'No active FX rate found for selected currency. '
                'Configure Exchange Rates before creating orders.',
            )
            return redirect('order_create')
        tax_rate = tax.rate_percent if tax else Decimal('0.00')

        promo_obj = None
        if promo_input:
            promo_obj = PromoCode.objects.filter(
                code__iexact=promo_input).first()

        sub_total = Decimal('0.00')
        plan_line_data = None

        payment_txn = None
        with db_transaction.atomic():
            order = SubscriptionOrder.objects.create(
                tenant=tenant,
                order_classification=classification,
                currency=currency,
                payment_method=pm,
                created_by=request.user,
                promo_code=None,
                tax_code=tax,
                sub_total=Decimal('0.00'),
                discount_amount=Decimal('0.00'),
                tax_amount=Decimal('0.00'),
                grand_total=Decimal('0.00'),
                exchange_rate_snapshot=fx,
                base_currency_equivalent=Decimal('0.00'),
                # Skip a separate draft step and move directly to payment flow.
                order_status='Pending_Payment',
            )

            if classification in _PLAN_CLASSIFICATIONS:
                plan_id = request.POST.get('plan', '').strip()
                try:
                    cycles = int(request.POST.get('number_of_cycles', '1'))
                except ValueError:
                    cycles = 1
                plan = SubscriptionPlan.objects.filter(
                    plan_id=plan_id,
                    is_active=True,
                    is_deleted=False,
                ).first()
                if not plan:
                    messages.error(request, 'Select a valid plan.')
                    order.delete()
                    return redirect('order_create')
                ppc = PlanPricingCycle.objects.filter(
                    plan=plan,
                    currency=currency,
                    number_of_cycles=cycles,
                ).first()
                if not ppc:
                    messages.error(
                        request,
                        'No pricing found for this plan, currency, and cycle count.',
                    )
                    order.delete()
                    return redirect('order_create')
                if classification == 'Downgrade':
                    err = validate_downgrade_order(tenant, plan)
                    if err:
                        messages.error(request, err)
                        order.delete()
                        return redirect('order_create')
                plan_price = ppc.price
                pro_rata = Decimal('0.00')
                if classification == 'Upgrade' and tenant.current_plan:
                    old_px = resolve_upgrade_credit_basis_price(
                        tenant.current_plan, currency.currency_code)
                    pro_rata = calculate_pro_rata_credit(tenant, old_px)
                line_total = (plan_price + pro_rata).quantize(Decimal('0.01'))
                OrderPlanLine.objects.create(
                    order=order,
                    plan=plan,
                    number_of_cycles=cycles,
                    plan_price=plan_price,
                    pro_rata_adjustment=pro_rata,
                    line_total=line_total,
                    plan_name_en_snapshot=plan.plan_name_en,
                    plan_name_ar_snapshot=plan.plan_name_ar or '',
                )
                sub_total += line_total
                plan_line_data = plan

            elif classification == 'Add_ons':
                policy = AddOnsPricingPolicy.objects.filter(
                    is_active=True).first()
                if not policy:
                    messages.error(
                        request,
                        'No active add-ons pricing policy. Configure one first.',
                    )
                    order.delete()
                    return redirect('order_create')
                actions = request.POST.getlist('addon_action')
                types = request.POST.getlist('addon_add_on_type')
                qtys = request.POST.getlist('addon_quantity')
                base_days = (
                    get_plan_cycle_days(tenant.current_plan)
                    if tenant.current_plan else get_standard_billing_cycle_days()
                )
                expiry = tenant.subscription_expiry_date or timezone.now().date()
                for action, add_type, qty_s in zip(actions, types, qtys):
                    if not add_type:
                        continue
                    try:
                        qty = int(qty_s)
                    except ValueError:
                        qty = 1
                    if qty < 1:
                        qty = 1
                    if action not in dict(OrderAddonLine.ACTION_TYPE_CHOICES):
                        action = 'Add'
                    unit = _billing_addon_unit_price(policy, add_type)
                    cycles_fr, prorata_unit_total = calculate_addon_prorata(
                        unit, base_days, expiry)
                    signed_qty = qty if action == 'Add' else -qty
                    line_piece = (
                        Decimal(str(signed_qty)) *
                        prorata_unit_total).quantize(Decimal('0.01'))
                    addon_label = dict(
                        OrderAddonLine.ADDON_TYPE_CHOICES,
                    ).get(add_type, add_type)
                    OrderAddonLine.objects.create(
                        order=order,
                        action_type=action,
                        add_on_type=add_type,
                        quantity=qty,
                        number_of_cycles=cycles_fr,
                        unit_price=unit,
                        pro_rata_adjustment=Decimal('0.00'),
                        line_total=line_piece,
                        add_on_type_label_snapshot=str(addon_label),
                    )
                    sub_total += line_piece
            else:
                messages.error(request, 'Unsupported classification.')
                order.delete()
                return redirect('order_create')

            if classification == 'Add_ons' and sub_total == 0:
                messages.error(
                    request,
                    'Add at least one add-on line with quantity.',
                )
                order.delete()
                return redirect('order_create')

            if promo_input:
                if not promo_obj:
                    messages.error(request, 'Invalid or Expired Code.')
                    order.delete()
                    return redirect('order_create')
                ok, err = promo_obj.is_valid_for_use(for_plan=plan_line_data)
                if not ok:
                    messages.error(
                        request,
                        err or 'Invalid or Expired Code.',
                    )
                    order.delete()
                    return redirect('order_create')

            discount = calculate_promo_discount(
                promo_obj, sub_total, for_plan=plan_line_data)
            taxable_base = (sub_total - discount).quantize(Decimal('0.01'))
            if taxable_base < 0:
                taxable_base = Decimal('0.00')
            tax_amount = (
                taxable_base * tax_rate / Decimal('100')
            ).quantize(Decimal('0.01'))
            grand_total = (taxable_base + tax_amount).quantize(Decimal('0.01'))
            base_equiv = (grand_total * fx).quantize(Decimal('0.01'))

            order.promo_code = promo_obj
            order.sub_total = sub_total
            order.discount_amount = discount
            order.tax_amount = tax_amount
            order.grand_total = grand_total
            order.base_currency_equivalent = base_equiv
            order.save(update_fields=[
                'promo_code', 'sub_total', 'discount_amount',
                'tax_amount', 'grand_total', 'base_currency_equivalent',
            ])

            refresh_order_projected_fields(order)
            order.save(update_fields=[
                'projected_plan', 'projected_expiry_date',
                'projected_max_users', 'projected_max_internal_trucks',
                'projected_max_external_trucks', 'projected_max_drivers',
            ])
            payment_txn = sync_or_create_order_payment_transaction(order)

        pm = order.payment_method
        if pm and pm.method_type == 'Online_Gateway':
            if complete_order_payment_as_system(order, request.user):
                messages.success(
                    request,
                    'Online payment captured. Order fulfilled.',
                )
            else:
                messages.warning(
                    request,
                    'Order submitted; online capture did not complete '
                    '(check amount or retry from order detail).',
                )
            payment_txn = Transaction.objects.filter(
                order=order,
                transaction_type='Order_Payment',
            ).order_by('-created_at').first()
        else:
            if pm and pm.method_type == 'Offline_Bank':
                try:
                    dispatch_internal_alerts(
                        'Bank_Transfer_Pending',
                        context_dict={
                            'order_id': str(order.order_id),
                            'tenant_id': str(order.tenant_id),
                            'company_name': order.tenant.company_name,
                            'amount': str(order.grand_total),
                            'currency': order.currency_id,
                            'message': (
                                f'Bank transfer pending for '
                                f'"{order.tenant.company_name}".'
                            ),
                        },
                    )
                except Exception:
                    logger.exception(
                        'Internal alert dispatch failed for pending bank transfer order %s',
                        order.order_id,
                    )
            messages.success(request, 'Order marked as pending payment.')

        return redirect('order_detail', pk=order.pk)


class OrderPreviewAjaxView(RootRequiredMixin, View):
    """Realtime order calculator for create form projected preview."""

    def get(self, request):
        tenant_id = (request.GET.get('tenant') or '').strip()
        classification = (request.GET.get('order_classification') or '').strip()
        currency_id = (request.GET.get('currency') or '').strip()
        plan_id = (request.GET.get('plan') or '').strip()
        promo_input = (request.GET.get('promo_code') or '').strip()
        try:
            cycles = int((request.GET.get('number_of_cycles') or '1').strip())
        except ValueError:
            cycles = 1
        cycles = max(1, cycles)

        tenant = TenantProfile.objects.filter(tenant_id=tenant_id).first()
        if not tenant:
            return JsonResponse({'ok': False, 'error': 'Select a valid tenant.'}, status=400)
        currency = Currency.objects.filter(
            currency_code=currency_id,
            is_active=True,
        ).first()
        if not currency:
            return JsonResponse({'ok': False, 'error': 'Select a valid currency.'}, status=400)
        if classification not in dict(SubscriptionOrder.CLASSIFICATION_CHOICES):
            return JsonResponse({'ok': False, 'error': 'Invalid order classification.'}, status=400)

        tax = get_tax_code_for_tenant(tenant, client_ip=get_client_ip(request))
        tax_rate = tax.rate_percent if tax else Decimal('0.00')
        fx = get_fx_snapshot(currency_id, strict=True)
        if fx is None:
            return JsonResponse(
                {
                    'ok': False,
                    'error': 'No active FX rate found for selected currency.',
                },
                status=400,
            )

        sub_total = Decimal('0.00')
        plan_price = Decimal('0.00')
        pro_rata = Decimal('0.00')
        plan_line_total = Decimal('0.00')
        selected_plan = None
        promo_obj = None
        addon_preview_rows = []

        if classification in _PLAN_CLASSIFICATIONS:
            if plan_id:
                selected_plan = SubscriptionPlan.objects.filter(
                    plan_id=plan_id,
                    is_active=True,
                    is_deleted=False,
                ).first()
            if selected_plan:
                ppc = PlanPricingCycle.objects.filter(
                    plan=selected_plan,
                    currency=currency,
                    number_of_cycles=cycles,
                ).first()
                if ppc:
                    plan_price = ppc.price
                    if classification == 'Upgrade' and tenant.current_plan:
                        old_px = resolve_upgrade_credit_basis_price(
                            tenant.current_plan,
                            currency.currency_code,
                        )
                        pro_rata = calculate_pro_rata_credit(tenant, old_px)
                    plan_line_total = (plan_price + pro_rata).quantize(Decimal('0.01'))
                    sub_total += plan_line_total

        elif classification == 'Add_ons':
            policy = AddOnsPricingPolicy.objects.filter(is_active=True).first()
            if policy:
                actions = request.GET.getlist('addon_action')
                types = request.GET.getlist('addon_add_on_type')
                qtys = request.GET.getlist('addon_quantity')
                base_days = (
                    get_plan_cycle_days(tenant.current_plan)
                    if tenant.current_plan else get_standard_billing_cycle_days()
                )
                expiry = tenant.subscription_expiry_date or timezone.now().date()
                for action, add_type, qty_s in zip(actions, types, qtys):
                    if not add_type:
                        continue
                    try:
                        qty = int(qty_s)
                    except ValueError:
                        qty = 1
                    qty = max(1, qty)
                    if action not in dict(OrderAddonLine.ACTION_TYPE_CHOICES):
                        action = 'Add'
                    unit = _billing_addon_unit_price(policy, add_type)
                    cycles_fr, prorata_unit_total = calculate_addon_prorata(
                        unit,
                        base_days,
                        expiry,
                    )
                    signed_qty = qty if action == 'Add' else -qty
                    line_total = (
                        Decimal(str(signed_qty)) *
                        prorata_unit_total
                    ).quantize(Decimal('0.01'))
                    sub_total += line_total
                    addon_preview_rows.append({
                        'action_type': action,
                        'add_on_type': add_type,
                        'quantity': qty,
                        'unit_price': str(unit.quantize(Decimal('0.01'))),
                        'number_of_cycles': str(cycles_fr.quantize(Decimal('0.01'))),
                        'pro_rata_adjustment': '0.00',
                        'line_total': str(line_total),
                    })

        promo_error = ''
        if promo_input:
            promo_obj = PromoCode.objects.filter(code__iexact=promo_input).first()
            if not promo_obj:
                promo_error = 'Invalid or Expired Code.'
            else:
                ok, err = promo_obj.is_valid_for_use(for_plan=selected_plan)
                if not ok:
                    promo_obj = None
                    promo_error = err or 'Invalid or Expired Code.'

        discount = calculate_promo_discount(
            promo_obj,
            sub_total,
            for_plan=selected_plan,
        ).quantize(Decimal('0.01'))
        taxable_base = (sub_total - discount).quantize(Decimal('0.01'))
        if taxable_base < 0:
            taxable_base = Decimal('0.00')
        tax_amount = (taxable_base * tax_rate / Decimal('100')).quantize(Decimal('0.01'))
        grand_total = (taxable_base + tax_amount).quantize(Decimal('0.01'))
        base_equiv = (grand_total * fx).quantize(Decimal('0.01'))

        proj_plan = tenant.current_plan
        proj_expiry = tenant.subscription_expiry_date
        proj_u = tenant.active_max_users
        proj_it = tenant.active_max_internal_trucks
        # MVP scope: keep backend value for compatibility, but hide UI wiring
        proj_et = tenant.active_max_external_trucks
        proj_d = tenant.active_max_drivers

        if classification == 'New_Subscription' and selected_plan:
            proj_plan = selected_plan
            proj_expiry = date.today() + timedelta(
                days=get_plan_cycle_days(selected_plan) * cycles,
            )
            proj_u = selected_plan.max_internal_users
            proj_it = selected_plan.max_internal_trucks
            proj_et = selected_plan.max_external_trucks
            proj_d = selected_plan.max_active_drivers
        elif classification == 'Renewal' and selected_plan:
            proj_plan = selected_plan
            extra = get_plan_cycle_days(selected_plan) * cycles
            if tenant.subscription_expiry_date:
                proj_expiry = tenant.subscription_expiry_date + timedelta(days=extra)
            else:
                proj_expiry = date.today() + timedelta(days=extra)
        elif classification == 'Upgrade' and selected_plan:
            proj_plan = selected_plan
            proj_expiry = date.today() + timedelta(
                days=get_plan_cycle_days(selected_plan) * cycles,
            )
            proj_u = selected_plan.max_internal_users
            proj_it = selected_plan.max_internal_trucks
            proj_et = selected_plan.max_external_trucks
            proj_d = selected_plan.max_active_drivers
        elif classification == 'Downgrade' and selected_plan:
            proj_plan = selected_plan
            proj_expiry = tenant.subscription_expiry_date
            proj_u = selected_plan.max_internal_users
            proj_it = selected_plan.max_internal_trucks
            proj_et = selected_plan.max_external_trucks
            proj_d = selected_plan.max_active_drivers
        elif classification == 'Add_ons':
            for row in addon_preview_rows:
                qty = int(row['quantity'])
                if row['action_type'] == 'Reduce':
                    qty = -qty
                if row['add_on_type'] == 'Extra_User':
                    proj_u += qty
                elif row['add_on_type'] == 'Extra_Internal_Truck':
                    proj_it += qty
                elif row['add_on_type'] == 'Extra_External_Truck':
                    proj_et += qty
                elif row['add_on_type'] == 'Extra_Driver':
                    proj_d += qty

        return JsonResponse({
            'ok': True,
            'totals': {
                'sub_total': str(sub_total.quantize(Decimal('0.01'))),
                'discount_amount': str(discount),
                'tax_amount': str(tax_amount),
                'grand_total': str(grand_total),
                'exchange_rate_snapshot': str(fx),
                'base_currency_equivalent': str(base_equiv),
            },
            'plan_preview': {
                'plan_price': str(plan_price.quantize(Decimal('0.01'))),
                'pro_rata_adjustment': str(pro_rata.quantize(Decimal('0.01'))),
                'line_total': str(plan_line_total.quantize(Decimal('0.01'))),
            },
            'projected': {
                'plan_id': str(proj_plan.plan_id) if proj_plan else '',
                'plan_name': proj_plan.plan_name_en if proj_plan else '',
                'expiry_date': proj_expiry.isoformat() if proj_expiry else '',
                'max_users': proj_u,
                'max_internal_trucks': proj_it,
                # 'max_external_trucks': proj_et,  # MVP scope: disable UI connection
                'max_drivers': proj_d,
            },
            'addon_rows': addon_preview_rows,
            'promo': {
                'code': promo_obj.code if promo_obj else '',
                'valid': bool(promo_obj) and not promo_error,
                'error': promo_error,
                'discount_amount': str(discount),
            },
        })


class OrderDetailView(LoginRequiredMixin, View):
    template_name = 'crm/orders/order_detail.html'

    def get(self, request, pk):
        order = get_object_or_404(
            SubscriptionOrder.objects.select_related(
                'tenant',
                'tenant__scheduled_downgrade_plan',
                'currency', 'payment_method', 'promo_code',
                'tax_code', 'created_by', 'projected_plan',
            ).prefetch_related('plan_lines__plan', 'addon_lines'),
            order_id=pk,
        )
        payment_txn = Transaction.objects.filter(
            order=order,
            transaction_type='Order_Payment',
        ).order_by('-created_at').first()
        invoice = order.invoices.order_by('-issue_date').first()
        downgrade_schedule = None
        if order.order_classification == 'Downgrade' and order.order_status == 'Paid':
            plan_line = order.plan_lines.first() if order.plan_lines.exists() else None
            plan_name = ''
            if plan_line:
                plan_name = (
                    (plan_line.plan_name_en_snapshot or '').strip()
                    or plan_line.plan.plan_name_en
                )
            eff_date = order.projected_expiry_date
            if plan_name and eff_date:
                downgrade_schedule = {
                    'plan_name': plan_name,
                    'effective_date': eff_date,
                }
        return render(request, self.template_name, {
            'order': order,
            'payment_txn': payment_txn,
            'invoice': invoice,
            'downgrade_schedule': downgrade_schedule,
        })


class OrderStatusUpdateView(RootRequiredMixin, View):
    def post(self, request, pk):
        order = get_object_or_404(SubscriptionOrder, order_id=pk)
        action = request.POST.get('action', '')
        if action == 'pending_payment':
            if order.order_status != 'Draft':
                messages.error(request, 'Only draft orders can be submitted.')
            else:
                order.order_status = 'Pending_Payment'
                order.save(update_fields=['order_status', 'updated_at'])
                sync_or_create_order_payment_transaction(order)
                order.refresh_from_db()
                pm = order.payment_method
                if pm and pm.method_type == 'Online_Gateway':
                    if complete_order_payment_as_system(order, request.user):
                        messages.success(
                            request,
                            'Online payment captured. Order fulfilled.',
                        )
                    else:
                        messages.warning(
                            request,
                            'Order submitted; online capture did not complete '
                            '(check amount or retry from order detail).',
                        )
                else:
                    if pm and pm.method_type == 'Offline_Bank':
                        try:
                            dispatch_internal_alerts(
                                'Bank_Transfer_Pending',
                                context_dict={
                                    'order_id': str(order.order_id),
                                    'tenant_id': str(order.tenant_id),
                                    'company_name': order.tenant.company_name,
                                    'amount': str(order.grand_total),
                                    'currency': order.currency_id,
                                    'message': (
                                        f'Bank transfer pending for '
                                        f'"{order.tenant.company_name}".'
                                    ),
                                },
                            )
                        except Exception:
                            logger.exception(
                                'Internal alert dispatch failed for pending bank transfer order %s',
                                order.order_id,
                            )
                    messages.success(
                        request,
                        'Order marked as pending payment.',
                    )
        elif action == 'cancel':
            if order.order_status not in ('Draft', 'Pending_Payment'):
                messages.error(request, 'This order cannot be cancelled.')
            else:
                order.order_status = 'Cancelled'
                order.save(update_fields=['order_status', 'updated_at'])
                Transaction.objects.filter(
                    order=order,
                    transaction_type='Order_Payment',
                    status='Pending',
                ).update(status='Failed')
                messages.success(request, 'Order cancelled.')
        elif action == 'mark_paid':
            if order.order_status != 'Pending_Payment':
                messages.error(
                    request,
                    'Only orders awaiting payment can be recorded as paid.',
                )
            else:
                fulfilled = False
                uploaded_attachment = request.FILES.get('attachment')
                with db_transaction.atomic():
                    ord_row = SubscriptionOrder.objects.select_for_update().get(
                        pk=order.pk)
                    sync_or_create_order_payment_transaction(ord_row)
                    txn = (
                        Transaction.objects.select_for_update()
                        .filter(
                            order=ord_row,
                            transaction_type='Order_Payment',
                            status='Pending',
                        )
                        .first()
                    )
                    if not txn:
                        messages.error(
                            request,
                            'No pending payment transaction found for this order.',
                        )
                    else:
                        pm = ord_row.payment_method
                        if uploaded_attachment:
                            txn.attachment = _rename_txn_attachment(
                                uploaded_attachment,
                                txn.transaction_id,
                            )
                            txn.save(update_fields=['attachment', 'updated_at'])
                        if pm and pm.method_type == 'Offline_Bank' and not txn.attachment:
                            messages.error(
                                request,
                                'Offline bank payments require an attachment/receipt '
                                'before recording payment.',
                            )
                        else:
                            txn.status = 'Completed'
                            txn.reviewed_by = request.user
                            txn.review_notes = None
                            txn.save(update_fields=[
                                'status', 'reviewed_by', 'review_notes',
                                'updated_at',
                            ])
                            ord_row.order_status = 'Paid'
                            ord_row.save(update_fields=[
                                'order_status', 'updated_at',
                            ])
                            fulfill_paid_order(ord_row, request.user, txn.amount)
                            log_audit_action(
                                request,
                                'Status_Change',
                                'Order Marked Paid',
                                str(ord_row.order_id),
                                new_instance=ord_row,
                            )
                            fulfilled = True
                if fulfilled:
                    messages.success(
                        request,
                        'Payment recorded. Invoice, subscription, and LTV updated.',
                    )
        else:
            messages.error(request, 'Unknown action.')
        return redirect('order_detail', pk=order.pk)


class TransactionListView(LoginRequiredMixin, View):
    template_name = 'crm/transactions/txn_list.html'

    def get(self, request):
        # Keep payment transactions consistent with paid orders.
        Transaction.objects.filter(
            transaction_type='Order_Payment',
            status='Pending',
            order__isnull=False,
            order__order_status='Paid',
        ).update(status='Completed', updated_at=timezone.now())
        Transaction.objects.filter(
            transaction_type='Order_Payment',
            status='Pending',
            order__isnull=False,
            order__invoices__isnull=False,
        ).update(status='Completed', updated_at=timezone.now())

        qs = Transaction.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc()
            )
        ).select_related(
            'tenant', 'currency', 'payment_method',
        )

        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(tenant__company_name__icontains=q) |
                Q(gateway_ref__icontains=q)
            )
        tt = request.GET.get('transaction_type', '').strip()
        if tt:
            qs = qs.filter(transaction_type=tt)
        st = request.GET.get('status', '').strip()
        if st:
            qs = qs.filter(status=st)
        cur = request.GET.get('currency', '').strip()
        if cur:
            qs = qs.filter(currency_id=cur)

        # Advanced Sorting Logic
        sort = request.GET.get('sort', 'date')
        direction = request.GET.get('dir', 'desc')
        
        sort_mapping = {
            'rank': 'default_rank',
            'tenant': 'tenant__company_name',
            'type': 'transaction_type',
            'currency': 'currency_id',
            'amount': 'amount',
            'base_equiv': 'base_currency_equivalent_amount',
            'status': 'status',
            'gateway_ref': 'gateway_ref',
            'date': 'created_at',
        }
        
        order_by_field = sort_mapping.get(sort, 'created_at')
        if direction == 'desc':
            qs = qs.order_by(F(order_by_field).desc())
        else:
            qs = qs.order_by(F(order_by_field).asc())

        paginator = Paginator(qs, 10)
        page = paginator.get_page(request.GET.get('page'))
        total_count = qs.count()
        start_index = page.start_index()
        for offset, txn in enumerate(page.object_list):
            # Show descending list ID so newest appears with highest number.
            txn.list_rank = total_count - (start_index + offset) + 1
        
        return render(request, self.template_name, {
            'transactions': page,
            'search_query': q,
            'type_filter': tt,
            'status_filter': st,
            'currency_filter': cur,
            'type_choices': Transaction.TYPE_CHOICES,
            'status_choices': Transaction.STATUS_CHOICES,
            'currencies': Currency.objects.filter(is_active=True).order_by(
                'currency_code'),
            'sort': sort,
            'dir': direction,
        })


class TransactionCreateView(RootRequiredMixin, View):
    template_name = 'crm/transactions/txn_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'tenants': TenantProfile.objects.order_by('company_name'),
            'currencies': Currency.objects.filter(is_active=True).order_by(
                'currency_code'),
        })

    def post(self, request):
        tenant_id = request.POST.get('tenant', '').strip()
        currency_id = request.POST.get('currency', '').strip()
        amount_s = request.POST.get('amount', '').strip()
        gateway_ref = request.POST.get('gateway_ref', '').strip()

        tenant = TenantProfile.objects.filter(tenant_id=tenant_id).first()
        if not tenant:
            messages.error(request, 'Select a valid tenant.')
            return redirect('transaction_create')
        currency = Currency.objects.filter(
            currency_code=currency_id, is_active=True).first()
        if not currency:
            messages.error(request, 'Select a valid currency.')
            return redirect('transaction_create')
        try:
            amount = Decimal(amount_s)
        except Exception:
            amount = Decimal('0')
        if amount < Decimal('0.01'):
            messages.error(request, 'Enter a valid amount.')
            return redirect('transaction_create')

        fx = get_fx_snapshot(currency_id, strict=True)
        if fx is None:
            messages.error(
                request,
                'No active FX rate found for selected currency. '
                'Configure Exchange Rates before recording transaction.',
            )
            return redirect('transaction_create')
        base_equiv = (amount * fx).quantize(Decimal('0.01'))
        attachment = request.FILES.get('attachment')
        payment_method = (
            PaymentMethod.objects.filter(
                is_active=True,
                supported_currencies__contains=[currency_id],
            )
            .order_by('display_order')
            .first()
        )
        if not payment_method:
            payment_method = PaymentMethod.objects.filter(
                is_active=True
            ).order_by('display_order').first()
        if not payment_method:
            messages.error(
                request,
                'No active payment method is configured. '
                'Configure at least one method before recording top-up.',
            )
            return redirect('transaction_create')

        with db_transaction.atomic():
            txn = Transaction.objects.create(
                tenant=tenant,
                order=None,
                transaction_type='Wallet_TopUp',
                payment_method=payment_method,
                currency=currency,
                amount=amount,
                exchange_rate_snapshot=fx,
                base_currency_equivalent_amount=base_equiv,
                status='Completed',
                gateway_ref=gateway_ref or None,
                attachment=attachment,
            )
            tenant.wallet_balance = (
                tenant.wallet_balance + amount
            ).quantize(Decimal('0.01'))
            tenant.save(update_fields=['wallet_balance', 'updated_at'])

        messages.success(request, 'Wallet top-up recorded.')
        return redirect('transaction_list')


class TransactionDetailView(LoginRequiredMixin, View):
    template_name = 'crm/transactions/txn_detail.html'

    def get(self, request, pk):
        txn = get_object_or_404(
            Transaction.objects.select_related(
                'tenant', 'currency', 'payment_method',
                'order', 'reviewed_by',
            ),
            transaction_id=pk,
        )
        if (
            txn.transaction_type == 'Order_Payment'
            and txn.status == 'Pending'
            and txn.order_id
            and txn.order
            and txn.order.order_status == 'Paid'
        ):
            txn.status = 'Completed'
            txn.save(update_fields=['status', 'updated_at'])
            txn.refresh_from_db()
        return render(request, self.template_name, {'txn': txn})


class TransactionApproveView(RootRequiredMixin, View):
    def post(self, request, pk):
        _write_txn_approve_debug(f'Approve requested for txn={pk} user={request.user.pk}')
        try:
            txn = get_object_or_404(
                Transaction.objects.select_related('payment_method'),
                transaction_id=pk,
            )
            _write_txn_approve_debug(
                f'Loaded txn={txn.transaction_id} status={txn.status} order_id={txn.order_id}'
            )
            if txn.status != 'Pending':
                _write_txn_approve_debug('Blocked: transaction is not pending')
                messages.error(request, 'Only pending transactions can be approved.')
                return redirect('transaction_detail', pk=pk)
            pm = txn.payment_method
            if not pm or pm.method_type != 'Offline_Bank':
                _write_txn_approve_debug(
                    f'Blocked: invalid payment method {getattr(pm, "method_type", None)}'
                )
                messages.error(
                    request,
                    'Only offline bank transfers use manual approval.',
                )
                return redirect('transaction_detail', pk=pk)
            uploaded_attachment = request.FILES.get('attachment')
            if uploaded_attachment:
                txn.attachment = _rename_txn_attachment(
                    uploaded_attachment,
                    txn.transaction_id,
                )
                txn.save(update_fields=['attachment', 'updated_at'])
                _write_txn_approve_debug('Attachment uploaded and saved')
            if not txn.attachment:
                _write_txn_approve_debug('Blocked: no attachment present')
                messages.error(
                    request,
                    'Offline bank transfers require an attachment/receipt before approval.',
                )
                return redirect('transaction_detail', pk=pk)

            order = None
            if txn.order_id:
                order = SubscriptionOrder.objects.filter(pk=txn.order_id).first()
            if not order or order.order_status not in ('Pending_Payment', 'Paid'):
                _write_txn_approve_debug(
                    f'Blocked: order invalid or wrong status={getattr(order, "order_status", None)}'
                )
                messages.error(request, 'Order is not awaiting payment.')
                return redirect('transaction_detail', pk=pk)

            # Persist status first (autocommit) so later failures do not revert approval.
            txn.status = 'Completed'
            txn.reviewed_by = request.user
            txn.review_notes = None
            txn.save(update_fields=[
                'status', 'reviewed_by', 'review_notes', 'updated_at',
            ])
            _write_txn_approve_debug('Transaction status updated to Completed')

            if order.order_status != 'Paid':
                order.order_status = 'Paid'
                order.save(update_fields=['order_status', 'updated_at'])
                _write_txn_approve_debug('Order marked Paid')
                try:
                    fulfill_paid_order(order, request.user, txn.amount)
                    _write_txn_approve_debug('Fulfillment completed')
                except Exception as exc:
                    _write_txn_approve_debug(f'Fulfillment ERROR: {exc!r}')
                    logger.exception('Fulfillment failed for txn %s', txn.transaction_id)
            else:
                _write_txn_approve_debug('Order already Paid; skipped fulfill')

            log_audit_action(
                request,
                'Status_Change',
                'Transaction Approval',
                str(txn.transaction_id),
                new_instance=txn,
            )
            _write_txn_approve_debug('Audit log written')

            post_commit_txn = Transaction.objects.get(pk=txn.pk)
            post_commit_order = SubscriptionOrder.objects.get(pk=order.pk)
            _write_txn_approve_debug(
                'Post-save check: '
                f'txn_status={post_commit_txn.status}, '
                f'order_status={post_commit_order.order_status}'
            )
            if post_commit_txn.status != 'Completed':
                messages.error(
                    request,
                    'Approval failed to persist transaction status. Check debug log.',
                )
                return redirect('transaction_detail', pk=pk)

            messages.success(request, 'Payment approved and order fulfilled.')
            invoice = (
                StandardInvoice.objects.filter(order_id=txn.order_id)
                .order_by('-updated_at', '-issue_date')
                .first()
            )
            if invoice:
                _write_txn_approve_debug(f'Redirecting to invoice_detail {invoice.pk}')
                return redirect('invoice_detail', pk=invoice.pk)
            _write_txn_approve_debug('Redirecting to invoice_list (no invoice found)')
            return redirect('invoice_list')
        except Exception as exc:
            _write_txn_approve_debug(f'ERROR for txn={pk}: {exc!r}')
            logger.exception('Transaction approve failed for %s', pk)
            messages.error(
                request,
                'Approval failed due to an internal error. Please check debug log.',
            )
            return redirect('transaction_detail', pk=pk)


class TransactionUploadAttachmentView(RootRequiredMixin, View):
    def post(self, request, pk):
        txn = get_object_or_404(Transaction, transaction_id=pk)
        if txn.status != 'Pending':
            messages.error(request, 'Attachment upload is only allowed for pending transactions.')
            return redirect('transaction_detail', pk=pk)
        pm = txn.payment_method
        if not pm or pm.method_type != 'Offline_Bank':
            messages.error(request, 'Attachment upload is only required for offline bank transfers.')
            return redirect('transaction_detail', pk=pk)
        uploaded = request.FILES.get('attachment')
        if not uploaded:
            messages.error(request, 'Please choose a receipt file to upload.')
            return redirect('transaction_detail', pk=pk)
        txn.attachment = _rename_txn_attachment(uploaded, txn.transaction_id)
        txn.save(update_fields=['attachment', 'updated_at'])
        messages.success(request, 'Receipt uploaded. You can now approve this transaction.')
        return redirect('transaction_detail', pk=pk)


class TransactionRejectView(RootRequiredMixin, View):
    def post(self, request, pk):
        notes = (request.POST.get('review_notes') or '').strip()
        if not notes:
            messages.error(request, 'Review notes are required to reject.')
            return redirect('transaction_detail', pk=pk)

        linked_order_cancelled = False
        with db_transaction.atomic():
            txn = get_object_or_404(
                Transaction.objects.select_for_update(),
                transaction_id=pk,
            )
            if txn.status != 'Pending':
                messages.error(request, 'Only pending transactions can be rejected.')
                return redirect('transaction_detail', pk=pk)
            txn.status = 'Rejected'
            txn.reviewed_by = request.user
            txn.review_notes = notes
            txn.save(update_fields=[
                'status', 'reviewed_by', 'review_notes', 'updated_at',
            ])
            pm = txn.payment_method
            should_cancel_linked_order = bool(pm and pm.method_type != 'Offline_Bank')
            if should_cancel_linked_order and txn.order_id:
                order = (
                    SubscriptionOrder.objects.select_for_update()
                    .filter(pk=txn.order_id)
                    .first()
                )
                if order and order.order_status == 'Pending_Payment':
                    order.order_status = 'Cancelled'
                    order.save(update_fields=['order_status', 'updated_at'])
                    linked_order_cancelled = True

        try:
            dispatch_internal_alerts(
                'Payment_Failed',
                context_dict={
                    'transaction_id': str(txn.transaction_id),
                    'order_id': str(txn.order_id) if txn.order_id else '',
                    'tenant_id': str(txn.tenant_id),
                    'company_name': txn.tenant.company_name,
                    'amount': str(txn.amount),
                    'currency': txn.currency_id,
                    'review_notes': txn.review_notes or '',
                    'message': (
                        f'Payment failed/rejected for transaction '
                        f'"{txn.tenant.company_name}".'
                    ),
                },
            )
        except Exception:
            logger.exception(
                'Internal alert dispatch failed for rejected transaction %s',
                txn.transaction_id,
            )

        if linked_order_cancelled:
            messages.success(request, 'Transaction rejected and linked order cancelled.')
        else:
            messages.success(request, 'Transaction rejected.')
        return redirect('transaction_detail', pk=pk)


class InvoiceListView(LoginRequiredMixin, View):
    template_name = 'crm/invoices/invoice_list.html'

    def get(self, request):
        # 1. Base Queryset
        qs = StandardInvoice.objects.select_related(
            'tenant', 'currency',
        )

        # 2. Filtering
        st = request.GET.get('status', 'All').strip()
        if st and st != 'All':
            qs = qs.filter(status=st)

        cur = request.GET.get('currency', '').strip()
        if cur:
            qs = qs.filter(currency_id=cur)

        tenant_id = request.GET.get('tenant', '').strip()
        if tenant_id:
            qs = qs.filter(tenant_id=tenant_id)

        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(invoice_number__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(tenant__company_name__icontains=q)
            )

        # 3. Stable Rank Annotation
        qs = qs.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('issue_date').desc()
            )
        )

        # 4. Sorting logic
        # Normalize values used by sorting to avoid null/blank inconsistencies.
        qs = qs.annotate(
            sort_currency_code=Coalesce(
                F('currency__currency_code'),
                Value(''),
                output_field=CharField(),
            ),
            sort_tax_amount=Coalesce(
                F('tax_amount'),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            sort_status_order=Case(
                When(status='Issued', then=Value(1)),
                When(status='Paid', then=Value(2)),
                When(status='Void', then=Value(3)),
                default=Value(99),
                output_field=IntegerField(),
            ),
        )

        sort_key = request.GET.get('sort', 'date')
        direction = request.GET.get('dir', 'desc')

        sort_mapping = {
            'number': ['invoice_number'],
            'tenant': ['tenant__company_name'],
            'total': ['grand_total'],
            'tax': ['sort_tax_amount'],
            'status': ['sort_status_order', 'status'],
            'date': ['issue_date'],
            'currency': ['sort_currency_code'],
            'rank': ['default_rank'],
        }

        active_sort_fields = sort_mapping.get(sort_key, ['issue_date'])
        ordering = []
        for field in active_sort_fields:
            if direction == 'desc':
                ordering.append(f'-{field}')
            else:
                ordering.append(field)
        
        # Secondary stable sort to prevent "shuffling" rows with same values.
        # Use the same direction as the active sort so ASC/DESC clicks are visible
        # even when many rows share the same primary value.
        if 'invoice_number' not in active_sort_fields:
            if direction == 'desc':
                ordering.append('-invoice_number')
            else:
                ordering.append('invoice_number')
        
        qs = qs.order_by(*ordering)

        # 4. Pagination
        paginator = Paginator(qs, 10) # Increased to 10 for better list view
        page = paginator.get_page(request.GET.get('page'))
        total_count = qs.count()
        start_index = page.start_index()
        for offset, invoice in enumerate(page.object_list):
            # Show descending list ID so newest appears with highest number.
            invoice.list_rank = total_count - (start_index + offset) + 1

        return render(request, self.template_name, {
            'invoices': page,
            'status_filter': st,
            'currency_filter': cur,
            'tenant_filter': tenant_id,
            'search_query': q,
            'current_sort': sort_key,
            'current_dir': direction,
            'status_choices': StandardInvoice.STATUS_CHOICES,
            'currencies': Currency.objects.filter(is_active=True).order_by(
                'currency_code'),
            'tenants': TenantProfile.objects.order_by('company_name'),
        })


def _invoice_attachment_session_key(invoice_pk):
    return f'invoice_uploaded_attachment_{invoice_pk}'


def _is_pdf_upload(uploaded_attachment):
    if not uploaded_attachment:
        return False
    name = (uploaded_attachment.name or '').lower()
    content_type = (getattr(uploaded_attachment, 'content_type', '') or '').lower()
    return name.endswith('.pdf') and (
        content_type in ('application/pdf', 'application/x-pdf', '')
    )


def _store_invoice_uploaded_attachment(request, invoice_pk, uploaded_attachment):
    key = _invoice_attachment_session_key(invoice_pk)
    old_meta = request.session.get(key)
    if old_meta and old_meta.get('path') and default_storage.exists(old_meta['path']):
        default_storage.delete(old_meta['path'])

    file_ext = os.path.splitext(uploaded_attachment.name or '')[1] or '.bin'
    stored_path = default_storage.save(
        f'invoice_uploads/{invoice_pk}/{uuid.uuid4().hex}{file_ext}',
        uploaded_attachment,
    )
    request.session[key] = {
        'path': stored_path,
        'name': uploaded_attachment.name or os.path.basename(stored_path),
        'content_type': uploaded_attachment.content_type or 'application/octet-stream',
    }
    request.session.modified = True


class InvoiceDetailView(LoginRequiredMixin, View):
    template_name = 'crm/invoices/invoice_detail.html'

    def get(self, request, pk):
        invoice = get_object_or_404(
            StandardInvoice.objects.select_related(
                'tenant', 'currency', 'tax_code', 'order',
            ).prefetch_related('line_items'),
            invoice_id=pk,
        )
        base_cfg = BaseCurrencyConfig.objects.filter(
            setting_id='GLOBAL-BASE-CURRENCY',
        ).first()
        legal_identity = LegalIdentity.objects.filter(
            identity_id='GLOBAL-LEGAL-IDENTITY',
        ).first()
        attachment_meta = request.session.get(_invoice_attachment_session_key(pk))
        if attachment_meta and not default_storage.exists(attachment_meta.get('path', '')):
            request.session.pop(_invoice_attachment_session_key(pk), None)
            attachment_meta = None
        return render(
            request,
            self.template_name,
            {
                'invoice': invoice,
                'legal_identity': legal_identity,
                'has_uploaded_attachment': bool(attachment_meta),
                'uploaded_attachment_name': (
                    attachment_meta.get('name', '') if attachment_meta else ''
                ),
                'base_currency_code': (
                    base_cfg.base_currency_id if base_cfg and base_cfg.base_currency_id else None
                ),
            },
        )


class InvoicePrintView(LoginRequiredMixin, View):
    template_name = 'crm/invoices/invoice_print.html'

    def get(self, request, pk):
        invoice = get_object_or_404(
            StandardInvoice.objects.select_related(
                'tenant', 'currency', 'tax_code', 'order',
            ).prefetch_related('line_items'),
            invoice_id=pk,
        )
        legal_identity = LegalIdentity.objects.filter(
            identity_id='GLOBAL-LEGAL-IDENTITY',
        ).first()
        
        # Render HTML to string
        html_content = render_to_string(
            self.template_name,
            {
                'invoice': invoice,
                'legal_identity': legal_identity,
                'bill_to': get_live_bill_to_snapshot(invoice),
            },
        )
        
        # Convert relative media URLs to absolute file paths for wkhtmltopdf
        if settings.MEDIA_URL in html_content:
            media_root_path = str(settings.MEDIA_ROOT).replace('\\', '/')
            if not media_root_path.endswith('/'):
                media_root_path += '/'
            html_content = html_content.replace(
                settings.MEDIA_URL,
                media_root_path
            )
        
        # Generate PDF using wkhtmltopdf
        try:
            # We use '-' for stdin and stdout
            # --enable-local-file-access is added to allow wkhtmltopdf to access local images/css if paths are used
            process = subprocess.Popen(
                ['wkhtmltopdf', '--quiet', '--enable-local-file-access', '-', '-'],
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
            pdf_content, error = process.communicate(input=html_content.encode('utf-8'))
            
            if process.returncode != 0:
                logger.error(f"wkhtmltopdf error: {error.decode('utf-8')}")
                # Fallback to HTML if PDF fails
                response = HttpResponse(html_content)
                response['Content-Disposition'] = f'attachment; filename="{invoice.invoice_number}.html"'
                return response
                
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{invoice.invoice_number}.pdf"'
            return response
        except Exception as e:
            logger.exception("PDF generation failed, falling back to HTML")
            response = HttpResponse(html_content)
            response['Content-Disposition'] = f'attachment; filename="{invoice.invoice_number}.html"'
            return response


class InvoiceSendEmailView(LoginRequiredMixin, View):
    def post(self, request, pk):
        invoice = get_object_or_404(
            StandardInvoice.objects.select_related('tenant', 'currency'),
            invoice_id=pk,
        )
        uploaded_attachment = request.FILES.get('email_attachment')
        attachment_meta = request.session.get(_invoice_attachment_session_key(pk))
        attachment_payload = None

        if request.POST.get('upload_only') == '1':
            if not uploaded_attachment:
                messages.error(request, 'Please choose a file to upload.')
                return redirect('invoice_detail', pk=pk)
            if not _is_pdf_upload(uploaded_attachment):
                messages.error(request, 'Only PDF files are allowed.')
                return redirect('invoice_detail', pk=pk)
            _store_invoice_uploaded_attachment(request, pk, uploaded_attachment)
            messages.success(request, 'Attachment uploaded successfully.')
            return redirect('invoice_detail', pk=pk)

        if uploaded_attachment:
            if not _is_pdf_upload(uploaded_attachment):
                messages.error(request, 'Only PDF files are allowed.')
                return redirect('invoice_detail', pk=pk)
            _store_invoice_uploaded_attachment(request, pk, uploaded_attachment)
            attachment_payload = [(
                uploaded_attachment.name,
                uploaded_attachment.read(),
                uploaded_attachment.content_type or 'application/octet-stream',
            )]
        elif attachment_meta and default_storage.exists(attachment_meta.get('path', '')):
            with default_storage.open(attachment_meta['path'], 'rb') as stored_file:
                attachment_payload = [(
                    attachment_meta.get('name') or 'invoice-attachment',
                    stored_file.read(),
                    attachment_meta.get('content_type') or 'application/octet-stream',
                )]
        else:
            messages.error(
                request,
                'Please upload a file before sending invoice email.',
            )
            return redirect('invoice_detail', pk=pk)

        # sent = send_invoice_paid_notification(invoice, use_async_tasks=False)
        # sent = send_transactional_email(
        #     invoice.tenant.primary_email,
        #     f'Invoice {invoice.invoice_number} issued',
        #     (
        #         f'Hello,\n\n'
        #         f'Your invoice {invoice.invoice_number} is issued.\n'
        #         f'Amount: {invoice.grand_total} {invoice.currency_id}\n\n'
        #         f'Thank you.'
        #     ),
        #     (
        #         '<p>Hello,</p>'
        #         f'<p>Your invoice <strong>{invoice.invoice_number}</strong> is issued.</p>'
        #         f'<p><strong>Amount:</strong> {invoice.grand_total} {invoice.currency_id}</p>'
        #         '<p>Thank you.</p>'
        #     ),
        #     trigger_source='Manual: Invoice Send Email',
        #     client_id=str(invoice.tenant_id),
        #     attachments=[(
        #         uploaded_attachment.name,
        #         uploaded_attachment.read(),
        #         uploaded_attachment.content_type or 'application/octet-stream',
        #     )],
        # )
        email_context = {
            'invoice_number': invoice.invoice_number,
            'invoice_amount': str(invoice.grand_total),
            'invoice_amount_display': f'{invoice.grand_total} {invoice.currency_id}',
            'currency_code': invoice.currency_id,
            'company_name': invoice.customer_name,
            'tenant_name': invoice.customer_name,
            'issue_date': invoice.issue_date.strftime('%Y-%m-%d') if invoice.issue_date else '',
            'due_date': invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '',
            'invoice_status': invoice.status or 'Issued',
            'invoice_sub_total': str(invoice.sub_total),
            'invoice_discount_amount': str(invoice.discount_amount),
            'invoice_tax_amount': str(invoice.tax_amount),
            'invoice_grand_total': str(invoice.grand_total),
            'invoice_taxable_amount': str(invoice.taxable_amount),
            'customer_tax_number': invoice.customer_tax_number or '',
            'customer_address': invoice.customer_address or '',
            'line_items': list(
                invoice.line_items.values(
                    'item_description',
                    'quantity',
                    'unit_price',
                    'tax_rate',
                    'tax_amount',
                    'line_total',
                )
            ),
        }
        ensure_default_notification_templates()
        sent = send_named_notification_email(
            'INVOICE_PAID',
            recipient_email=invoice.tenant.primary_email,
            context_dict=email_context,
            language='en',
            default_subject=f'Invoice {invoice.invoice_number} issued',
            trigger_source=f'Manual Template: INVOICE_PAID for {invoice.invoice_number}',
            attachments=attachment_payload,
        )
        if not sent:
            sent = dispatch_event_notification(
                'Invoice_Paid',
                recipient_email=invoice.tenant.primary_email,
                context_dict=email_context,
                use_async_tasks=False,
                attachments=attachment_payload,
            )
        if sent:
            messages.success(
                request,
                f'Invoice email queued for {invoice.tenant.primary_email}.',
            )
        else:
            messages.error(
                request,
                'Invoice email could not be sent. Check gateway/template configuration.',
            )
        return redirect('invoice_detail', pk=pk)


class InvoiceUploadAttachmentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        get_object_or_404(StandardInvoice, invoice_id=pk)
        uploaded_attachment = request.FILES.get('email_attachment')
        if not uploaded_attachment:
            messages.error(request, 'Please choose a file to upload.')
            return redirect('invoice_detail', pk=pk)
        if not _is_pdf_upload(uploaded_attachment):
            messages.error(request, 'Only PDF files are allowed.')
            return redirect('invoice_detail', pk=pk)
        _store_invoice_uploaded_attachment(request, pk, uploaded_attachment)
        messages.success(request, 'Attachment uploaded successfully.')
        return redirect('invoice_detail', pk=pk)


class InvoiceVoidView(RootRequiredMixin, View):
    def post(self, request, pk):
        from superadmin.billing_helpers import generate_credit_note_from_invoice

        invoice = get_object_or_404(
            StandardInvoice.objects.prefetch_related('line_items'),
            invoice_id=pk,
        )
        if invoice.status != 'Issued':
            messages.error(request, 'Only issued invoices can be voided.')
            return redirect('invoice_detail', pk=pk)
        with db_transaction.atomic():
            inv_row = StandardInvoice.objects.select_for_update().get(pk=invoice.pk)
            inv_row.status = 'Void'
            inv_row.save(update_fields=['status', 'updated_at'])
            generate_credit_note_from_invoice(inv_row, request.user)
        messages.success(request, 'Invoice voided (credit note generated).')
        return redirect('invoice_detail', pk=pk)


class SupportCategoryListView(LoginRequiredMixin, View):
    template_name = 'support/categories/category_list.html'

    def get(self, request):
        sort_by = request.GET.get('sort', 'rank').strip()
        sort_dir = request.GET.get('dir', 'desc').strip().lower()
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'desc'

        qs = SupportCategory.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('name_en').asc(),
            )
        )
        q = request.GET.get('q', '').strip()
        status_filter = request.GET.get('is_active', 'All').strip()
        if q:
            qs = qs.filter(
                Q(name_en__icontains=q) |
                Q(name_ar__icontains=q)
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': ['default_rank'],
            'category_id': ['category_id'],
            'name_en': ['name_en'],
            'name_ar': ['name_ar'],
            'status': ['is_active'],
        }
        active_sort_fields = sort_mapping.get(sort_by, ['default_rank'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else f'-{f}')
        qs = qs.order_by(*ordering)

        total_count = qs.count()
        paginator = Paginator(qs, 10)
        categories = paginator.get_page(request.GET.get('page'))
        start_index = categories.start_index()
        for offset, category in enumerate(categories.object_list):
            # Show descending list ID so top rows have higher numbers.
            category.list_rank = total_count - (start_index + offset) + 1
        return render(
            request,
            self.template_name,
            {
                'categories': categories,
                'search_query': q,
                'status_filter': status_filter,
                'current_sort': sort_by,
                'current_dir': sort_dir,
            },
        )


class SupportCategoryCreateView(LoginRequiredMixin, View):
    template_name = 'support/categories/category_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {
                'form': SupportCategoryForm(),
                'is_edit': False,
                'category': None,
            },
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = SupportCategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.created_by = request.user
            category.save()
            messages.success(request, 'Support category created successfully.')
            return redirect('support_category_list')
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': False,
                'category': None,
            },
        )


class SupportCategoryUpdateView(LoginRequiredMixin, View):
    template_name = 'support/categories/category_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        category = get_object_or_404(SupportCategory, pk=pk)
        return render(
            request,
            self.template_name,
            {
                'form': SupportCategoryForm(instance=category),
                'is_edit': True,
                'category': category,
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        category = get_object_or_404(SupportCategory, pk=pk)
        form = SupportCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Support category updated successfully.')
            return redirect('support_category_list')
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': True,
                'category': category,
            },
        )


class SupportCategoryToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        category = get_object_or_404(SupportCategory, pk=pk)
        if category.is_active:
            # TODO: Check if category has open tickets before
            #        deactivating — warn but allow
            category.is_active = False
            status_text = 'deactivated'
        else:
            category.is_active = True
            status_text = 'activated'
        category.save(update_fields=['is_active', 'updated_at'])
        messages.success(
            request,
            f"Support category '{category.name_en}' {status_text}.",
        )
        return redirect('support_category_list')


class SupportCategoryDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        messages.error(
            request,
            'Categories cannot be deleted. Deactivate instead.',
        )
        return redirect('support_category_list')

    def get(self, request, pk):
        return self.post(request, pk)


class CannedResponseListView(LoginRequiredMixin, View):
    template_name = 'support/canned/canned_list.html'

    def get(self, request):
        sort_by = request.GET.get('sort', 'rank').strip()
        sort_dir = request.GET.get('dir', 'desc').strip().lower()
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'desc'

        qs = CannedResponse.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('title').asc(),
            )
        )
        q = request.GET.get('q', '').strip()
        status_filter = request.GET.get('is_active', 'All').strip()
        if q:
            qs = qs.filter(
                Q(title__icontains=q) |
                Q(message_body__icontains=q)
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': ['default_rank'],
            'template_id': ['template_id'],
            'title': ['title'],
            'message': ['message_body'],
            'status': ['is_active'],
        }
        active_sort_fields = sort_mapping.get(sort_by, ['default_rank'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else f'-{f}')
        qs = qs.order_by(*ordering)

        total_count = qs.count()
        paginator = Paginator(qs, 10)
        canned_responses = paginator.get_page(request.GET.get('page'))
        start_index = canned_responses.start_index()
        for offset, canned in enumerate(canned_responses.object_list):
            # Show descending list ID so top rows have higher numbers.
            canned.list_rank = total_count - (start_index + offset) + 1
        return render(
            request,
            self.template_name,
            {
                'canned_responses': canned_responses,
                'search_query': q,
                'status_filter': status_filter,
                'current_sort': sort_by,
                'current_dir': sort_dir,
            },
        )


class CannedResponseCreateView(LoginRequiredMixin, View):
    template_name = 'support/canned/canned_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {
                'form': CannedResponseForm(),
                'is_edit': False,
                'canned': None,
            },
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = CannedResponseForm(request.POST)
        if form.is_valid():
            canned = form.save(commit=False)
            canned.created_by = request.user
            canned.save()
            messages.success(request, 'Canned response created successfully.')
            return redirect('canned_response_list')
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': False,
                'canned': None,
            },
        )


class CannedResponseUpdateView(LoginRequiredMixin, View):
    template_name = 'support/canned/canned_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        canned = get_object_or_404(CannedResponse, pk=pk)
        return render(
            request,
            self.template_name,
            {
                'form': CannedResponseForm(instance=canned),
                'is_edit': True,
                'canned': canned,
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        canned = get_object_or_404(CannedResponse, pk=pk)
        form = CannedResponseForm(request.POST, instance=canned)
        if form.is_valid():
            form.save()
            messages.success(request, 'Canned response updated successfully.')
            return redirect('canned_response_list')
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': True,
                'canned': canned,
            },
        )


class CannedResponseToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        canned = get_object_or_404(CannedResponse, pk=pk)
        canned.is_active = not canned.is_active
        canned.save(update_fields=['is_active', 'updated_at'])
        messages.success(request, 'Canned response status updated successfully.')
        return redirect('canned_response_list')


class CannedResponseDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        messages.error(
            request,
            'Canned responses cannot be deleted. Deactivate instead.',
        )
        return redirect('canned_response_list')

    def get(self, request, pk):
        return self.post(request, pk)


class SubscriptionFAQListView(LoginRequiredMixin, View):
    template_name = 'support/faqs/faq_list.html'

    def get(self, request):
        sort_by = request.GET.get('sort', 'rank').strip()
        sort_dir = request.GET.get('dir', 'desc').strip().lower()
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'desc'

        qs = SubscriptionFAQ.objects.annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('display_order').asc(),
            )
        )
        q = request.GET.get('q', '').strip()
        status_filter = request.GET.get('is_active', 'All').strip()
        if q:
            qs = qs.filter(Q(question__icontains=q) | Q(answer__icontains=q))
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'rank': ['default_rank'],
            'question': ['question'],
            'order': ['display_order'],
            'status': ['is_active'],
        }
        active_sort_fields = sort_mapping.get(sort_by, ['default_rank'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else f'-{f}')
        qs = qs.order_by(*ordering)

        total_count = qs.count()
        paginator = Paginator(qs, 10)
        faqs = paginator.get_page(request.GET.get('page'))
        start_index = faqs.start_index()
        for offset, faq in enumerate(faqs.object_list):
            faq.list_rank = total_count - (start_index + offset) + 1
        return render(
            request,
            self.template_name,
            {
                'faqs': faqs,
                'search_query': q,
                'status_filter': status_filter,
                'current_sort': sort_by,
                'current_dir': sort_dir,
            },
        )


class SubscriptionFAQCreateView(LoginRequiredMixin, View):
    template_name = 'support/faqs/faq_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {
                'form': SubscriptionFAQForm(),
                'is_edit': False,
                'faq': None,
            },
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        form = SubscriptionFAQForm(request.POST)
        if form.is_valid():
            faq = form.save(commit=False)
            faq.created_by = request.user
            faq.updated_by = request.user
            faq.save()
            messages.success(request, 'FAQ created successfully.')
            return redirect('subscription_faq_list')
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': False,
                'faq': None,
            },
        )


class SubscriptionFAQUpdateView(LoginRequiredMixin, View):
    template_name = 'support/faqs/faq_form.html'

    def get(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        faq = get_object_or_404(SubscriptionFAQ, pk=pk)
        return render(
            request,
            self.template_name,
            {
                'form': SubscriptionFAQForm(instance=faq),
                'is_edit': True,
                'faq': faq,
            },
        )

    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        faq = get_object_or_404(SubscriptionFAQ, pk=pk)
        form = SubscriptionFAQForm(request.POST, instance=faq)
        if form.is_valid():
            updated = form.save(commit=False)
            updated.updated_by = request.user
            updated.save()
            messages.success(request, 'FAQ updated successfully.')
            return redirect('subscription_faq_list')
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'is_edit': True,
                'faq': faq,
            },
        )


class SubscriptionFAQToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        faq = get_object_or_404(SubscriptionFAQ, pk=pk)
        faq.is_active = not faq.is_active
        faq.updated_by = request.user
        faq.save(update_fields=['is_active', 'updated_by', 'updated_at'])
        messages.success(
            request,
            f"FAQ '{faq.question}' {'activated' if faq.is_active else 'deactivated'}.",
        )
        return redirect('subscription_faq_list')


class TicketListView(LoginRequiredMixin, View):
    template_name = 'support/tickets/ticket_list.html'

    def get(self, request):
        sort_by = request.GET.get('sort', 'rank').strip()
        sort_dir = request.GET.get('dir', 'desc').strip().lower()
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'desc'

        tickets_qs = SupportTicket.objects.select_related(
            'tenant',
            'category',
            'assigned_to',
        ).annotate(
            default_rank=Window(
                expression=RowNumber(),
                order_by=F('created_at').desc(),
            )
        )

        q = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', '').strip()
        priority_filter = request.GET.get('priority', '').strip()
        category_filter = request.GET.get('category', '').strip()
        assigned_filter = request.GET.get('assigned_to', '').strip()

        if q:
            tickets_qs = tickets_qs.filter(
                Q(ticket_no__icontains=q) |
                Q(subject__icontains=q) |
                Q(tenant__company_name__icontains=q)
            )
        if status_filter:
            tickets_qs = tickets_qs.filter(status=status_filter)
        if priority_filter:
            tickets_qs = tickets_qs.filter(priority=priority_filter)
        if category_filter:
            tickets_qs = tickets_qs.filter(category_id=category_filter)
        if assigned_filter == 'unassigned':
            tickets_qs = tickets_qs.filter(assigned_to__isnull=True)
        elif assigned_filter:
            tickets_qs = tickets_qs.filter(assigned_to_id=assigned_filter)

        sort_mapping = {
            'rank': ['default_rank'],
            'ticket_no': ['ticket_no'],
            'subject': ['subject'],
            'category': ['category__name_en'],
            'priority': ['priority'],
            'status': ['status'],
            'assignee': ['assigned_to__first_name', 'assigned_to__last_name'],
            'created': ['created_at'],
        }
        active_sort_fields = sort_mapping.get(sort_by, ['default_rank'])
        ordering = []
        for f in active_sort_fields:
            ordering.append(f if sort_dir == 'asc' else f'-{f}')
        tickets_qs = tickets_qs.order_by(*ordering)

        total_count = tickets_qs.count()
        paginator = Paginator(tickets_qs, 10)
        tickets = paginator.get_page(request.GET.get('page'))
        start_index = tickets.start_index()
        for offset, ticket in enumerate(tickets.object_list):
            # Show descending list ID so top rows have higher numbers.
            ticket.list_rank = total_count - (start_index + offset) + 1

        context = {
            'tickets': tickets,
            'search_query': q,
            'status_filter': status_filter,
            'priority_filter': priority_filter,
            'category_filter': category_filter,
            'assigned_filter': assigned_filter,
            'total_tickets_count': SupportTicket.objects.count(),
            'open_tickets_count': SupportTicket.objects.filter(status='New').count(),
            'in_progress_tickets_count': SupportTicket.objects.filter(status='In_Progress').count(),
            'closed_tickets_count': SupportTicket.objects.filter(status='Closed').count(),
            'categories': SupportCategory.objects.order_by('name_en'),
            'admins': AdminUser.objects.filter(status='Active').order_by(
                'first_name', 'last_name'
            ),
            'status_choices': SupportTicket.STATUS_CHOICES,
            'priority_choices': SupportTicket.PRIORITY_CHOICES,
            'current_sort': sort_by,
            'current_dir': sort_dir,
        }
        return render(request, self.template_name, context)


class TicketCreateView(LoginRequiredMixin, View):
    template_name = 'support/tickets/ticket_form.html'

    def get(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp
        return render(
            request,
            self.template_name,
            {'form': SupportTicketForm()},
        )

    def post(self, request):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        form = SupportTicketForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        ticket = form.save(commit=False)
        ticket.ticket_no = SupportTicket.generate_ticket_no()
        ticket.status = 'New'
        ticket.created_by = str(
            getattr(request.user, 'admin_id', getattr(request.user, 'id', ''))
        )
        ticket.save()

        TicketReply.objects.create(
            ticket=ticket,
            sender_type='System_Bot',
            sender_id='SYSTEM',
            message_body=(
                f"Ticket {ticket.ticket_no} has been "
                f"created. Our support team will review "
                f"your issue shortly."
            ),
            is_internal=False,
        )

        if ticket.assigned_to:
            ticket.status = 'In_Progress'
            ticket.save(update_fields=['status'])

        messages.success(request, 'Support ticket created successfully.')
        return redirect('ticket_detail', ticket_no=ticket.ticket_no)


def _resolve_support_ticket(identifier):
    """
    Resolve support ticket by ticket_no first, then fallback to UUID primary key
    for legacy links/forms that may still carry the old URL format.
    """
    ticket = SupportTicket.objects.filter(ticket_no=identifier).first()
    if ticket:
        return ticket
    return get_object_or_404(SupportTicket, pk=identifier)


class TicketDetailView(LoginRequiredMixin, View):
    template_name = 'support/tickets/ticket_detail.html'

    def get(self, request, ticket_no):
        ticket = _resolve_support_ticket(ticket_no)
        ticket = SupportTicket.objects.select_related(
            'tenant',
            'category',
            'assigned_to',
        ).get(pk=ticket.pk)
        replies = ticket.replies.select_related('ticket').all()

        context = {
            'ticket': ticket,
            'replies': replies,
            'reply_form': AdminReplyForm(),
            'assign_form': TicketAssignForm(instance=ticket),
            'priority_form': TicketPriorityForm(instance=ticket),
            'canned_responses': CannedResponse.objects.filter(
                is_active=True
            ).order_by('title'),
        }
        return render(request, self.template_name, context)


class TicketAdminReplyView(LoginRequiredMixin, View):
    def post(self, request, ticket_no):
        ticket = _resolve_support_ticket(ticket_no)
        if ticket.status == 'Closed':
            messages.error(
                request,
                'Cannot reply to a closed ticket. Reopen workflow is required.',
            )
            return redirect('ticket_detail', ticket_no=ticket.ticket_no)
        form = AdminReplyForm(request.POST, request.FILES)

        if not form.is_valid():
            messages.error(request, 'Please correct the reply form errors.')
            replies = ticket.replies.select_related('ticket').all()
            context = {
                'ticket': ticket,
                'replies': replies,
                'reply_form': form,
                'assign_form': TicketAssignForm(instance=ticket),
                'priority_form': TicketPriorityForm(instance=ticket),
                'canned_responses': CannedResponse.objects.filter(
                    is_active=True
                ).order_by('title'),
            }
            return render(request, 'support/tickets/ticket_detail.html', context)

        reply = form.save(commit=False)
        reply.ticket = ticket
        reply.sender_type = 'Admin_Support'
        reply.sender_id = str(
            getattr(request.user, 'admin_id', getattr(request.user, 'id', ''))
        )
        if (request.POST.get('submit_mode') or '').strip().lower() == 'internal':
            reply.is_internal = True
        reply.save()

        if not reply.is_internal:
            ticket.status = 'Pending_Client'
            ticket.save(update_fields=['status'])

        messages.success(request, 'Reply submitted successfully.')
        return redirect('ticket_detail', ticket_no=ticket.ticket_no)

    def get(self, request, ticket_no):
        return redirect('ticket_detail', ticket_no=ticket_no)


class TicketAssignView(LoginRequiredMixin, View):
    def post(self, request, ticket_no):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        ticket = _resolve_support_ticket(ticket_no)
        form = TicketAssignForm(request.POST, instance=ticket)
        if not form.is_valid():
            messages.error(request, 'Please select a valid assignee.')
            return redirect('ticket_detail', ticket_no=ticket.ticket_no)

        ticket = form.save(commit=False)
        if ticket.assigned_to and ticket.status != 'Closed':
            ticket.status = 'In_Progress'
        ticket.save()

        assignee_display = (
            f'{ticket.assigned_to.first_name} {ticket.assigned_to.last_name}'.strip()
            if ticket.assigned_to
            else 'Unassigned'
        )
        TicketReply.objects.create(
            ticket=ticket,
            sender_type='System_Bot',
            sender_id='SYSTEM',
            message_body=f'Ticket assigned to {assignee_display}.',
            is_internal=True,
        )
        messages.success(request, 'Ticket assignment updated.')
        return redirect('ticket_detail', ticket_no=ticket.ticket_no)

    def get(self, request, ticket_no):
        return redirect('ticket_detail', ticket_no=ticket_no)



class TicketPriorityOverrideView(LoginRequiredMixin, View):
    def post(self, request, ticket_no):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        ticket = _resolve_support_ticket(ticket_no)
        form = TicketPriorityForm(request.POST, instance=ticket)
        if not form.is_valid():
            messages.error(request, 'Please select a valid priority.')
            return redirect('ticket_detail', ticket_no=ticket.ticket_no)

        ticket = form.save()
        if ticket.priority in ['High', 'Critical']:
            try:
                dispatch_internal_alerts(
                    'High_Priority_Ticket',
                    context_dict={
                        'ticket_id': str(ticket.ticket_id),
                        'ticket_no': ticket.ticket_no,
                        'priority': ticket.priority,
                        'tenant': ticket.tenant.company_name,
                        'message': (
                            f'Ticket "{ticket.ticket_no}" priority changed '
                            f'to {ticket.priority}.'
                        ),
                    },
                )
            except Exception:
                logger.exception(
                    'Internal alert dispatch failed on ticket priority update %s',
                    ticket.ticket_id,
                )
        TicketReply.objects.create(
            ticket=ticket,
            sender_type='System_Bot',
            sender_id='SYSTEM',
            message_body=f'Priority changed to {ticket.priority} by admin.',
            is_internal=True,
        )
        messages.success(request, 'Ticket priority updated.')
        return redirect('ticket_detail', ticket_no=ticket.ticket_no)

    def get(self, request, ticket_no):
        return redirect('ticket_detail', ticket_no=ticket_no)


class TicketForceCloseView(LoginRequiredMixin, View):
    def post(self, request, ticket_no):
        redirect_resp = _require_root_or_redirect(request)
        if redirect_resp:
            return redirect_resp

        ticket = _resolve_support_ticket(ticket_no)
        ticket.status = 'Closed'
        ticket.closed_at = timezone.now()
        ticket.save(update_fields=['status', 'closed_at'])

        TicketReply.objects.create(
            ticket=ticket,
            sender_type='System_Bot',
            sender_id='SYSTEM',
            message_body=(
                'This ticket has been closed by the support team. '
                'If your issue persists, please open a new ticket.'
            ),
            is_internal=False,
        )
        messages.success(request, 'Ticket has been force closed.')
        return redirect('ticket_detail', ticket_no=ticket.ticket_no)

    def get(self, request, ticket_no):
        return redirect('ticket_detail', ticket_no=ticket_no)


class GlobalSearchView(LoginRequiredMixin, TemplateView):
    template_name = 'search/search_results.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query = self.request.GET.get('q', '').strip()

        results = {
            'users': [],
            'tenants': [],
            'tickets': [],
            'invoices': [],
            'roles': [],
            'plans': [],
            'promo_codes': [],
        }

        if query:
            # Users
            results['users'] = AdminUser.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(email__icontains=query) |
                Q(phone_number__icontains=query)
            ).distinct()[:10]

            # Tenants
            results['tenants'] = TenantProfile.objects.filter(
                Q(company_name__icontains=query) |
                Q(registration_number__icontains=query) |
                Q(tax_number__icontains=query) |
                Q(primary_email__icontains=query) |
                Q(primary_phone__icontains=query)
            ).distinct()[:10]

            # Tickets
            results['tickets'] = SupportTicket.objects.filter(
                Q(subject__icontains=query) |
                Q(ticket_no__icontains=query)
            ).distinct()[:10]

            # Invoices
            results['invoices'] = StandardInvoice.objects.filter(
                Q(invoice_number__icontains=query) |
                Q(customer_name__icontains=query) |
                Q(customer_tax_number__icontains=query)
            ).distinct()[:10]

            # Roles
            results['roles'] = Role.objects.filter(
                Q(role_name_en__icontains=query) |
                Q(role_name_ar__icontains=query)
            ).distinct()[:10]

            # Plans
            results['plans'] = SubscriptionPlan.objects.filter(
                Q(plan_name_en__icontains=query) |
                Q(plan_name_ar__icontains=query)
            ).distinct()[:10]

            # Promo Codes
            results['promo_codes'] = PromoCode.objects.filter(
                Q(code__icontains=query)
            ).distinct()[:10]

        context['query'] = query
        context['results'] = results
        total_count = 0
        for key, val in results.items():
            if hasattr(val, 'count'):
                total_count += val.count()
            else:
                total_count += len(val)
        
        context['total_count'] = total_count
        context['page_title'] = f"Search results for '{query}'"
        return context


# ── CMS: Home Page (iroad_frontend) ─────────────────────────────────

from iroad_frontend.models import (
    AboutApproachPillar,
    AboutFaqItem,
    AboutHowWorkStep,
    AboutPageContent,
    ContactPageContent,
    ContactSubmission,
    HomeMapLocation,
    HomePageContent,
    HomePricingBenefit,
    HomePricingTier,
    HomeServiceCard,
    HomeTestimonial,
    PricingFaqItem,
    PricingInteractiveStep,
    PricingPageContent,
    PrivacyPolicyPageContent,
    TermsConditionsPageContent,
)
from iroad_frontend.cms_forms import (
    AboutApproachPillarForm,
    AboutFaqItemForm,
    AboutHowWorkStepForm,
    AboutPageContentForm,
    ContactPageContentForm,
    HomeMapLocationForm,
    HomePageContentForm,
    HomePricingBenefitForm,
    HomePricingTierForm,
    HomeServiceCardForm,
    HomeTestimonialForm,
    PricingFaqItemForm,
    PricingInteractiveStepForm,
    PricingPageContentForm,
    PrivacyPolicyPageContentForm,
    TermsConditionsPageContentForm,
)


def _cms_next_child_order(queryset):
    last = queryset.order_by('-order').first()
    if last is None:
        return 0
    return (last.order or 0) + 1


class HomePageCMSView(LoginRequiredMixin, View):
    """
    Singleton edit view for Home Page main content.
    GET: show form pre-filled with current content
    POST: save and redirect back with success message
    """
    template_name = 'superadmin/cms/home_page_cms.html'

    def _get_home(self):
        return HomePageContent.get_singleton()

    def get(self, request):
        home = self._get_home()
        form = HomePageContentForm(instance=home)
        return render(request, self.template_name, {
            'form': form,
            'home': home,
            'page_title': 'Home Page CMS',
        })

    def post(self, request):
        home = self._get_home()
        form = HomePageContentForm(
            request.POST,
            request.FILES,
            instance=home,
        )
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = (
                f'{request.user.first_name} '
                f'{request.user.last_name}'
            )
            obj.save()
            messages.success(
                request,
                'Home page content updated successfully.',
            )
            return redirect('home_page_cms')
        return render(request, self.template_name, {
            'form': form,
            'home': home,
            'page_title': 'Home Page CMS',
        })


class HomeServiceCardListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/service_card_list.html'

    def get(self, request):
        home = HomePageContent.get_singleton()
        qs = home.service_cards.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(title_en__icontains=search_query)
                | Q(title_ar__icontains=search_query)
                | Q(summary_en__icontains=search_query)
                | Q(summary_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'title': 'title_en',
            'title_ar': 'title_ar',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        cards = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'cards': cards,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'Service Cards',
            },
        )


class HomeServiceCardCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/service_card_form.html'

    def get(self, request):
        form = HomeServiceCardForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Service Card',
        })

    def post(self, request):
        form = HomeServiceCardForm(request.POST, request.FILES)
        if form.is_valid():
            card = form.save(commit=False)
            card.home = HomePageContent.get_singleton()
            card.order = _cms_next_child_order(card.home.service_cards.all())
            card.save()
            messages.success(request, 'Service card created.')
            return redirect('home_service_card_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Service Card',
        })


class HomeServiceCardUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/service_card_form.html'

    def get(self, request, pk):
        card = get_object_or_404(HomeServiceCard, pk=pk)
        form = HomeServiceCardForm(instance=card)
        return render(request, self.template_name, {
            'form': form,
            'card': card,
            'page_title': 'Edit Service Card',
        })

    def post(self, request, pk):
        card = get_object_or_404(HomeServiceCard, pk=pk)
        form = HomeServiceCardForm(
            request.POST, request.FILES, instance=card)
        if form.is_valid():
            form.save()
            messages.success(request, 'Service card updated.')
            return redirect('home_service_card_list')
        return render(request, self.template_name, {
            'form': form,
            'card': card,
            'page_title': 'Edit Service Card',
        })


class HomePricingTierListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_tier_list.html'

    def get(self, request):
        home = HomePageContent.get_singleton()
        qs = home.pricing_tiers.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(name_en__icontains=search_query)
                | Q(name_ar__icontains=search_query)
                | Q(summary_en__icontains=search_query)
                | Q(summary_ar__icontains=search_query)
                | Q(price_display_en__icontains=search_query)
                | Q(price_display_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'name': 'name_en',
            'name_ar': 'name_ar',
            'price': 'price_display_en',
            'featured': 'is_featured',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        tiers = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'tiers': tiers,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'Pricing Tiers',
            },
        )


class HomePricingTierCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_tier_form.html'

    def get(self, request):
        form = HomePricingTierForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Pricing Tier',
        })

    def post(self, request):
        form = HomePricingTierForm(request.POST, request.FILES)
        if form.is_valid():
            tier = form.save(commit=False)
            tier.home = HomePageContent.get_singleton()
            tier.order = _cms_next_child_order(tier.home.pricing_tiers.all())
            tier.save()
            messages.success(request, 'Pricing tier created.')
            return redirect('home_pricing_tier_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Pricing Tier',
        })


class HomePricingTierUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_tier_form.html'

    def get(self, request, pk):
        tier = get_object_or_404(HomePricingTier, pk=pk)
        form = HomePricingTierForm(instance=tier)
        return render(request, self.template_name, {
            'form': form,
            'tier': tier,
            'page_title': 'Edit Pricing Tier',
        })

    def post(self, request, pk):
        tier = get_object_or_404(HomePricingTier, pk=pk)
        form = HomePricingTierForm(
            request.POST, request.FILES, instance=tier)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pricing tier updated.')
            return redirect('home_pricing_tier_list')
        return render(request, self.template_name, {
            'form': form,
            'tier': tier,
            'page_title': 'Edit Pricing Tier',
        })


class HomePricingBenefitListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/home_pricing_benefit_list.html'

    def get(self, request):
        home = HomePageContent.get_singleton()
        qs = home.pricing_benefits.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(text_en__icontains=search_query)
                | Q(text_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'text': 'text_en',
            'text_ar': 'text_ar',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        benefits = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'benefits': benefits,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'Pricing Benefits',
            },
        )


class HomePricingBenefitCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/home_pricing_benefit_form.html'

    def get(self, request):
        form = HomePricingBenefitForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Pricing Benefit',
        })

    def post(self, request):
        form = HomePricingBenefitForm(request.POST, request.FILES)
        if form.is_valid():
            row = form.save(commit=False)
            row.home = HomePageContent.get_singleton()
            row.order = _cms_next_child_order(row.home.pricing_benefits.all())
            row.save()
            messages.success(request, 'Pricing benefit created.')
            return redirect('home_pricing_benefit_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Pricing Benefit',
        })


class HomePricingBenefitUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/home_pricing_benefit_form.html'

    def get(self, request, pk):
        benefit = get_object_or_404(HomePricingBenefit, pk=pk)
        form = HomePricingBenefitForm(instance=benefit)
        return render(request, self.template_name, {
            'form': form,
            'benefit': benefit,
            'page_title': 'Edit Pricing Benefit',
        })

    def post(self, request, pk):
        benefit = get_object_or_404(HomePricingBenefit, pk=pk)
        form = HomePricingBenefitForm(
            request.POST, request.FILES, instance=benefit)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pricing benefit updated.')
            return redirect('home_pricing_benefit_list')
        return render(request, self.template_name, {
            'form': form,
            'benefit': benefit,
            'page_title': 'Edit Pricing Benefit',
        })


class HomePricingBenefitDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        benefit = get_object_or_404(HomePricingBenefit, pk=pk)
        benefit.delete()
        messages.success(request, 'Pricing benefit deleted.')
        return redirect('home_pricing_benefit_list')


class HomeTestimonialListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/testimonial_list.html'

    def get(self, request):
        home = HomePageContent.get_singleton()
        qs = home.testimonials.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(author_name_en__icontains=search_query)
                | Q(author_name_ar__icontains=search_query)
                | Q(author_role_en__icontains=search_query)
                | Q(author_role_ar__icontains=search_query)
                | Q(quote_en__icontains=search_query)
                | Q(quote_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'author': 'author_name_en',
            'author_ar': 'author_name_ar',
            'role': 'author_role_en',
            'role_ar': 'author_role_ar',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        testimonials = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'testimonials': testimonials,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'Testimonials',
            },
        )


class HomeTestimonialCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/testimonial_form.html'

    def get(self, request):
        form = HomeTestimonialForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Testimonial',
        })

    def post(self, request):
        form = HomeTestimonialForm(request.POST, request.FILES)
        if form.is_valid():
            testimonial = form.save(commit=False)
            testimonial.home = HomePageContent.get_singleton()
            testimonial.order = _cms_next_child_order(
                testimonial.home.testimonials.all())
            testimonial.save()
            messages.success(request, 'Testimonial created.')
            return redirect('home_testimonial_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Testimonial',
        })


class HomeTestimonialUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/testimonial_form.html'

    def get(self, request, pk):
        testimonial = get_object_or_404(HomeTestimonial, pk=pk)
        form = HomeTestimonialForm(instance=testimonial)
        return render(request, self.template_name, {
            'form': form,
            'testimonial': testimonial,
            'page_title': 'Edit Testimonial',
        })

    def post(self, request, pk):
        testimonial = get_object_or_404(HomeTestimonial, pk=pk)
        form = HomeTestimonialForm(
            request.POST, request.FILES, instance=testimonial)
        if form.is_valid():
            form.save()
            messages.success(request, 'Testimonial updated.')
            return redirect('home_testimonial_list')
        return render(request, self.template_name, {
            'form': form,
            'testimonial': testimonial,
            'page_title': 'Edit Testimonial',
        })


class HomeMapLocationListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/map_location_list.html'

    def get(self, request):
        home = HomePageContent.get_singleton()
        qs = home.map_locations.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(title_en__icontains=search_query)
                | Q(title_ar__icontains=search_query)
                | Q(subtitle_en__icontains=search_query)
                | Q(subtitle_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'title': 'title_en',
            'title_ar': 'title_ar',
            'subtitle': 'subtitle_en',
            'subtitle_ar': 'subtitle_ar',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        locations = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'locations': locations,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'Map Locations',
            },
        )


class HomeMapLocationCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/map_location_form.html'

    def get(self, request):
        form = HomeMapLocationForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Map Location',
        })

    def post(self, request):
        form = HomeMapLocationForm(request.POST, request.FILES)
        if form.is_valid():
            loc = form.save(commit=False)
            loc.home = HomePageContent.get_singleton()
            loc.order = _cms_next_child_order(loc.home.map_locations.all())
            loc.save()
            messages.success(request, 'Map location created.')
            return redirect('home_map_location_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Map Location',
        })


class HomeMapLocationUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/map_location_form.html'

    def get(self, request, pk):
        loc = get_object_or_404(HomeMapLocation, pk=pk)
        form = HomeMapLocationForm(instance=loc)
        return render(request, self.template_name, {
            'form': form,
            'location': loc,
            'page_title': 'Edit Map Location',
        })

    def post(self, request, pk):
        loc = get_object_or_404(HomeMapLocation, pk=pk)
        form = HomeMapLocationForm(
            request.POST, request.FILES, instance=loc)
        if form.is_valid():
            form.save()
            messages.success(request, 'Map location updated.')
            return redirect('home_map_location_list')
        return render(request, self.template_name, {
            'form': form,
            'location': loc,
            'page_title': 'Edit Map Location',
        })


# ── CMS: About Page (iroad_frontend) ───────────────────────────────


class AboutPageCMSView(LoginRequiredMixin, View):
    """
    Singleton edit view for About Page main content.
    """
    template_name = 'superadmin/cms/about_page_cms.html'

    def _get_about(self):
        return AboutPageContent.get_singleton()

    def get(self, request):
        about = self._get_about()
        form = AboutPageContentForm(instance=about)
        return render(request, self.template_name, {
            'form': form,
            'about': about,
            'page_title': 'About Page CMS',
        })

    def post(self, request):
        about = self._get_about()
        form = AboutPageContentForm(
            request.POST,
            request.FILES,
            instance=about,
        )
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = (
                f'{request.user.first_name} '
                f'{request.user.last_name}'
            )
            obj.save()
            messages.success(
                request,
                'About page content updated successfully.',
            )
            return redirect('about_page_cms')
        return render(request, self.template_name, {
            'form': form,
            'about': about,
            'page_title': 'About Page CMS',
        })


class AboutApproachPillarListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_pillar_list.html'

    def get(self, request):
        about = AboutPageContent.get_singleton()
        qs = about.approach_pillars.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(title_en__icontains=search_query)
                | Q(title_ar__icontains=search_query)
                | Q(body_en__icontains=search_query)
                | Q(body_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'title': 'title_en',
            'title_ar': 'title_ar',
            'body': 'body_en',
            'body_ar': 'body_ar',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        pillars = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'pillars': pillars,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'Approach Pillars',
            },
        )


class AboutApproachPillarCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_pillar_form.html'

    def get(self, request):
        form = AboutApproachPillarForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Approach Pillar',
        })

    def post(self, request):
        form = AboutApproachPillarForm(request.POST, request.FILES)
        if form.is_valid():
            pillar = form.save(commit=False)
            pillar.about = AboutPageContent.get_singleton()
            pillar.order = _cms_next_child_order(
                pillar.about.approach_pillars.all())
            pillar.save()
            messages.success(request, 'Approach pillar created.')
            return redirect('about_pillar_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Approach Pillar',
        })


class AboutApproachPillarUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_pillar_form.html'

    def get(self, request, pk):
        pillar = get_object_or_404(AboutApproachPillar, pk=pk)
        form = AboutApproachPillarForm(instance=pillar)
        return render(request, self.template_name, {
            'form': form,
            'pillar': pillar,
            'page_title': 'Edit Approach Pillar',
        })

    def post(self, request, pk):
        pillar = get_object_or_404(AboutApproachPillar, pk=pk)
        form = AboutApproachPillarForm(
            request.POST, request.FILES, instance=pillar)
        if form.is_valid():
            form.save()
            messages.success(request, 'Approach pillar updated.')
            return redirect('about_pillar_list')
        return render(request, self.template_name, {
            'form': form,
            'pillar': pillar,
            'page_title': 'Edit Approach Pillar',
        })


class AboutHowWorkStepListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_how_step_list.html'

    def get(self, request):
        about = AboutPageContent.get_singleton()
        qs = about.how_work_steps.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(step_number__icontains=search_query)
                | Q(title_en__icontains=search_query)
                | Q(title_ar__icontains=search_query)
                | Q(body_en__icontains=search_query)
                | Q(body_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'step': 'step_number',
            'title': 'title_en',
            'title_ar': 'title_ar',
            'body': 'body_en',
            'body_ar': 'body_ar',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        steps = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'steps': steps,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'How It Works Steps',
            },
        )


class AboutHowWorkStepCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_how_step_form.html'

    def get(self, request):
        form = AboutHowWorkStepForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add How It Works Step',
        })

    def post(self, request):
        form = AboutHowWorkStepForm(request.POST, request.FILES)
        if form.is_valid():
            step = form.save(commit=False)
            step.about = AboutPageContent.get_singleton()
            step.order = _cms_next_child_order(
                step.about.how_work_steps.all())
            step.save()
            messages.success(request, 'How it works step created.')
            return redirect('about_how_step_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add How It Works Step',
        })


class AboutHowWorkStepUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_how_step_form.html'

    def get(self, request, pk):
        step = get_object_or_404(AboutHowWorkStep, pk=pk)
        form = AboutHowWorkStepForm(instance=step)
        return render(request, self.template_name, {
            'form': form,
            'step': step,
            'page_title': 'Edit How It Works Step',
        })

    def post(self, request, pk):
        step = get_object_or_404(AboutHowWorkStep, pk=pk)
        form = AboutHowWorkStepForm(
            request.POST, request.FILES, instance=step)
        if form.is_valid():
            form.save()
            messages.success(request, 'How it works step updated.')
            return redirect('about_how_step_list')
        return render(request, self.template_name, {
            'form': form,
            'step': step,
            'page_title': 'Edit How It Works Step',
        })


class AboutFaqItemListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_faq_list.html'

    def get(self, request):
        about = AboutPageContent.get_singleton()
        qs = about.faq_items.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(question_en__icontains=search_query)
                | Q(question_ar__icontains=search_query)
                | Q(answer_en__icontains=search_query)
                | Q(answer_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'question': 'question_en',
            'question_ar': 'question_ar',
            'answer': 'answer_en',
            'answer_ar': 'answer_ar',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        faqs = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'faqs': faqs,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'About FAQ Items',
            },
        )


class AboutFaqItemCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_faq_form.html'

    def get(self, request):
        form = AboutFaqItemForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add About FAQ',
        })

    def post(self, request):
        form = AboutFaqItemForm(request.POST, request.FILES)
        if form.is_valid():
            faq = form.save(commit=False)
            faq.about = AboutPageContent.get_singleton()
            faq.order = _cms_next_child_order(faq.about.faq_items.all())
            faq.save()
            messages.success(request, 'About FAQ item created.')
            return redirect('about_faq_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add About FAQ',
        })


class AboutFaqItemUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/about_faq_form.html'

    def get(self, request, pk):
        faq = get_object_or_404(AboutFaqItem, pk=pk)
        form = AboutFaqItemForm(instance=faq)
        return render(request, self.template_name, {
            'form': form,
            'faq': faq,
            'page_title': 'Edit About FAQ',
        })

    def post(self, request, pk):
        faq = get_object_or_404(AboutFaqItem, pk=pk)
        form = AboutFaqItemForm(request.POST, request.FILES, instance=faq)
        if form.is_valid():
            form.save()
            messages.success(request, 'About FAQ item updated.')
            return redirect('about_faq_list')
        return render(request, self.template_name, {
            'form': form,
            'faq': faq,
            'page_title': 'Edit About FAQ',
        })


class PricingFaqItemListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_faq_list.html'

    def get(self, request):
        pricing = PricingPageContent.get_singleton()
        qs = pricing.faq_items.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(question_en__icontains=search_query)
                | Q(question_ar__icontains=search_query)
                | Q(answer_en__icontains=search_query)
                | Q(answer_ar__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'question': 'question_en',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        items = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'items': items,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'Pricing FAQ Items',
            },
        )


class PricingFaqItemCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_faq_form.html'

    def get(self, request):
        form = PricingFaqItemForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Pricing FAQ',
        })

    def post(self, request):
        form = PricingFaqItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            pricing = PricingPageContent.get_singleton()
            item.pricing = pricing
            item.order = _cms_next_child_order(pricing.faq_items.all())
            item.save()
            messages.success(request, 'Pricing FAQ item created.')
            return redirect('pricing_faq_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Pricing FAQ',
        })


class PricingFaqItemUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_faq_form.html'

    def get(self, request, pk):
        item = get_object_or_404(PricingFaqItem, pk=pk)
        form = PricingFaqItemForm(instance=item)
        return render(request, self.template_name, {
            'form': form,
            'item': item,
            'page_title': 'Edit Pricing FAQ',
        })

    def post(self, request, pk):
        item = get_object_or_404(PricingFaqItem, pk=pk)
        form = PricingFaqItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pricing FAQ item updated.')
            return redirect('pricing_faq_list')
        return render(request, self.template_name, {
            'form': form,
            'item': item,
            'page_title': 'Edit Pricing FAQ',
        })


# ── CMS: Pricing Page (iroad_frontend) ─────────────────────────────


class PricingPageCMSView(LoginRequiredMixin, View):
    """
    Singleton edit view for Pricing Page main content.
    """
    template_name = 'superadmin/cms/pricing_page_cms.html'

    def _get_pricing(self):
        return PricingPageContent.get_singleton()

    def get(self, request):
        pricing = self._get_pricing()
        form = PricingPageContentForm(instance=pricing)
        return render(request, self.template_name, {
            'form': form,
            'pricing': pricing,
            'page_title': 'Pricing Page CMS',
        })

    def post(self, request):
        pricing = self._get_pricing()
        form = PricingPageContentForm(
            request.POST,
            request.FILES,
            instance=pricing,
        )
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = (
                f'{request.user.first_name} '
                f'{request.user.last_name}'
            )
            obj.save()
            messages.success(
                request,
                'Pricing page content updated successfully.',
            )
            return redirect('pricing_page_cms')
        return render(request, self.template_name, {
            'form': form,
            'pricing': pricing,
            'page_title': 'Pricing Page CMS',
        })


class PricingInteractiveStepListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_interactive_list.html'

    def get(self, request):
        pricing = PricingPageContent.get_singleton()
        qs = pricing.interactive_steps.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'order')
        direction = request.GET.get('dir', 'asc')

        if search_query:
            qs = qs.filter(
                Q(title_en__icontains=search_query)
                | Q(title_ar__icontains=search_query)
                | Q(subtitle_en__icontains=search_query)
                | Q(subtitle_ar__icontains=search_query)
                | Q(body_en__icontains=search_query)
                | Q(body_ar__icontains=search_query)
                | Q(detail_url__icontains=search_query),
            )
        if status_filter == 'Active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'Inactive':
            qs = qs.filter(is_active=False)

        sort_mapping = {
            'order': 'order',
            'title': 'title_en',
            'title_ar': 'title_ar',
            'subtitle': 'subtitle_en',
            'subtitle_ar': 'subtitle_ar',
            'status': 'is_active',
        }
        order_field = sort_mapping.get(sort, 'order')
        prefix = '' if direction == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        steps = paginator.get_page(request.GET.get('page', 1))
        return render(
            request,
            self.template_name,
            {
                'steps': steps,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'page_title': 'Pricing Interactive Steps',
            },
        )


class PricingInteractiveStepCreateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_interactive_form.html'

    def get(self, request):
        form = PricingInteractiveStepForm()
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Interactive Step',
        })

    def post(self, request):
        form = PricingInteractiveStepForm(request.POST, request.FILES)
        if form.is_valid():
            step = form.save(commit=False)
            step.pricing = PricingPageContent.get_singleton()
            step.order = _cms_next_child_order(
                step.pricing.interactive_steps.all())
            step.save()
            messages.success(request, 'Interactive step created.')
            return redirect('pricing_interactive_list')
        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Add Interactive Step',
        })


class PricingInteractiveStepUpdateView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/pricing_interactive_form.html'

    def get(self, request, pk):
        step = get_object_or_404(PricingInteractiveStep, pk=pk)
        form = PricingInteractiveStepForm(instance=step)
        return render(request, self.template_name, {
            'form': form,
            'step': step,
            'page_title': 'Edit Interactive Step',
        })

    def post(self, request, pk):
        step = get_object_or_404(PricingInteractiveStep, pk=pk)
        form = PricingInteractiveStepForm(
            request.POST, request.FILES, instance=step)
        if form.is_valid():
            form.save()
            messages.success(request, 'Interactive step updated.')
            return redirect('pricing_interactive_list')
        return render(request, self.template_name, {
            'form': form,
            'step': step,
            'page_title': 'Edit Interactive Step',
        })


# ── CMS: Contact Page (iroad_frontend) ────────────────────────────


class ContactPageCMSView(LoginRequiredMixin, View):
    """Singleton edit view for Contact Page CMS."""

    template_name = 'superadmin/cms/contact_page_cms.html'

    def _get_contact(self):
        return ContactPageContent.get_singleton()

    def get(self, request):
        contact = self._get_contact()
        form = ContactPageContentForm(instance=contact)
        return render(request, self.template_name, {
            'form': form,
            'contact': contact,
            'page_title': 'Contact Page CMS',
        })

    def post(self, request):
        contact = self._get_contact()
        form = ContactPageContentForm(
            request.POST,
            request.FILES,
            instance=contact,
        )
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = (
                f'{request.user.first_name} '
                f'{request.user.last_name}'
            )
            obj.save()
            messages.success(
                request,
                'Contact page content updated successfully.',
            )
            return redirect('contact_page_cms')
        return render(request, self.template_name, {
            'form': form,
            'contact': contact,
            'page_title': 'Contact Page CMS',
        })


class PrivacyPolicyCMSView(LoginRequiredMixin, View):
    """Singleton edit view for Privacy Policy (Website CMS)."""

    template_name = 'superadmin/cms/privacy_policy_cms.html'

    def _get_page(self):
        return PrivacyPolicyPageContent.get_singleton()

    def get(self, request):
        page = self._get_page()
        form = PrivacyPolicyPageContentForm(instance=page)
        return render(request, self.template_name, {
            'form': form,
            'page': page,
            'page_title': 'Privacy Policy CMS',
        })

    def post(self, request):
        page = self._get_page()
        form = PrivacyPolicyPageContentForm(
            request.POST,
            request.FILES,
            instance=page,
        )
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = (
                f'{request.user.first_name} '
                f'{request.user.last_name}'
            )
            obj.save()
            messages.success(
                request,
                'Privacy policy content updated successfully.',
            )
            return redirect('privacy_policy_cms')
        return render(request, self.template_name, {
            'form': form,
            'page': page,
            'page_title': 'Privacy Policy CMS',
        })


class TermsConditionsCMSView(LoginRequiredMixin, View):
    """Singleton edit view for Terms & Conditions (Website CMS)."""

    template_name = 'superadmin/cms/terms_conditions_cms.html'

    def _get_page(self):
        return TermsConditionsPageContent.get_singleton()

    def get(self, request):
        page = self._get_page()
        form = TermsConditionsPageContentForm(instance=page)
        return render(request, self.template_name, {
            'form': form,
            'page': page,
            'page_title': 'Terms & Conditions CMS',
        })

    def post(self, request):
        page = self._get_page()
        form = TermsConditionsPageContentForm(
            request.POST,
            request.FILES,
            instance=page,
        )
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = (
                f'{request.user.first_name} '
                f'{request.user.last_name}'
            )
            obj.save()
            messages.success(
                request,
                'Terms & conditions content updated successfully.',
            )
            return redirect('terms_conditions_cms')
        return render(request, self.template_name, {
            'form': form,
            'page': page,
            'page_title': 'Terms & Conditions CMS',
        })


class ContactSubmissionListView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/contact_submission_list.html'

    def _list_redirect(self, request):
        """Rebuild list URL from POST hidden fields (preserve filters / page)."""
        params = {}
        q = (request.POST.get('q') or '').strip()
        if q:
            params['q'] = q
        status = (request.POST.get('status') or 'All').strip()
        if status not in ('', 'All'):
            params['status'] = status
        sort = (request.POST.get('sort') or 'submitted_at').strip()
        dir_ = (request.POST.get('dir') or 'desc').strip()
        if sort:
            params['sort'] = sort
        if dir_:
            params['dir'] = dir_
        page = request.POST.get('page')
        if page and str(page).isdigit() and int(page) > 1:
            params['page'] = page
        base = reverse('contact_submission_list')
        if params:
            return redirect(f'{base}?{urlencode(params)}')
        return redirect(base)

    def get(self, request):
        qs = ContactSubmission.objects.all()
        search_query = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', 'All')
        sort = request.GET.get('sort', 'submitted_at')
        direction = request.GET.get('dir', 'desc')

        if search_query:
            qs = qs.filter(
                Q(email__icontains=search_query)
                | Q(first_name__icontains=search_query)
                | Q(last_name__icontains=search_query)
                | Q(phone__icontains=search_query)
                | Q(message__icontains=search_query),
            )
        if status_filter == 'Unread':
            qs = qs.filter(is_read=False)
        elif status_filter == 'Read':
            qs = qs.filter(is_read=True)

        sort_mapping = {
            'id': 'pk',
            'submitted_at': 'submitted_at',
            'email': 'email',
            'name': 'last_name',
            'phone': 'phone',
            'read': 'is_read',
        }
        order_field = sort_mapping.get(sort, 'submitted_at')
        prefix = '' if direction == 'asc' else '-'
        if order_field == 'last_name':
            qs = qs.order_by(
                f'{prefix}last_name',
                f'{prefix}first_name',
                f'{prefix}pk',
            )
        else:
            qs = qs.order_by(f'{prefix}{order_field}', f'{prefix}pk')

        paginator = Paginator(qs, 10)
        submissions = paginator.get_page(request.GET.get('page', 1))
        unread_count = ContactSubmission.objects.filter(is_read=False).count()
        return render(
            request,
            self.template_name,
            {
                'submissions': submissions,
                'search_query': search_query,
                'status_filter': status_filter,
                'current_sort': sort,
                'current_dir': direction,
                'unread_count': unread_count,
                'page_title': 'Contact Submissions',
            },
        )

    def post(self, request):
        if request.POST.get('action') == 'mark_all_read':
            updated = ContactSubmission.objects.filter(is_read=False).update(
                is_read=True,
            )
            messages.success(
                request,
                f'Marked {updated} submission(s) as read.',
            )
        elif request.POST.get('action') == 'mark_read':
            raw_pk = request.POST.get('submission_id')
            if raw_pk:
                try:
                    pk = int(raw_pk)
                except (TypeError, ValueError):
                    pk = None
                if pk is not None:
                    sub = get_object_or_404(ContactSubmission, pk=pk)
                    sub.is_read = not sub.is_read
                    sub.save(update_fields=['is_read'])
                    state = 'read' if sub.is_read else 'unread'
                    messages.success(
                        request,
                        f'Submission marked as {state}.',
                    )
        return self._list_redirect(request)


class ContactSubmissionDetailView(LoginRequiredMixin, View):
    template_name = 'superadmin/cms/contact_submission_detail.html'

    def get(self, request, pk):
        submission = get_object_or_404(ContactSubmission, pk=pk)
        if not submission.is_read:
            submission.is_read = True
            submission.save(update_fields=['is_read'])
        return render(request, self.template_name, {
            'submission': submission,
            'page_title': f'Contact — {submission.email}',
        })

    def post(self, request, pk):
        submission = get_object_or_404(ContactSubmission, pk=pk)
        if request.POST.get('action') == 'toggle_read':
            submission.is_read = not submission.is_read
            submission.save(update_fields=['is_read'])
            state = 'read' if submission.is_read else 'unread'
            messages.success(request, f'Submission marked as {state}.')
        return redirect('contact_submission_detail', pk=pk)
